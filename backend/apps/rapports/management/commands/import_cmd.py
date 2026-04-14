"""
Management command : import complet des données ASELBY 2026
depuis les fichiers Excel.

Usage:
    python manage.py import_aselby_2026 --fichiers /chemin/vers/dossier/excel
    python manage.py import_aselby_2026 --fichiers /chemin/vers/dossier/excel --dry-run

Logique des comptes utilisateurs :
    - Admins (bureau) : username=tontine_2026  mdp=tontine_2026  (1 compte partagé bureau)
    - Membres : username=premier_nom  mdp=aselby_2026  (ou leur tel si dispo)
    - Réinitialisation mdp possible via la fonctionnalité "mot de passe oublié"
"""

import os
import unicodedata
from decimal import Decimal
from datetime import date

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction


def normaliser_username(nom):
    """
    Extrait le premier mot du nom, en minuscules, sans accents.
    Ex: 'BAKOP ACHILLE' -> 'bakop'
        'DJOUKWE NANGUEN GUY BLAISE' -> 'djoukwe'
    """
    if not nom:
        return ''
    premier_mot = str(nom).strip().split()[0].lower()
    # Supprimer les accents
    nfkd = unicodedata.normalize('NFKD', premier_mot)
    return ''.join(c for c in nfkd if not unicodedata.combining(c))


def nettoyer_telephone(tel):
    """Nettoie et formate un numéro de téléphone camerounais."""
    if not tel:
        return ''
    tel_str = str(tel).replace('.0', '').replace(' ', '').strip()
    # Garder seulement les chiffres
    chiffres = ''.join(c for c in tel_str if c.isdigit())
    return chiffres if len(chiffres) >= 9 else ''


class Command(BaseCommand):
    help = "Import complet des données ASELBY 2026 depuis les fichiers Excel"

    def add_arguments(self, parser):
        parser.add_argument(
            '--fichiers',
            type=str,
            required=True,
            help='Chemin vers le dossier contenant les fichiers Excel ASELBY 2026'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Simuler sans écrire en base'
        )
        parser.add_argument(
            '--reset',
            action='store_true',
            help='Supprimer les données existantes avant import'
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            raise CommandError("openpyxl requis : pip install openpyxl")

        dossier = options['fichiers']
        dry_run = options['dry_run']
        reset   = options['reset']

        if not os.path.isdir(dossier):
            raise CommandError(f"Dossier introuvable : {dossier}")

        self.stdout.write(self.style.WARNING(
            f"\n{'[DRY-RUN] ' if dry_run else ''}Import ASELBY 2026 depuis : {dossier}\n"
        ))

        # Chemins des fichiers Excel
        fichiers = {
            'adherents'  : os.path.join(dossier, 'ASELBY2026LISTEADHERENT.xlsx'),
            'fonds'      : os.path.join(dossier, 'ASELBY2026LISTEFONDSCAISSE.xlsx'),
            'tabbord'    : os.path.join(dossier, 'ASELBY2026TABBORD.xlsx'),
            't35'        : os.path.join(dossier, 'ASELBY2026TONTINE60000.xlsx'),
            't75'        : os.path.join(dossier, 'ASELBY2026TONTINE75000.xlsx'),
            't100'       : os.path.join(dossier, 'ASELBY2026TONTINE100000.xlsx'),
            'basecalcul' : os.path.join(dossier, 'ASELBY2026BASECALCULINTERET.xlsx'),
        }

        for nom, chemin in fichiers.items():
            if not os.path.exists(chemin):
                raise CommandError(f"Fichier manquant : {chemin}")

        with transaction.atomic():
            if dry_run:
                # On utilisera un savepoint qu'on annulera à la fin
                pass

            if reset and not dry_run:
                self._reset_donnees()

            # 1. Configuration exercice 2026
            config = self._import_config_exercice(dry_run)

            # 2. Adhérents
            adherents_map = self._import_adherents(fichiers['adherents'], dry_run)

            # 3. Comptes utilisateurs
            self._creer_comptes_utilisateurs(adherents_map, dry_run)

            # 4. Niveaux tontine
            niveaux_map = self._creer_niveaux_tontine(config, dry_run)

            # 5. Fonds de caisse initiaux
            self._import_fonds_initiaux(fichiers['fonds'], adherents_map, config, dry_run)

            # 6. Parts tontine (janvier 2026)
            self._import_parts_tontine(
                fichiers['tabbord'], fichiers['t35'], fichiers['t75'], fichiers['t100'],
                adherents_map, niveaux_map, config, dry_run
            )

            # 7. Mouvements fonds janvier 2026
            self._import_mouvements_janvier(
                fichiers['basecalcul'], adherents_map, config, dry_run
            )

            if dry_run:
                transaction.set_rollback(True)
                self.stdout.write(self.style.WARNING("\n[DRY-RUN] Aucune donnée écrite en base."))
            else:
                self.stdout.write(self.style.SUCCESS("\nImport terminé avec succès !"))

    # ----------------------------------------------------------------
    def _reset_donnees(self):
        from apps.adherents.models import Adherent
        from apps.users.models import Utilisateur
        from apps.parametrage.models import ConfigExercice

        self.stdout.write("  Suppression des données existantes...")
        Utilisateur.objects.filter(role='MEMBRE').delete()
        Adherent.objects.all().delete()
        ConfigExercice.objects.filter(annee=2026).delete()
        self.stdout.write(self.style.SUCCESS("  OK"))

    # ----------------------------------------------------------------
    def _import_config_exercice(self, dry_run):
        from apps.parametrage.models import ConfigExercice

        self.stdout.write("\n[1/7] Configuration exercice 2026...")

        defaults = ConfigExercice.get_config_2026()
        defaults['est_ouvert'] = True
        defaults['date_ouverture'] = date(2026, 1, 1)

        if not dry_run:
            config, created = ConfigExercice.objects.update_or_create(
                annee=2026, defaults=defaults
            )
            action = "créée" if created else "mise à jour"
            self.stdout.write(self.style.SUCCESS(f"  Configuration 2026 {action}"))
        else:
            self.stdout.write(f"  [DRY-RUN] ConfigExercice 2026 serait créée avec {len(defaults)} paramètres")
            config = ConfigExercice(annee=2026, **defaults)

        return config

    # ----------------------------------------------------------------
    def _import_adherents(self, chemin, dry_run):
        import openpyxl
        from apps.adherents.models import Adherent

        self.stdout.write("\n[2/7] Import des adhérents...")

        wb = openpyxl.load_workbook(chemin, data_only=True)
        ws = wb['LISTADHERENT']

        adherents_map = {}
        crees = 0
        maj = 0

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 3:
                continue
            if not row[0] or not str(row[0]).startswith('AS'):
                continue

            matricule  = str(row[0]).strip()
            nom_prenom = str(row[2]).strip() if row[2] else ''
            statut     = str(row[8]).strip() if row[8] else 'ACTIF'
            if statut not in ('ACTIF', 'INACTIF'):
                statut = 'ACTIF'

            data = {
                'numero_ordre' : int(row[1]) if row[1] else 0,
                'nom_prenom'   : nom_prenom,
                'fonction'     : str(row[3]).strip() if row[3] else '',
                'telephone1'   : nettoyer_telephone(row[5]),
                'telephone2'   : nettoyer_telephone(row[6]),
                'residence'    : str(row[7]).strip() if row[7] else '',
                'statut'       : statut,
            }

            if not dry_run:
                adherent, created = Adherent.objects.update_or_create(
                    matricule=matricule, defaults=data
                )
                if created:
                    crees += 1
                else:
                    maj += 1
            else:
                from apps.adherents.models import Adherent as A
                adherent = A(matricule=matricule, **data)

            adherents_map[matricule] = adherent
            self.stdout.write(
                f"  {'[+]' if not dry_run else '[?]'} {matricule} | "
                f"{nom_prenom[:30]:30} | {statut}"
            )

        if not dry_run:
            self.stdout.write(self.style.SUCCESS(
                f"  Adhérents : {crees} créés, {maj} mis à jour"
            ))
        else:
            self.stdout.write(f"  [DRY-RUN] {len(adherents_map)} adhérents seraient importés")

        return adherents_map

    # ----------------------------------------------------------------
    def _creer_comptes_utilisateurs(self, adherents_map, dry_run):
        from apps.users.models import Utilisateur

        self.stdout.write("\n[3/7] Création des comptes utilisateurs...")

        # Compte bureau partagé
        if not dry_run:
            bureau, created = Utilisateur.objects.update_or_create(
                username='tontine_2026',
                defaults={
                    'nom_complet' : 'Bureau ASELBY',
                    'role'        : Utilisateur.BUREAU,
                    'is_staff'    : True,
                    'est_actif'   : True,
                }
            )
            if created:
                bureau.set_password('tontine_2026')
                bureau.save()
                self.stdout.write(self.style.SUCCESS(
                    "  [+] Compte bureau : username=tontine_2026  mdp=tontine_2026"
                ))
        else:
            self.stdout.write("  [DRY-RUN] Compte bureau tontine_2026 serait créé")

        # Comptes membres
        crees = 0
        doublons = []

        for matricule, adherent in adherents_map.items():
            if matricule == 'AS201648':
                # ASELBY = compte institution, pas de compte membre
                continue

            username = normaliser_username(adherent.nom_prenom)
            if not username:
                username = matricule.lower()

            # Vérifier doublon de username
            if username in doublons:
                username = f"{username}_{matricule[-2:]}"
            doublons.append(username)

            # Mot de passe par défaut : numéro de téléphone si dispo, sinon aselby_2026
            tel = getattr(adherent, 'telephone1', '')
            mdp_defaut = tel if len(tel) >= 9 else 'aselby_2026'

            if not dry_run:
                user, created = Utilisateur.objects.update_or_create(
                    username=username,
                    defaults={
                        'nom_complet' : adherent.nom_prenom,
                        'role'        : Utilisateur.MEMBRE,
                        'adherent'    : adherent,
                        'est_actif'   : adherent.statut == 'ACTIF',
                    }
                )
                if created:
                    user.set_password(mdp_defaut)
                    user.save()
                    crees += 1

                source_mdp = f"tel ({tel})" if tel and len(tel) >= 9 else "aselby_2026"
                self.stdout.write(
                    f"  {'[+]' if created else '[=]'} {matricule} | "
                    f"username={username:20} | mdp={source_mdp}"
                )
            else:
                source_mdp = f"tel ({tel})" if tel and len(tel) >= 9 else "aselby_2026"
                self.stdout.write(
                    f"  [?] {matricule} | username={username:20} | mdp={source_mdp}"
                )

        if not dry_run:
            self.stdout.write(self.style.SUCCESS(f"  Comptes membres : {crees} créés"))

    # ----------------------------------------------------------------
    def _creer_niveaux_tontine(self, config, dry_run):
        from apps.tontines.models import NiveauTontine

        self.stdout.write("\n[4/7] Création des niveaux de tontine...")

        niveaux = [
            {
                'code'                      : 'T35',
                'taux_mensuel'              : Decimal('35000'),
                'versement_mensuel_par_part': Decimal('60000'),
                'diviseur_interet'          : 35,
            },
            {
                'code'                      : 'T75',
                'taux_mensuel'              : Decimal('75000'),
                'versement_mensuel_par_part': Decimal('75000'),
                'diviseur_interet'          : 1,
            },
            {
                'code'                      : 'T100',
                'taux_mensuel'              : Decimal('100000'),
                'versement_mensuel_par_part': Decimal('100000'),
                'diviseur_interet'          : 1,
            },
        ]

        niveaux_map = {}
        for n in niveaux:
            if not dry_run:
                niveau, created = NiveauTontine.objects.update_or_create(
                    config_exercice=config,
                    code=n['code'],
                    defaults=n
                )
                niveaux_map[n['code']] = niveau
                self.stdout.write(self.style.SUCCESS(
                    f"  [+] {n['code']} : {n['taux_mensuel']:,.0f} FCFA/mois "
                    f"(versement {n['versement_mensuel_par_part']:,.0f})"
                ))
            else:
                niveaux_map[n['code']] = NiveauTontine(**n)
                self.stdout.write(f"  [?] {n['code']} serait créé")

        return niveaux_map

    # ----------------------------------------------------------------
    def _import_fonds_initiaux(self, chemin, adherents_map, config, dry_run):
        import openpyxl
        from apps.fonds.models import MouvementFonds

        self.stdout.write("\n[5/7] Import des fonds de caisse initiaux...")

        wb = openpyxl.load_workbook(chemin, data_only=True)
        ws = wb['FONDSCAISSE']

        crees = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 2:
                continue
            if not row[0] or not str(row[0]).startswith('AS'):
                continue

            matricule = str(row[0]).strip()
            if matricule not in adherents_map:
                continue

            # Colonnes : matricule, ordre, nom, fonds_depart, reconduction, capital_compose, nouveau_fonds
            fonds_depart      = Decimal(str(row[3])) if row[3] else Decimal('0')
            reconduction      = Decimal(str(row[4])) if row[4] else Decimal('0')
            capital_compose   = Decimal(str(row[5])) if row[5] else Decimal('0')

            if not dry_run:
                adherent = adherents_map[matricule]
                mvt, created = MouvementFonds.objects.update_or_create(
                    adherent=adherent,
                    mois=0,       # mois=0 = situation initiale (report de l'exercice précédent)
                    annee=2026,
                    defaults={
                        'config_exercice'         : config,
                        'reconduction'             : reconduction,
                        'capital_compose_precedent': fonds_depart,
                        'fonds_definitif'          : fonds_depart,
                        'capital_compose'          : fonds_depart + capital_compose,
                    }
                )
                if created:
                    crees += 1
                self.stdout.write(
                    f"  [+] {matricule} | fonds={fonds_depart:>12,.0f} | "
                    f"capital_composé={capital_compose:>12,.0f}"
                )
            else:
                self.stdout.write(
                    f"  [?] {matricule} | fonds={fonds_depart:>12,.0f}"
                )

        if not dry_run:
            self.stdout.write(self.style.SUCCESS(f"  Fonds : {crees} enregistrements créés"))

    # ----------------------------------------------------------------
    def _import_parts_tontine(self, chemin_tabbord, chemin_t35, chemin_t75, chemin_t100,
                               adherents_map, niveaux_map, config, dry_run):
        import openpyxl
        from apps.tontines.models import SessionTontine, ParticipationTontine

        self.stdout.write("\n[6/7] Import des parts de tontine (Janvier 2026)...")

        # Créer les sessions janvier 2026
        sessions = {}
        date_seance_janv = date(2026, 1, 18)  # 3ème dimanche de janvier 2026

        for code, niveau in niveaux_map.items():
            if dry_run:
                sessions[code] = None
                self.stdout.write(f"  [?] Session {code} Janvier 2026 serait créée")
            else:
                session, _ = SessionTontine.objects.update_or_create(
                    niveau=niveau,
                    mois=1,
                    annee=2026,
                    defaults={'date_seance': date_seance_janv}
                )
                sessions[code] = session
                self.stdout.write(self.style.SUCCESS(
                    f"  [+] Session {code} - Janvier 2026 ({date_seance_janv})"
                ))

        # Lire les parts depuis le tableau de bord
        wb = openpyxl.load_workbook(chemin_tabbord, data_only=True)
        ws = wb['HISTOJANV26']

        crees = 0
        for i in range(6, 44):
            mat = ws.cell(row=i, column=1).value
            if not mat or not str(mat).startswith('AS'):
                continue
            if mat not in adherents_map:
                continue

            parts_t35  = int(ws.cell(row=i, column=10).value or 0)
            parts_t75  = int(ws.cell(row=i, column=12).value or 0)
            parts_t100 = int(ws.cell(row=i, column=14).value or 0)

            # Versements réels
            versement_banque  = Decimal(str(ws.cell(row=i, column=4).value or 0))
            versement_especes = Decimal(str(ws.cell(row=i, column=5).value or 0))

            # Mode
            if versement_banque > 0:
                mode = 'BANQUE'
            elif versement_especes > 0:
                mode = 'ESPECES'
            else:
                mode = 'ECHEC'

            adherent = adherents_map[mat]

            parts = [('T35', parts_t35), ('T75', parts_t75), ('T100', parts_t100)]
            for code, nb_parts in parts:
                if nb_parts == 0:
                    continue
                if not dry_run and sessions[code]:
                    niveau    = niveaux_map[code]
                    montant   = Decimal(str(nb_parts)) * niveau.versement_mensuel_par_part
                    if mode == 'ECHEC':
                        montant = Decimal('0')

                    ParticipationTontine.objects.update_or_create(
                        session=sessions[code],
                        adherent=adherent,
                        defaults={
                            'nombre_parts'  : nb_parts,
                            'mode_versement': mode,
                            'montant_verse' : montant if mode != 'ECHEC' else Decimal('0'),
                        }
                    )
                    crees += 1

                self.stdout.write(
                    f"  {'[+]' if not dry_run else '[?]'} {mat} | "
                    f"{code} {nb_parts} part(s) | {mode}"
                )

        if not dry_run:
            self.stdout.write(self.style.SUCCESS(f"  Participations : {crees} créées"))

    # ----------------------------------------------------------------
    def _import_mouvements_janvier(self, chemin, adherents_map, config, dry_run):
        import openpyxl
        from apps.fonds.models import MouvementFonds

        self.stdout.write("\n[7/7] Import des mouvements fonds Janvier 2026...")

        wb = openpyxl.load_workbook(chemin, data_only=True)
        ws = wb['MVTJANV26']

        crees = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 5:
                continue
            if not row[0] or not str(row[0]).startswith('AS'):
                continue

            mat = str(row[0]).strip()
            if mat not in adherents_map:
                continue

            # Colonnes BASECALCUL :
            # 0=mat 1=nom 2=fonds_départ 3=reconduction 4=retrait 5=fonds_def
            # 6=base_calcul 7=interet 8=capital 9=sanction 10=reste 11=epargne
            # 12=fonds_rlmt 13=frais_excep 14=collation
            fonds_def     = Decimal(str(row[5])) if row[5] else Decimal('0')
            base_calcul   = Decimal(str(row[6])) if row[6] else Decimal('0')
            interet       = Decimal(str(row[7])) if row[7] else Decimal('0')
            capital       = Decimal(str(row[8])) if row[8] else Decimal('0')
            reste         = Decimal(str(row[10])) if row[10] else Decimal('0')
            epargne       = Decimal(str(row[11])) if row[11] else Decimal('0')
            fonds_rlt     = Decimal(str(row[12])) if row[12] else Decimal('0')
            frais_excep   = Decimal(str(row[13])) if row[13] else Decimal('0')
            collation_val = Decimal(str(row[14])) if row[14] else Decimal('0')

            if not dry_run:
                adherent = adherents_map[mat]
                mvt, created = MouvementFonds.objects.update_or_create(
                    adherent=adherent,
                    mois=1,
                    annee=2026,
                    defaults={
                        'config_exercice'         : config,
                        'reste'                   : reste,
                        'epargne_nette'           : epargne,
                        'fonds_roulement'         : fonds_rlt,
                        'frais_exceptionnels'     : frais_excep,
                        'collation'               : collation_val,
                        'fonds_definitif'         : fonds_def,
                        'base_calcul_interet'     : base_calcul,
                        'interet_attribue'        : interet,
                        'capital_compose'         : capital,
                    }
                )
                if created:
                    crees += 1

                self.stdout.write(
                    f"  [+] {mat} | fonds_def={fonds_def:>12,.0f} | "
                    f"base={base_calcul:>12,.0f} | interet={interet:>8,.2f}"
                )
            else:
                self.stdout.write(
                    f"  [?] {mat} | fonds_def={fonds_def:>12,.0f}"
                )

        if not dry_run:
            self.stdout.write(self.style.SUCCESS(
                f"  Mouvements janvier : {crees} créés"
            ))

        self.stdout.write(self.style.SUCCESS(
            "\n" + "="*60 +
            "\nImport ASELBY 2026 terminé !\n" +
            "="*60 +
            "\n\nComptes créés :" +
            "\n  Bureau  : username=tontine_2026  mdp=tontine_2026" +
            "\n  Membres : username=<premier_nom>  mdp=<tel> ou aselby_2026" +
            "\n\nLes membres peuvent changer leur mot de passe via" +
            "\n'Mot de passe oublié' sur la page de connexion.\n"
        ))
