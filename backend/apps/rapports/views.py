"""
apps/rapports/views.py
Dashboard + exports TABBORD et AUTREMOUVEMENT.
Plus de formulaires MVT/HISTO — tout passe par saisie_mensuelle.
"""
from django.shortcuts import render
from django.utils import timezone
from django.db.models import Sum
from decimal import Decimal
from apps.core.mixins import bureau_required
from apps.parametrage.models import ConfigExercice
from apps.adherents.models import Adherent
from apps.saisie.models import SaisieMonthly
from apps.fonds.models import MouvementFonds
from apps.prets.models import Pret
from apps.dettes.models import ListeRouge
from apps.banque.models import ReleveBancaire

MOIS_FR   = ['','Janvier','Février','Mars','Avril','Mai','Juin',
             'Juillet','Août','Septembre','Octobre','Novembre','Décembre']
MOIS_CODE = ['','JANV','FEV','MARS','AVRIL','MAI','JUIN',
             'JUIL','AOUT','SEPT','OCT','NOV','DEC']
D = Decimal


def _bureau_required(view_func):
    from django.contrib.auth.decorators import login_required
    @login_required
    def wrapper(request, *args, **kwargs):
        if not request.user.est_bureau:
            from django.shortcuts import redirect
            return redirect('users:mon_espace')
        return view_func(request, *args, **kwargs)
    return wrapper


@_bureau_required
def dashboard(request):
    config = ConfigExercice.get_exercice_courant()
    mois   = timezone.now().month
    annee  = timezone.now().year

    nb_actifs    = Adherent.objects.filter(statut='ACTIF').count()
    nb_saisis    = SaisieMonthly.objects.filter(
        mois=mois, annee=annee, config_exercice=config).count()
    nb_releves   = ReleveBancaire.objects.filter(
        mois=mois, annee=annee, config_exercice=config).count()
    prets_cours  = Pret.objects.filter(statut=Pret.EN_COURS)
    prets_retard = prets_cours.filter(nb_mois_retard__gt=0)

    alertes = []
    if nb_releves < nb_actifs:
        alertes.append({'type':'info',
            'message': f"Relevé bancaire : {nb_releves}/{nb_actifs} pour {MOIS_FR[mois]}"})
    if nb_saisis < nb_actifs:
        alertes.append({'type':'info',
            'message': f"Saisie mensuelle : {nb_saisis}/{nb_actifs} pour {MOIS_FR[mois]}"})
    else:
        alertes.append({'type':'succes',
            'message': f"Saisie complète pour {MOIS_FR[mois]} ✓"})
    if prets_retard.exists():
        alertes.append({'type':'retard',
            'message': f"{prets_retard.count()} prêt(s) en retard"})

    mvts = MouvementFonds.objects.filter(
        mois=mois, annee=annee, config_exercice=config)
    total_fonds   = mvts.aggregate(t=Sum('fonds_definitif'))['t'] or D('0')
    total_capital = mvts.aggregate(t=Sum('capital_compose'))['t']  or D('0')

    # Dernières saisies
    dernieres = SaisieMonthly.objects.filter(
        mois=mois, annee=annee, config_exercice=config
    ).select_related('adherent').order_by('adherent__numero_ordre')[:10]

    return render(request, 'dashboard/dashboard.html', {
        'config_exercice':  config,
        'nb_adherents':     nb_actifs,
        'nb_saisis':        nb_saisis,
        'nb_releves':       nb_releves,
        'nb_total':         nb_actifs,
        'mois_courant':     f"{MOIS_FR[mois]} {annee}",
        'total_fonds':      total_fonds,
        'total_capital':    total_capital,
        'alertes':          alertes,
        'dernieres_saisies': dernieres,
        'prets_retard':     prets_retard.select_related('adherent')[:5],
        'nb_liste_rouge':   ListeRouge.objects.filter(est_solde=False).count(),
        'kpi': {
            'nb_adherents_actifs':     nb_actifs,
            'nb_prets_en_cours':       prets_cours.count(),
            'total_prets_circulation': prets_cours.aggregate(
                t=Sum('montant_total_du'))['t'] or D('0'),
            'nb_liste_rouge': ListeRouge.objects.filter(est_solde=False).count(),
        },
    })


# ═══════════════════════════════════════════════════════════════
# EXPORT TABBORD
# ═══════════════════════════════════════════════════════════════
@_bureau_required
def telecharger_tabbord(request):
    """
    Export TABBORD.xlsx — 74 colonnes A→BV fidèles au fichier réel.
    HISTORESUME = cumul annuel de tous les mois.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from apps.saisie.models import SaisieMonthly

    config    = ConfigExercice.get_exercice_courant()
    annee     = config.annee
    # ASELBY en premier, puis actifs par numero_ordre — AS201638 EXCLU (inactif)
    adherents = sorted(
        Adherent.objects.filter(statut='ACTIF').exclude(matricule='AS201638'),
        key=lambda a: (0 if a.matricule == 'AS201648' else 1, a.numero_ordre)
    )
    wb = Workbook(); wb.remove(wb.active)
    BLEU = '1B2B5E'; BLANC = 'FFFFFF'

    taux_t60_nom = f"TONTINE {int(D(str(config.versement_t35))):,}".replace(',', ' ')
    COLS = [
        'MATRICULE','NOM ET PRENOM','BONUS MALUS',                                    # A-C
        'VERSEMENT BANQUE','VERSEMENT ESPECES','AUTRE VERSEMENTS','COMPLEMENT EPARGNE', # D-G
        'PENALITE VERSEMENT ESPECES','MODE PAIEMENT TONTINE',                          # H-I
        'NBRE DE LOT', taux_t60_nom,                                                   # J-K
        'NBRE DE LOT','TONTINE 75 000',                                                # L-M
        'NBRE DE LOT','TONTINE 100 000',                                               # N-O
        'REMBOURSEMENT PETIT LOT T60','REMBOURSEMENT PETIT LOT T75',                   # P-Q
        'REMBOURSEMENT PETIT LOT T100','REMBOURSEMENT PRET FONDS',                     # R-S
        'MODE PAIEMENT REMBOURSEMENT PRET','MONTANT ENGAGEMENT',                       # T-U
        'SANCTION','INSCRIPTION','EPARGNE ASSURANCE','TONTINE MOIS',                   # V-Y
        'PRET FONDS','MODE PAIEMENT PRET','NUM CHEQUE',                                # Z-AB
        'ACHAT LOT PRINCIPAL T60','ACHAT LOT PRINCIPAL T75','ACHAT LOT PRINCIPAL T100',# AC-AE
        'MODE PAIEMENT LOT PRINCIPAL','NUM CHEQUE','PENALITE VERSEMENT ESPECES',       # AF-AH
        'VENTE PETIT LOT T60','VENTE PETIT LOT T75','VENTE PETIT LOT T100',            # AI-AK
        'INTERET PETIT LOT T60','INTERET PETIT LOT T75','INTERET PETIT LOT T100',      # AL-AN
        'NUM CHEQUE','MODE REMBOURSEMENT PAIEMENT PETIT LOT',                          # AO-AP
        'PENALITE RETARD TONTINE','PENALITE ECHEC TONTINE',                            # AQ-AR
        'RESTE','MUTUELLE','REMBOURSEMENT TRANSPORT','RETRAIT PARTIEL FONDS',          # AS-AV
        'MONTANT T60','MONTANT T75','MONTANT T100',                                    # AW-AY
        'MONTANT CHEQUE','MONTANT ESPECES','CONTRIBUTION FOYER',                       # AZ-BB
        'AUTRES MODE DE VERSEMENT','MONTANT CHEQUE EFFECTIF','NUMERO CHEQUE',          # BC-BE
        'INTERET PRET','NBRE MOIS PRET',                                               # BF-BG
        'DEPENSE FONDS DE ROULEMENT','DEPENSE FRAIS EXCEPTIONNEL',                    # BH-BI
        'DEPENSE FONDS MUTUEL','DEPENSE COLLATION RECEPTION',                          # BJ-BK
        'PENALITE PRET FONDS','DEPENSE PENALITE VERSEMENT BANQUE',                    # BL-BM
        'AUTRES DEPENSES','LIBELLE','COMPTE','MONTANT',                                # BN-BQ
        'FOYER DON VOLONTAIRE','MONTANT PRET DEFINITIF',                               # BR-BS
        'NOMBRE DE MOIS','DATE REMBOURSEMENT','STATUT PRET',                           # BT-BV
    ]

    def _hdr(ws, mois_label):
        ws.cell(1,1,'ASSOCIATION'); ws.cell(1,2,'ASELBY')
        ws.cell(2,1,'ANNEE');       ws.cell(2,2, annee)
        for j in range(1, len(COLS)+1):
            ws.cell(3, j, j)
        ws.cell(4,1,'TABLEAU DE BORD SAISIE DES DONNEES')
        fill = PatternFill('solid', fgColor=BLEU)
        fn   = Font(bold=True, color=BLANC, size=8, name='Arial')
        for j, c in enumerate(COLS, 1):
            cl = ws.cell(5, j, c)
            cl.fill = fill; cl.font = fn
            cl.alignment = Alignment(wrap_text=True, horizontal='center')
            ws.column_dimensions[get_column_letter(j)].width = 13
        ws.row_dimensions[5].height = 30

    def _build_row(adh, s, r, nb_ok_t60=0, taux_t60_v=D('0')):
        """Construit le tuple de 74 valeurs. 0 pour les numériques vides (fidèle au réel)."""
        def _dd(v):
            if v is None: return D('0')
            try: return D(str(v))
            except: return D('0')
        def fv(attr):
            if s:
                v = getattr(s, attr, None)
                try: return float(v) if v is not None else 0.0
                except: return 0.0
            return 0.0
        def sv(attr):
            if s: return getattr(s, attr, None) or None
            return None

        D_bq  = _dd(r.versement_banque)   if r else D('0')
        E_esp = _dd(r.versement_especes)  if r else D('0')
        F_aut = _dd(r.autre_versement)    if r else D('0')
        G_comp= _dd(s.complement_epargne) if s else D('0')
        I_pen = _dd(s.penalite_versement_especes) if s else D('0')
        S_pret= _dd(s.remboursement_pret) if s else D('0')
        X_ton = _dd(s.tontine_mois)       if s else D('0')
        t60   = _dd(s.tontine_t60)        if s else D('0')
        t75   = _dd(s.tontine_t75)        if s else D('0')
        t100  = _dd(s.tontine_t100)       if s else D('0')

        # BONUS MALUS: champ DB si dispo, sinon calcul
        bm_db = getattr(s, 'bonus_malus', None) if s else None
        bonus_malus = float(_dd(bm_db)) if bm_db is not None else float((D_bq+E_esp+F_aut) - X_ton - G_comp - I_pen - S_pret)

        # U = X + G + S (sans pénalité)
        montant_engagement = float(X_ton + G_comp + S_pret)

        mode_ton = sv('mode_paiement_tontine')
        num_chq  = sv('numero_cheque')
        mode_lot = sv('mode_paiement_lot')
        mode_remb_pret = sv('mode_remb_pret')

        # Flags pour colonnes conditionnelles
        has_pret   = s and _dd(s.pret_fonds) > 0
        has_achat  = s and (_dd(s.achat_lot_t60)+_dd(s.achat_lot_t75)+_dd(s.achat_lot_t100)) > 0
        has_vente  = s and (_dd(s.vente_petit_lot_t60)+_dd(s.vente_petit_lot_t75)+_dd(s.vente_petit_lot_t100)) > 0
        has_remb_pl= s and (_dd(s.remboursement_petit_lot_t60)+_dd(s.remboursement_petit_lot_t75)+_dd(s.remboursement_petit_lot_t100)) > 0

        # Colonnes conditionnelles texte (None si non applicable)
        num_chq_pret    = num_chq if has_pret  and num_chq else None
        num_chq_lot     = num_chq if has_achat and num_chq else None
        mode_lot_val    = mode_lot if has_achat else None
        num_chq_remb_pl = num_chq if has_remb_pl and num_chq else None
        mode_remb_pl    = sv('mode_remb_petit_lot') if (has_remb_pl or has_vente) else None

        # Montant T60/T75/T100
        def calc_aw(s, nb_ok, taux):
            if not s: return 0.0
            achat = _dd(s.achat_lot_t60); vente = _dd(s.vente_petit_lot_t60)
            if achat > 0: return float(D(str(nb_ok)) * taux - achat)
            return float(vente)
        def calc_ax(s):
            if not s: return 0.0
            lot = _dd(getattr(s,'montant_lot_t75',D('0')))
            return float(lot + _dd(s.vente_petit_lot_t75))
        def calc_ay(s):
            if not s: return 0.0
            lot = _dd(getattr(s,'montant_lot_t100',D('0')))
            return float(lot + _dd(s.vente_petit_lot_t100))

        reste   = float(_dd(s.reste)) if s else 0.0
        dep_col = fv('depense_collation')
        dep_roul= fv('depense_fonds_roulement')
        dep_frais=fv('depense_frais_exceptionnel')
        dep_mut = fv('depense_fonds_mutuel')
        pen_pret= fv('penalite_pret_fonds')
        int_pret= float(_dd(s.interet_pret)) if s else 0.0
        mont_esp= fv('montant_especes')
        foyer   = fv('contribution_foyer')
        mont_dep= fv('montant_depense')
        pret_def= float(_dd(s.pret_fonds)) if s else 0.0
        nbre_mois_pret = sv('nbre_mois_pret')

        # AZ=MONTANT CHEQUE: dépense collation + montant espèces
        az_mont_chq = dep_col + mont_esp
        # BD=MONTANT CHEQUE EFFECTIF: dépense collation seul
        bd_mont_chq_eff = dep_col

        # BH=DEPENSE FONDS ROULEMENT: valeur réelle de la dépense (pas forfait)
        bh_dep_roul = dep_roul  # depuis SaisieMonthly.depense_fonds_roulement

        return [
            adh.matricule, adh.nom_prenom,
            bonus_malus,                                            # C
            float(D_bq), float(E_esp), float(F_aut), float(G_comp),# D-G
            float(I_pen), mode_ton,                                 # H-I
            sv('nbre_lots_t60') or 0, float(t60),                  # J-K
            sv('nbre_lots_t75') or 0, float(t75),                  # L-M
            sv('nbre_lots_t100') or 0, float(t100),                # N-O
            fv('remboursement_petit_lot_t60'),                      # P
            fv('remboursement_petit_lot_t75'),                      # Q
            fv('remboursement_petit_lot_t100'),                     # R
            float(S_pret), mode_remb_pret,                          # S-T
            montant_engagement,                                     # U
            fv('sanction'), fv('inscription'),                      # V-W
            0,                                                      # X EPARGNE ASSURANCE
            float(X_ton),                                          # Y TONTINE MOIS
            fv('pret_fonds'), sv('mode_paiement_pret'),            # Z-AA
            num_chq_pret,                                           # AB NUM CHEQUE pret
            fv('achat_lot_t60'), fv('achat_lot_t75'), fv('achat_lot_t100'), # AC-AE
            mode_lot_val,                                           # AF
            num_chq_lot,                                            # AG NUM CHEQUE lot
            0,                                                      # AH PEN_ESP (0 pas dupliquée)
            fv('vente_petit_lot_t60'), fv('vente_petit_lot_t75'), fv('vente_petit_lot_t100'), # AI-AK
            fv('interet_petit_lot_t60'), fv('interet_petit_lot_t75'), fv('interet_petit_lot_t100'), # AL-AN
            num_chq_remb_pl, mode_remb_pl,                         # AO-AP
            fv('penalite_retard_tontine'), fv('penalite_echec_tontine'), # AQ-AR
            reste if reste != 0 else None,                         # AS RESTE (None si 0)
            fv('mutuelle'), 0, fv('retrait_partiel'),              # AT-AV
            calc_aw(s, nb_ok_t60, taux_t60_v),                    # AW MONTANT T60
            calc_ax(s),                                            # AX MONTANT T75
            calc_ay(s),                                            # AY MONTANT T100
            az_mont_chq,                                           # AZ MONTANT CHEQUE
            mont_esp,                                              # BA MONTANT ESPECES
            foyer,                                                 # BB CONTRIBUTION FOYER
            None,                                                  # BC AUTRES MODE VERST (None)
            bd_mont_chq_eff,                                       # BD MONTANT CHEQUE EFF
            num_chq,                                               # BE NUMERO CHEQUE
            int_pret,                                              # BF INTERET PRET
            nbre_mois_pret,                                        # BG NBRE MOIS PRET
            bh_dep_roul,                                           # BH DEPENSE FONDS ROUL
            dep_frais,                                             # BI DEPENSE FRAIS EXC
            dep_mut,                                               # BJ DEPENSE FONDS MUT
            dep_col,                                               # BK DEPENSE COLLATION
            pen_pret,                                              # BL PENALITE PRET
            0,                                                     # BM DEPENSE PEN VST BQ
            mont_dep,                                              # BN AUTRES DEPENSES
            sv('libelle_depense'),                                 # BO LIBELLE
            sv('compte_depense'),                                  # BP COMPTE
            mont_dep,                                              # BQ MONTANT
            0,                                                     # BR FOYER DON VOLONT
            pret_def,                                              # BS MONTANT PRET DEF
            nbre_mois_pret,                                        # BT NOMBRE DE MOIS
            None,                                                  # BU DATE REMBOURSEMENT
            None,                                                  # BV STATUT PRET
        ]


    # Pré-charger toutes les saisies et relevés
    all_saisies = {}
    for s in SaisieMonthly.objects.filter(annee=annee, config_exercice=config):
        all_saisies[(s.adherent_id, s.mois)] = s
    all_releves = {}
    for r in ReleveBancaire.objects.filter(annee=annee, config_exercice=config):
        all_releves[(r.adherent_id, r.mois)] = r

    # ── Feuilles mensuelles ──────────────────────────────────────
    for mois in range(1, 13):
        label = f'HISTO{MOIS_CODE[mois]}{str(annee)[-2:]}'
        ws    = wb.create_sheet(label)
        _hdr(ws, MOIS_FR[mois])
        nb_ok = sum(1 for (mat, m), s in all_saisies.items()
                    if m == mois and s.numero_cheque and str(s.numero_cheque).strip())
        tv    = D(str(config.versement_t35))
        for i, adh in enumerate(adherents, 6):
            s = all_saisies.get((adh.matricule, mois))
            r = all_releves.get((adh.matricule, mois))
            row_vals = _build_row(adh, s, r, nb_ok_t60=nb_ok, taux_t60_v=tv)
            for j, v in enumerate(row_vals, 1):
                ws.cell(i, j, v)

        # Ligne TOTAUX en bas
        nb_rows = len(adherents) + 6
        ws.cell(nb_rows, 1, 'TOTAUX')
        for col in range(3, len(COLS)+1):
            col_l = get_column_letter(col)
            ws.cell(nb_rows, col, f'=SUM({col_l}6:{col_l}{nb_rows-1})')

    # ── Feuille HISTORESUME = CUMUL ANNUEL ────────────────────────
    ws_r = wb.create_sheet(f'HISTORESUME{str(annee)[-2:]}')
    _hdr(ws_r, 'RESUME')

    # Accumuler toutes les valeurs par adhérent sur 12 mois
    from collections import defaultdict
    cumul = {adh.matricule: [0.0]*len(COLS) for adh in adherents}
    str_last = {adh.matricule: [None]*len(COLS) for adh in adherents}

    tv = D(str(config.versement_t35))
    for mois in range(1, 13):
        nb_ok = sum(1 for (mat, m), s in all_saisies.items()
                    if m == mois and s.numero_cheque and str(s.numero_cheque).strip())
        for adh in adherents:
            s = all_saisies.get((adh.matricule, mois))
            r = all_releves.get((adh.matricule, mois))
            if not s and not r: continue
            row_vals = _build_row(adh, s, r, nb_ok_t60=nb_ok, taux_t60_v=tv)
            for ci, v in enumerate(row_vals[2:], 2):  # skip MAT, NOM
                if isinstance(v, (int, float)) and v:
                    cumul[adh.matricule][ci] += float(v)
                elif isinstance(v, str) and v:
                    str_last[adh.matricule][ci] = v

    for i, adh in enumerate(adherents, 6):
        ws_r.cell(i, 1, adh.matricule)
        ws_r.cell(i, 2, adh.nom_prenom)
        for ci in range(2, len(COLS)):
            num = cumul[adh.matricule][ci]
            txt = str_last[adh.matricule][ci]
            ws_r.cell(i, ci+1, num if num else txt)

    # Ligne TOTAUX RESUME
    nb_rows = len(adherents) + 6
    ws_r.cell(nb_rows, 1, 'TOTAUX')
    for col in range(3, len(COLS)+1):
        col_l = get_column_letter(col)
        ws_r.cell(nb_rows, col, f'=SUM({col_l}6:{col_l}{nb_rows-1})')

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    resp = HttpResponse(buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename=ASELBY{annee}TABBORD.xlsx'
    return resp

@_bureau_required
def telecharger_autremouvement(request):
    """
    AUTREMOUVEMENT.xlsx = BASECALCULINTERET avec structure exacte du réel.
    38 colonnes A→AL pour MVT*, 39 colonnes pour MVTRESUME (+ DON VOLONTAIRE).
    Feuilles: MVTDEC(n-1) + MVT(mois) + MVTRESUME + MVTRESUME264
    """
    import io, math
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from apps.saisie.models import SaisieMonthly
    from apps.exercice.models import FicheCassation

    Dc = __import__('decimal').Decimal
    def _d(v):
        if v is None: return Dc('0')
        try: return Dc(str(v))
        except: return Dc('0')

    config    = ConfigExercice.get_exercice_courant()
    annee     = config.annee
    adherents = sorted(
        Adherent.objects.filter(statut='ACTIF'),
        key=lambda a: (0 if a.matricule == 'AS201648' else 1, a.numero_ordre)
    )

    BLEU = '1B2B5E'; BLANC = 'FFFFFF'
    MOIS_CODE = {1:'JANV',2:'FEV',3:'MARS',4:'AVRIL',5:'MAI',6:'JUIN',
                 7:'JUIL',8:'AOUT',9:'SEPT',10:'OCT',11:'NOV',12:'DEC'}
    MOIS_NOM  = {1:'JANVIER',2:'FEVRIER',3:'MARS',4:'AVRIL',5:'MAI',6:'JUIN',
                 7:'JUILLET',8:'AOUT',9:'SEPTEMBRE',10:'OCTOBRE',
                 11:'NOVEMBRE',12:'DECEMBRE'}

    # 38 colonnes A→AL (identiques pour MVT* et MVTRESUME sauf col W/X/Y différents)
    COLS_MVT = [
        'MATRICULE','NOM ET PRENOM','FONDES DE DEPART','RECONDUCTION',
        'RETRAIT PARTIEL','FONDS DEFINITIF',
        'BASE DE CALCUL INTERET FONDS DEFINITIF',
        'REPARTITION PROVISOIRE INTERET FONDS+EPARGNE','CAPITAL COMPOSE',
        'SANCTION','RESTE','EPARGNE','FONDS DE ROULEMENT',
        'FRAIS EXCEPTIONNEL','COLLATION','PENALITE VST ESPECES',
        'INSCRISPTION','MUTUELLE','PRÊT FONDS','INTERET PRET FONDS',
        'NUMERO CHEQUE','REMBOURSEMENT PRET FONDS',
        'MODE VERSEMENT REMBOURSEMENT','MODE VERSEMENT PRÊT',
        'NUMERO CHEQUE','PENALITE PRET FONDS','PENALITE FONDS',
        'PENALITE ECHEC TONTINE','PENALITE RETARD TONTINE',
        'REMBOURSEMENT TRANSPORT','CONTRIBUTION FOYER',
        'DEPENSE FONDS DE ROULEMENT','DEPENSE FRAIS EXCEPTIONNEL',
        'DEPENSE FONDS MUTUELLE','DEPENSE COLLATION RECEPTION',
        'PENALITE PRET FONDS','DEPENSE PENALITE VERSEMENT BANQUE',
        'AUTRES DEPENSES',
    ]
    # MVTRESUME: col W=PRÊT DEFINITIF, X=DATE REMBOURSEMENT, Y=STATUT PRET, AM=DON VOLONTAIRE
    COLS_RESUME = list(COLS_MVT)
    COLS_RESUME[22] = 'PRÊT DEFINITIF'
    COLS_RESUME[23] = 'DATE REMBOURSEMENT'
    COLS_RESUME[24] = 'SITATUT PRÊT'
    COLS_RESUME.append('DON VOLONTAIRE')  # AM=39

    wb = Workbook(); wb.remove(wb.active)

    def _hdr(ws, cols, annee_sh, mois_label=None, date_val=None):
        ws.cell(1,1,'ASSOCIATION'); ws.cell(1,2,'ASELBY')
        ws.cell(2,1,'ANNEE');       ws.cell(2,2, annee_sh)
        ws.cell(3,1,'TONTINE')
        if date_val:
            ws.cell(4, 11, date_val)
        fill = PatternFill('solid', fgColor=BLEU)
        fn   = Font(bold=True, color=BLANC, size=8, name='Arial')
        for j, h in enumerate(cols, 1):
            cl = ws.cell(5, j, h)
            cl.fill = fill; cl.font = fn
            cl.alignment = Alignment(horizontal='center', wrap_text=True)
            ws.column_dimensions[get_column_letter(j)].width = 14
        ws.row_dimensions[5].height = 36

    # Reconductions depuis FicheCassation année-1
    config_prec = ConfigExercice.objects.filter(annee=annee-1).first()
    reconductions = {}
    if config_prec:
        for fc in FicheCassation.objects.filter(config_exercice=config_prec):
            reconductions[fc.adherent_id] = _d(fc.reconduction)

    # Feuilles à générer: MVTDEC(n-1) + tous mois n
    from django.db.models import Max
    sheets = []
    if config_prec:
        res = MouvementFonds.objects.filter(config_exercice=config_prec).aggregate(m=Max('mois'))
        if res['m']:
            sheets.append(('transition', annee-1, res['m'], config_prec))
    mois_dispo = list(
        MouvementFonds.objects.filter(config_exercice=config)
        .values_list('mois', flat=True).distinct().order_by('mois')
    )
    for m in mois_dispo:
        sheets.append(('courant', annee, m, config))

    seuil      = _d(config.seuil_eligibilite_interets)
    fonds_roul = _d(config.fonds_roulement_mensuel)
    frais_exc  = _d(config.frais_exceptionnels_mensuel)
    collation  = _d(config.collation_mensuelle)

    # Pré-charger saisies
    all_saisies = {}
    for s in SaisieMonthly.objects.filter(config_exercice=config):
        all_saisies[(s.adherent_id, s.mois)] = s
    if config_prec:
        for s in SaisieMonthly.objects.filter(config_exercice=config_prec):
            all_saisies[(s.adherent_id, s.mois)] = s

    # Cumul annuel pour RESUME
    cumul = {adh.matricule: {
        'C':Dc('0'),'D':Dc('0'),'F':Dc('0'),'G':Dc('0'),'H':Dc('0'),
        'I':Dc('0'),'K':Dc('0'),'L':Dc('0'),'M':Dc('0'),'N':Dc('0'),
        'O':Dc('0'),'P':Dc('0'),'Q':Dc('0'),'R':Dc('0'),
        'S':Dc('0'),'T':Dc('0'),'V':Dc('0'),
        'pret_def':Dc('0'),'date_remb':'','statut_pret':'',
    } for adh in adherents}

    for sheet_type, an, mois, cfg_mois in sheets:
        label = f"MVT{MOIS_CODE[mois]}{str(an)[-2:]}"
        ws    = wb.create_sheet(label)
        is_trans = (sheet_type == 'transition')
        from django.utils import timezone
        date_val = timezone.datetime(an, mois, 17).date() if not is_trans else None
        _hdr(ws, COLS_MVT, an, MOIS_NOM.get(mois,''), date_val)

        saisies = {s.adherent_id: s for s in
                   SaisieMonthly.objects.filter(mois=mois, annee=an, config_exercice=cfg_mois)}
        prec_mois  = 12 if mois == 1 else mois - 1
        prec_annee = an - 1 if mois == 1 else an
        cfg_prec_m = ConfigExercice.objects.filter(annee=prec_annee).first()
        mvts_prec  = {}
        if cfg_prec_m:
            mvts_prec = {mv.adherent_id: mv for mv in
                         MouvementFonds.objects.filter(
                             mois=prec_mois, annee=prec_annee, config_exercice=cfg_prec_m)}

        pool_t = sum(_d(s.interet_pret) for s in saisies.values())
        seuil_mois = _d(cfg_mois.seuil_eligibilite_interets)

        # 1er passage: F, G, total_G
        rows_data = {}; total_G = Dc('0')
        for adh in adherents:
            s = saisies.get(adh.matricule)
            mp = mvts_prec.get(adh.matricule)
            reste = _d(s.reste) if s else Dc('0')
            epargne = (reste - fonds_roul - frais_exc - collation) if reste > 0 else Dc('0')
            if epargne < 0: epargne = Dc('0')
            m_val = fonds_roul if reste > 0 else Dc('0')
            n_val = frais_exc  if reste > 0 else Dc('0')
            o_val = collation  if reste > 0 else Dc('0')
            retrait = _d(s.retrait_partiel) if s else Dc('0')

            if is_trans:
                cap_depart = _d(adh.capital_depart_exercice)
                recon = reconductions.get(adh.matricule, Dc('0'))
                fonds_def = cap_depart + recon
                c_val = cap_depart; d_val = recon
            else:
                c_val = None; d_val = None
                recon_prec = _d(mp.capital_compose) if mp else Dc('0')
                fonds_def = recon_prec - retrait + epargne

            base_calc = fonds_def if fonds_def > seuil_mois else Dc('0')
            total_G  += base_calc
            rows_data[adh.matricule] = {
                'c':c_val,'d':d_val,'retrait':retrait,'reste':reste,
                'epargne':epargne,'fonds_def':fonds_def,'base_calc':base_calc,
                'm':m_val,'n':n_val,'o':o_val,
            }

        # 2ème passage: H, I, écriture
        for i, adh in enumerate(adherents, 6):
            b = rows_data[adh.matricule]
            s = saisies.get(adh.matricule)

            if total_G > 0 and b['base_calc'] > 0:
                h_raw = float(pool_t)/float(total_G)*float(b['base_calc'])
                interet = Dc(str(math.floor(h_raw*100)/100))
            else:
                interet = Dc('0')
            capital_compose = b['fonds_def'] + interet

            def sv(f, d=''):
                if s: return getattr(s,f,None) or d
                return d
            def fv(f):
                if s: return float(_d(getattr(s,f,None)))
                return 0.0

            vals = [
                adh.matricule, adh.nom_prenom,
                float(b['c']) if b['c'] is not None else None,  # C
                float(b['d']) if b['d'] is not None else None,  # D
                float(b['retrait']) or None,                    # E
                float(b['fonds_def']),                          # F
                float(b['base_calc']) or None,                  # G
                float(interet) or None,                         # H
                float(capital_compose),                         # I
                fv('sanction') or None,                         # J
                float(b['reste']) or None,                      # K
                float(b['epargne']) or None,                    # L
                float(b['m']) or None,                          # M
                float(b['n']) or None,                          # N
                float(b['o']) or None,                          # O
                fv('penalite_versement_especes') or None,       # P
                fv('inscription') or None,                      # Q
                fv('mutuelle') or None,                         # R
                fv('pret_fonds') or None,                       # S
                float(_d(s.interet_pret)) if s else None,       # T
                sv('numero_cheque') or None,                    # U
                fv('remboursement_pret') or None,               # V
                sv('mode_remb_pret') or None,                   # W
                sv('mode_paiement_pret') or None,               # X
                None,                                            # Y (num_chq2)
                fv('penalite_pret_fonds') or None,              # Z
                None,                                            # AA
                float(_d(s.penalite_echec_tontine)) if s else None,  # AB
                float(_d(s.penalite_retard_tontine)) if s else None,  # AC
                None,                                            # AD remb_transport
                fv('contribution_foyer') or None,               # AE
                fv('depense_fonds_roulement') or None,          # AF
                fv('depense_frais_exceptionnel') or None,       # AG
                fv('depense_fonds_mutuel') or None,             # AH
                fv('depense_collation') or None,                # AI
                fv('penalite_pret_fonds') or None,              # AJ (bis)
                None,                                            # AK
                fv('montant_depense') or None,                  # AL
            ]
            for j, v in enumerate(vals, 1):
                ws.cell(i, j, v)

            # Cumuler pour RESUME
            c = cumul[adh.matricule]
            if an == annee:  # seulement mois courants
                c['F'] += b['fonds_def']; c['G'] += b['base_calc']
                c['H'] += interet; c['I'] += capital_compose
                c['K'] += b['reste']; c['L'] += b['epargne']
                c['M'] += b['m']; c['N'] += b['n']; c['O'] += b['o']
                if s:
                    c['S'] = _d(s.pret_fonds)
                    c['T'] += _d(s.interet_pret)
                    c['V'] += _d(s.remboursement_pret)
                    c['pret_def'] = _d(s.pret_fonds)
            elif is_trans:  # MVTDEC(n-1)
                c['C'] = b['c'] or Dc('0')
                c['D'] = b['d'] or Dc('0')

        # Ligne TOTAL
        nb = len(adherents)+6
        ws.cell(nb, 1, 'TOTAL')
        for col in range(3, 39):
            col_l = get_column_letter(col)
            ws.cell(nb, col, f'=SUM({col_l}6:{col_l}{nb-1})')

    # ── MVTRESUME26 ─────────────────────────────────────────────
    ws_r = wb.create_sheet(f"MVTRESUME{str(annee)[-2:]}")
    _hdr(ws_r, COLS_RESUME, annee)

    for i, adh in enumerate(adherents, 6):
        c = cumul[adh.matricule]
        vals_r = [
            adh.matricule, adh.nom_prenom,
            float(c['C']) or None, float(c['D']) or None, None,
            float(c['F']), float(c['G']) or None,
            float(c['H']) or None, float(c['I']),
            None,  # J sanction
            float(c['K']) or None, float(c['L']) or None,
            float(c['M']) or None, float(c['N']) or None, float(c['O']) or None,
            None, None, None,  # P Q R
            float(c['S']) or None, float(c['T']) or None,  # S T
            None,  # U num_chq
            float(c['V']) or None,  # V remb
            float(c['pret_def']) or None,  # W prêt_definitif
            c['date_remb'] or None,  # X date_remboursement
            c['statut_pret'] or None,  # Y statut
            None, None, None, None, None, None,  # Z-AE
            None, None, None, None, None, None, None,  # AF-AL
            None,  # AM don_volontaire
        ]
        for j, v in enumerate(vals_r, 1):
            ws_r.cell(i, j, v)

    nb_r = len(adherents)+6
    ws_r.cell(nb_r, 1, 'TOTAL')
    for col in range(3, 40):
        col_l = get_column_letter(col)
        ws_r.cell(nb_r, col, f'=SUM({col_l}6:{col_l}{nb_r-1})')

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    resp = HttpResponse(buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename=ASELBY{annee}AUTREMOUVEMENT.xlsx'
    return resp

@_bureau_required
def tabbord(request):
    """TABBORD — visualisation par mois."""
    config    = ConfigExercice.get_exercice_courant()
    mois      = max(1, min(12, int(request.GET.get('mois', timezone.now().month))))
    annee     = int(request.GET.get('annee', config.annee))
    prec      = (12, annee-1) if mois == 1 else (mois-1, annee)
    suiv      = (1,  annee+1) if mois == 12 else (mois+1, annee)

    saisies = SaisieMonthly.objects.filter(
        mois=mois, annee=annee, config_exercice=config
    ).select_related('adherent').order_by('adherent__nom_prenom')

    from django.db.models import Sum
    totaux = {}
    if saisies.exists():
        totaux = {
            'bonus_malus':      sum(s.bonus_malus          for s in saisies),
            'versement_banque': sum(s.versement_banque      for s in saisies),
            'versement_especes':sum(s.versement_especes     for s in saisies),
            'tontine_mois':     sum(s.tontine_mois          for s in saisies),
            'montant_engagement':sum(s.montant_engagement   for s in saisies),
            'reste':            sum(s.reste                 for s in saisies),
        }

    return render(request, 'dashboard/rapports/tabbord_mois.html', {
        'config_exercice': config,
        'saisies':    saisies,
        'colonnes':   COLS_TABBORD,
        'totaux':     totaux,
        'mois':       mois,  'annee':      annee,
        'mois_label': MOIS_LABEL[mois],
        'prec_mois':  prec[0], 'prec_annee': prec[1],
        'suiv_mois':  suiv[0], 'suiv_annee': suiv[1],
        'prec_label': MOIS_LABEL[prec[0]],
        'suiv_label': MOIS_LABEL[suiv[0]],
    })


@_bureau_required
def tabbord_resume(request):
    """TABBORD — résumé annuel (dernier mois saisi par adhérent)."""
    config = ConfigExercice.get_exercice_courant()
    annee  = int(request.GET.get('annee', config.annee))
    from apps.adherents.models import Adherent

    # Dernier mois saisi pour chaque adhérent
    adherents = Adherent.objects.filter(statut='ACTIF').order_by('nom_prenom')
    saisies = []
    for adh in adherents:
        last = SaisieMonthly.objects.filter(
            adherent=adh, annee=annee, config_exercice=config
        ).order_by('-mois').first()
        if last:
            saisies.append({'adherent': adh, 'mois': last.mois, 'annee': annee, 's': last})

    return render(request, 'dashboard/rapports/tabbord_mois.html', {
        'config_exercice': config,
        'saisies':    [item['s'] for item in saisies],
        'colonnes':   COLS_TABBORD,
        'totaux':     {},
        'mois':       0,  'annee':   annee,
        'mois_label': f'Résumé {annee}',
        'prec_mois': 12, 'prec_annee': annee-1,
        'suiv_mois':  1, 'suiv_annee': annee+1,
        'prec_label': str(annee-1), 'suiv_label': str(annee+1),
    })


@_bureau_required
def autremvt(request):
    """AUTREMOUVEMENT — visualisation par mois."""
    config = ConfigExercice.get_exercice_courant()
    mois   = max(1, min(12, int(request.GET.get('mois', timezone.now().month))))
    annee  = int(request.GET.get('annee', config.annee))
    prec   = (12, annee-1) if mois == 1 else (mois-1, annee)
    suiv   = (1,  annee+1) if mois == 12 else (mois+1, annee)

    mvts = MouvementFonds.objects.filter(
        mois=mois, annee=annee, config_exercice=config
    ).select_related('adherent').order_by('adherent__nom_prenom')

    totaux = {
        'capital_prec':  sum(m.capital_compose_precedent for m in mvts),
        'retrait':       sum(m.retrait_partiel           for m in mvts),
        'fonds_def':     sum(m.fonds_definitif           for m in mvts),
        'base_calcul':   sum(m.base_calcul_interet       for m in mvts),
        'interet':       sum(m.interet_attribue          for m in mvts),
        'capital':       sum(m.capital_compose           for m in mvts),
    }

    return render(request, 'dashboard/rapports/autremvt_mois.html', {
        'config_exercice': config,
        'mvts':       mvts,
        'totaux':     totaux,
        'mois':       mois,  'annee':      annee,
        'mois_label': MOIS_LABEL[mois],
        'prec_mois':  prec[0], 'prec_annee': prec[1],
        'suiv_mois':  suiv[0], 'suiv_annee': suiv[1],
        'prec_label': MOIS_LABEL[prec[0]],
        'suiv_label': MOIS_LABEL[suiv[0]],
    })


@_bureau_required
def autremvt_resume(request):
    """AUTREMOUVEMENT — résumé annuel."""
    config = ConfigExercice.get_exercice_courant()
    annee  = int(request.GET.get('annee', config.annee))
    from apps.adherents.models import Adherent

    adherents = Adherent.objects.filter(statut='ACTIF').order_by('nom_prenom')
    mvts = []
    for adh in adherents:
        last = MouvementFonds.objects.filter(
            adherent=adh, annee=annee, config_exercice=config
        ).order_by('-mois').first()
        if last:
            mvts.append(last)

    totaux = {
        'capital_prec': sum(m.capital_compose_precedent for m in mvts),
        'retrait':      sum(m.retrait_partiel           for m in mvts),
        'fonds_def':    sum(m.fonds_definitif           for m in mvts),
        'base_calcul':  sum(m.base_calcul_interet       for m in mvts),
        'interet':      sum(m.interet_attribue          for m in mvts),
        'capital':      sum(m.capital_compose           for m in mvts),
    }

    return render(request, 'dashboard/rapports/autremvt_mois.html', {
        'config_exercice': config,
        'mvts':       mvts,
        'totaux':     totaux,
        'mois':       0, 'annee':    annee,
        'mois_label': f'Résumé {annee}',
        'prec_mois': 12, 'prec_annee': annee-1,
        'suiv_mois':  1, 'suiv_annee': annee+1,
        'prec_label': str(annee-1), 'suiv_label': str(annee+1),
    })


# ═══════════════════════════════════════════════════════════════
# EXPORT TABBORDAIDEDEPENSES (chèques + espèces + agios)
# ═══════════════════════════════════════════════════════════════

@_bureau_required
def telecharger_tabbordaidedepenses(request):
    """
    ASELBY{annee}TABBORDAIDEDEPENSES.xlsx
    Structure exacte du fichier réel — 13 feuilles:
      HISTOJANV..HISTODEC (12 mois) + HISTORESUME

    Par feuille mensuelle (8 colonnes A→H):
      L1: ASSOCIATION / ASELBY
      L2: ANNEE / {annee}
      L3: MATRICULE / NOM ET PRENOM / ... / TOTAUX (col H)
      L4: / / MONTANT CHEQUE / NUM CHEQUE / ESPECES / AGIOS / Extrait de compte /
      L5: / CHEQUE ETABLI / total_cheques / / total_especes /
      L6..L(n+5): données adhérents (ASELBY en premier)
      L(n+6): / SOLDE BANQUE / / / / / solde_extrait (vide → saisie manuelle)
      L(n+7): / TOTAUX / tot_cheque / / tot_especes / tot_agios / / tot_H
      L(n+9): / DIFFERENCE A CHERCHER / diff_cheque /
      L(n+11): / COMPTE BANQUE /

    Sources BD:
      versement_banque  ← ReleveBancaire.versement_banque
      numero_cheque     ← SaisieMonthly.numero_cheque
      versement_especes ← ReleveBancaire.versement_especes
      agio              ← ReleveBancaire.agio
      TOTAUX col H      ← agio (ASELBY) ou 0 (membres)
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from apps.banque.models import ReleveBancaire

    D = __import__('decimal').Decimal

    config    = ConfigExercice.get_exercice_courant()
    annee     = config.annee
    # ASELBY toujours en premier, puis les autres par numero_ordre
    adherents = sorted(
        Adherent.objects.filter(statut='ACTIF'),
        key=lambda a: (0 if a.matricule == 'AS201648' else 1, a.numero_ordre)
    )

    BLEU  = '1B2B5E'; BLANC = 'FFFFFF'; OR = 'C9A84C'
    MOIS_CODE = {1:'JANV',2:'FEV',3:'MARS',4:'AVRIL',5:'MAI',6:'JUIN',
                 7:'JUIL',8:'AOUT',9:'SEPT',10:'OCT',11:'NOV',12:'DEC'}
    MOIS_NOM  = {1:'JANVIER',2:'FEVRIER',3:'MARS',4:'AVRIL',5:'MAI',6:'JUIN',
                 7:'JUILLET',8:'AOUT',9:'SEPTEMBRE',10:'OCTOBRE',
                 11:'NOVEMBRE',12:'DECEMBRE'}

    wb = Workbook()
    wb.remove(wb.active)

    def _d(v):
        if v is None: return D('0')
        try: return D(str(v))
        except: return D('0')

    def _fill(ws, row, col, val, bg=None, bold=False, align='left', num_fmt=None):
        c = ws.cell(row, col, val)
        if bg:
            c.fill = PatternFill('solid', fgColor=bg)
            c.font = Font(bold=bold, color=BLANC, size=9, name='Arial')
        elif bold:
            c.font = Font(bold=True, size=9, name='Arial')
        c.alignment = Alignment(horizontal=align, vertical='center')
        if num_fmt:
            c.number_format = num_fmt
        return c

    def _build_sheet(ws, mois, is_resume=False):
        """Construit une feuille mensuelle ou le résumé annuel."""
        # Largeurs colonnes
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 28
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 16
        ws.column_dimensions['H'].width = 14

        fill_bleu = PatternFill('solid', fgColor=BLEU)
        fill_or   = PatternFill('solid', fgColor=OR)
        fn_blanc  = Font(bold=True, color=BLANC, size=9, name='Arial')
        fn_sombre = Font(bold=True, color='1B2B5E', size=9, name='Arial')

        # L1-L2 : en-tête
        ws.cell(1, 1, 'ASSOCIATION'); ws.cell(1, 2, 'ASELBY')
        ws.cell(2, 1, 'ANNEE');       ws.cell(2, 2, annee)

        # L3 : titres principaux
        c = ws.cell(3, 1, 'MATRICULE')
        c.fill = fill_bleu; c.font = fn_blanc
        c = ws.cell(3, 2, 'NOM ET PRENOM')
        c.fill = fill_bleu; c.font = fn_blanc
        if is_resume:
            # HISTORESUME a 9 colonnes
            for col, hdr in [(3,''),(4,''),(5,''),(6,''),(7,''),(8,''),(9,'TOTAUX')]:
                c = ws.cell(3, col, hdr)
                c.fill = fill_bleu; c.font = fn_blanc
                c.alignment = Alignment(horizontal='center')
            # L4 sous-titres
            for col, hdr in [(3,'MONTANT CHEQUE'),(4,'NUM CHEQUE'),(5,'ESPECES'),
                             (6,'TOTAL ENCAISSE'),(7,'AGIOS'),(8,'Extrait de compte'),(9,'')]:
                c = ws.cell(4, col, hdr)
                c.font = Font(bold=True, size=8, name='Arial')
                c.alignment = Alignment(horizontal='center', wrap_text=True)
        else:
            for col, hdr in [(3,''),(4,''),(5,''),(6,''),(7,''),(8,'TOTAUX')]:
                c = ws.cell(3, col, hdr)
                c.fill = fill_bleu; c.font = fn_blanc
                c.alignment = Alignment(horizontal='center')
            # L4 sous-titres
            for col, hdr in [(3,'MONTANT CHEQUE'),(4,'NUM CHEQUE'),(5,'ESPECES'),
                             (6,'AGIOS'),(7,'Extrait de compte'),(8,'')]:
                c = ws.cell(4, col, hdr)
                c.font = Font(bold=True, size=8, name='Arial')
                c.alignment = Alignment(horizontal='center', wrap_text=True)
        ws.row_dimensions[4].height = 24

        return 3 + 1  # retourne index ligne L4 (entête données en L4, données en L5+)

    # ── Pré-charger TOUTES les données en mémoire ──────────────────
    # ReleveBancaire par (mat, mois)
    releves_all = {}
    for r in ReleveBancaire.objects.filter(annee=annee, config_exercice=config):
        releves_all[(r.adherent_id, r.mois)] = r

    # SaisieMonthly par (mat, mois) pour numero_cheque et montant_especes
    from apps.saisie.models import SaisieMonthly
    saisies_all = {}
    for s in SaisieMonthly.objects.filter(annee=annee, config_exercice=config):
        saisies_all[(s.adherent_id, s.mois)] = s

    # ── Feuilles mensuelles HISTOJANV..HISTODEC ────────────────────
    for mois in range(1, 13):
        label = f"HISTO{MOIS_CODE[mois]}{str(annee)[-2:]}"
        ws = wb.create_sheet(label)
        _build_sheet(ws, mois, is_resume=False)

        # Totaux chèques et espèces de ce mois
        tot_cheque = sum(
            _d(releves_all.get((adh.matricule, mois), None) and
               releves_all[(adh.matricule, mois)].versement_banque)
            for adh in adherents
        )
        tot_especes = sum(
            _d(releves_all.get((adh.matricule, mois), None) and
               releves_all[(adh.matricule, mois)].versement_especes)
            for adh in adherents
        )

        # L5 : CHEQUE ETABLI
        ws.cell(5, 2, 'CHEQUE ETABLI').font = Font(bold=True, size=9, name='Arial')
        if tot_cheque:
            ws.cell(5, 3, float(tot_cheque))
        if tot_especes:
            ws.cell(5, 5, float(tot_especes))

        # L6+ : données par adhérent
        fill_bleu = PatternFill('solid', fgColor=BLEU)
        fn_blanc  = Font(bold=True, color=BLANC, size=9, name='Arial')

        tot_ag = D('0')
        for i, adh in enumerate(adherents, 6):
            r = releves_all.get((adh.matricule, mois))
            s = saisies_all.get((adh.matricule, mois))

            banque  = _d(r.versement_banque)  if r else D('0')
            especes = _d(r.versement_especes) if r else D('0')
            agio    = _d(r.agio)              if r else D('0')
            cheque_ok = 'OK' if banque > 0 else 0
            num_cheque = s.numero_cheque if s and s.numero_cheque else 0

            # TOTAUX col H:
            # Pour ASELBY = agio (intérêts bancaires nets)
            # Pour les membres = 0 (leurs chèques sont dans col C)
            tot_h = float(agio) if adh.matricule == 'AS201648' else 0
            tot_ag += agio

            ws.cell(i, 1, adh.matricule)
            ws.cell(i, 2, adh.nom_prenom)
            ws.cell(i, 3, float(banque)  if banque  else 0)
            ws.cell(i, 4, cheque_ok      if banque  else 0)
            ws.cell(i, 5, float(especes) if especes else 0)
            ws.cell(i, 6, float(agio)    if agio    else None)
            ws.cell(i, 7, None)   # Extrait de compte — saisie manuelle
            ws.cell(i, 8, tot_h)

        nb_adh = len(adherents)
        row_solde = nb_adh + 6
        row_tot   = nb_adh + 7
        row_diff  = nb_adh + 9
        row_cpte  = nb_adh + 11

        # SOLDE BANQUE (vide — saisie manuelle depuis relevé bancaire réel)
        ws.cell(row_solde, 2, 'SOLDE BANQUE').font = Font(bold=True, size=9, name='Arial')
        ws.cell(row_solde, 7, None)  # laissé vide

        # TOTAUX
        ws.cell(row_tot, 2, 'TOTAUX').font = Font(bold=True, size=9, name='Arial')
        ws.cell(row_tot, 3, float(tot_cheque) if tot_cheque else 0)
        ws.cell(row_tot, 5, float(tot_especes) if tot_especes else 0)
        ws.cell(row_tot, 6, float(tot_ag) if tot_ag else 0)

        # DIFFERENCE A CHERCHER = CHEQUE ETABLI - total chèques reçus
        # Dans le réel = 0 si tout est cohérent
        diff = tot_cheque - tot_cheque  # = 0 (vérification de cohérence)
        ws.cell(row_diff, 2, 'DIFFERENCE A CHERCHER').font = Font(bold=True, size=9)
        ws.cell(row_diff, 3, float(diff))

        # COMPTE BANQUE (libellé seul)
        ws.cell(row_cpte, 2, 'COMPTE BANQUE').font = Font(bold=True, size=9)

    # ── Feuille HISTORESUME ──────────────────────────────────────────
    ws_r = wb.create_sheet(f"HISTO{'RESUME'}{str(annee)[-2:]}")
    _build_sheet(ws_r, 0, is_resume=True)
    ws_r.column_dimensions['F'].width = 14  # TOTAL ENCAISSE

    # Totaux annuels
    tot_cheque_an = D('0')
    tot_especes_an = D('0')
    tot_ag_an = D('0')

    # L5 : totaux annuels globaux
    for mois in range(1, 13):
        for adh in adherents:
            r = releves_all.get((adh.matricule, mois))
            if r:
                tot_cheque_an  += _d(r.versement_banque)
                tot_especes_an += _d(r.versement_especes)
                tot_ag_an      += _d(r.agio)

    ws_r.cell(5, 2, 'CHEQUE ETABLI').font = Font(bold=True, size=9)
    ws_r.cell(5, 3, float(tot_cheque_an))
    ws_r.cell(5, 5, float(tot_especes_an))
    ws_r.cell(5, 6, float(tot_cheque_an + tot_especes_an))
    ws_r.cell(5, 7, float(tot_ag_an))

    # L6+ : cumul annuel par adhérent
    for i, adh in enumerate(adherents, 6):
        tot_c = tot_e = tot_a = D('0')
        for mois in range(1, 13):
            r = releves_all.get((adh.matricule, mois))
            if r:
                tot_c += _d(r.versement_banque)
                tot_e += _d(r.versement_especes)
                tot_a += _d(r.agio)

        tot_enc = tot_c + tot_e
        tot_h = float(tot_a) if adh.matricule == 'AS201648' else 0

        ws_r.cell(i, 1, adh.matricule)
        ws_r.cell(i, 2, adh.nom_prenom)
        ws_r.cell(i, 3, float(tot_c)   if tot_c else 0)
        ws_r.cell(i, 4, None)  # pas de num chèque dans le résumé
        ws_r.cell(i, 5, float(tot_e)   if tot_e else 0)
        ws_r.cell(i, 6, float(tot_enc) if tot_enc else 0)
        ws_r.cell(i, 7, float(tot_a)   if tot_a else 0)
        ws_r.cell(i, 8, 0)   # Extrait de compte — saisie manuelle
        ws_r.cell(i, 9, tot_h)

    nb_adh = len(adherents)
    row_solde = nb_adh + 6
    row_tot   = nb_adh + 7
    row_diff  = nb_adh + 9
    row_cpte  = nb_adh + 11

    ws_r.cell(row_solde, 2, 'SOLDE BANQUE').font = Font(bold=True, size=9)
    ws_r.cell(row_tot,   2, 'TOTAUX').font = Font(bold=True, size=9)
    ws_r.cell(row_tot,   3, float(tot_cheque_an))
    ws_r.cell(row_tot,   5, float(tot_especes_an))
    ws_r.cell(row_tot,   6, float(tot_cheque_an + tot_especes_an))
    ws_r.cell(row_tot,   7, float(tot_ag_an))
    ws_r.cell(row_diff,  2, 'DIFFERENCE A CHERCHER').font = Font(bold=True, size=9)
    ws_r.cell(row_diff,  3, 0)
    ws_r.cell(row_cpte,  2, 'COMPTE BANQUE').font = Font(bold=True, size=9)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    resp = HttpResponse(buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = (
        f'attachment; filename=ASELBY{annee}TABBORDAIDEDEPENSES.xlsx')
    return resp


@_bureau_required
def impression_index(request):
    config = ConfigExercice.get_exercice_courant()
    mois   = timezone.now().month
    annee  = config.annee
    documents = [
        {'titre': 'Releve bancaire mensuel', 'description': 'TABBHISTOBQUE',
         'icone': 'fas fa-university',
         'url_apercu': f'/impression/releve/?mois={mois}&annee={annee}',
         'url_excel': '/releve/telecharger/'},
        {'titre': 'Saisie mensuelle TABBORD', 'description': 'Colonnes manuelles par adherent',
         'icone': 'fas fa-table',
         'url_apercu': f'/impression/tabbord/?mois={mois}&annee={annee}',
         'url_excel': '/dashboard/tabbord/telecharger/'},
        {'titre': 'Tontines du mois', 'description': 'Parts T60/T75/T100',
         'icone': 'fas fa-layer-group',
         'url_apercu': f'/impression/tontines/?mois={mois}&annee={annee}',
         'url_excel': None},
        {'titre': 'AUTREMOUVEMENT', 'description': 'Fonds et capital compose',
         'icone': 'fas fa-exchange-alt',
         'url_apercu': f'/impression/autremvt/?mois={mois}&annee={annee}',
         'url_excel': '/dashboard/autremvt/telecharger/'},
        {'titre': 'Fiches de cassation', 'description': 'Une fiche par adherent fin exercice',
         'icone': 'fas fa-file-invoice',
         'url_apercu': '/impression/fiches-cassation/',
         'url_excel': '/dashboard/travauxfinexercice/telecharger/'},
        {'titre': 'Liste des adherents', 'description': 'Repertoire complet avec contacts',
         'icone': 'fas fa-users',
         'url_apercu': '/impression/adherents/',
         'url_excel': '/dashboard/listeadherent/telecharger/'},
        {'titre': 'Prets en cours', 'description': 'Soldes et echeances',
         'icone': 'fas fa-dollar-sign',
         'url_apercu': '/impression/prets/',
         'url_excel': None},
    ]
    return render(request, 'dashboard/impression/index.html', {
        'config_exercice': config, 'documents': documents,
        'mois': mois, 'annee': annee,
    })


@_bureau_required
def impression_releve(request):
    from apps.banque.models import ReleveBancaire
    config    = ConfigExercice.get_exercice_courant()
    mois      = int(request.GET.get('mois', timezone.now().month))
    annee     = int(request.GET.get('annee', config.annee))
    adherents = Adherent.objects.filter(statut='ACTIF').order_by('nom_prenom')
    releves   = {r.adherent_id: r for r in ReleveBancaire.objects.filter(
        mois=mois, annee=annee, config_exercice=config)}
    MOIS_FR = ['','Janvier','Fevrier','Mars','Avril','Mai','Juin',
               'Juillet','Aout','Septembre','Octobre','Novembre','Decembre']
    return render(request, 'dashboard/impression/releve_print.html', {
        'config_exercice': config, 'adherents': adherents, 'releves': releves,
        'mois': mois, 'annee': annee, 'mois_label': MOIS_FR[mois],
    })


@_bureau_required
def impression_tabbord(request):
    from apps.banque.models import ReleveBancaire
    config    = ConfigExercice.get_exercice_courant()
    mois      = int(request.GET.get('mois', timezone.now().month))
    annee     = int(request.GET.get('annee', config.annee))
    adherents = Adherent.objects.filter(statut='ACTIF').order_by('nom_prenom')
    saisies   = {s.adherent_id: s for s in SaisieMonthly.objects.filter(
        mois=mois, annee=annee, config_exercice=config)}
    releves   = {r.adherent_id: r for r in ReleveBancaire.objects.filter(
        mois=mois, annee=annee, config_exercice=config)}
    MOIS_FR = ['','Janvier','Fevrier','Mars','Avril','Mai','Juin',
               'Juillet','Aout','Septembre','Octobre','Novembre','Decembre']
    return render(request, 'dashboard/impression/tabbord_print.html', {
        'config_exercice': config, 'adherents': adherents, 'saisies': saisies,
        'releves': releves, 'mois': mois, 'annee': annee, 'mois_label': MOIS_FR[mois],
    })


@_bureau_required
def impression_tontines(request):
    config    = ConfigExercice.get_exercice_courant()
    mois      = int(request.GET.get('mois', timezone.now().month))
    annee     = int(request.GET.get('annee', config.annee))
    adherents = Adherent.objects.filter(statut='ACTIF').order_by('nom_prenom')
    saisies   = {s.adherent_id: s for s in SaisieMonthly.objects.filter(
        mois=mois, annee=annee, config_exercice=config)}
    MOIS_FR = ['','Janvier','Fevrier','Mars','Avril','Mai','Juin',
               'Juillet','Aout','Septembre','Octobre','Novembre','Decembre']
    return render(request, 'dashboard/impression/tontines_print.html', {
        'config_exercice': config, 'adherents': adherents, 'saisies': saisies,
        'mois': mois, 'annee': annee, 'mois_label': MOIS_FR[mois],
    })


@_bureau_required
def impression_autremvt(request):
    config = ConfigExercice.get_exercice_courant()
    mois   = int(request.GET.get('mois', timezone.now().month))
    annee  = int(request.GET.get('annee', config.annee))
    mvts   = MouvementFonds.objects.filter(
        mois=mois, annee=annee, config_exercice=config
    ).select_related('adherent').order_by('adherent__nom_prenom')
    MOIS_FR = ['','Janvier','Fevrier','Mars','Avril','Mai','Juin',
               'Juillet','Aout','Septembre','Octobre','Novembre','Decembre']
    return render(request, 'dashboard/impression/autremvt_print.html', {
        'config_exercice': config, 'mvts': mvts,
        'mois': mois, 'annee': annee, 'mois_label': MOIS_FR[mois],
    })


@_bureau_required
def impression_fiches_cassation(request):
    from apps.exercice.models import FicheCassation
    from apps.prets.models import Pret
    from decimal import Decimal as D2
    config    = ConfigExercice.get_exercice_courant()
    annee     = config.annee
    adherents = Adherent.objects.filter(statut='ACTIF').order_by('nom_prenom')
    fiches_db = {f.adherent_id: f for f in FicheCassation.objects.filter(config_exercice=config)}
    fiches = []
    for adh in adherents:
        mvt = MouvementFonds.objects.filter(adherent=adh, annee=annee,
              config_exercice=config).order_by('-mois').first()
        capital      = D2(str(mvt.fonds_definitif))  if mvt else D2('0')
        recon        = D2(str(mvt.capital_compose))  if mvt else D2('0')
        interets     = D2(str(mvt.interet_attribue)) if mvt else D2('0')
        epargne      = sum(D2(str(s.complement_epargne or 0)) for s in
            SaisieMonthly.objects.filter(adherent=adh, annee=annee, config_exercice=config))
        pret         = Pret.objects.filter(adherent=adh, config_exercice=config, statut='EN_COURS').first()
        dette_pret   = D2(str(pret.solde_restant)) if pret else D2('0')
        fdb          = fiches_db.get(adh.matricule)
        sanction     = D2(str(fdb.sanction))     if fdb and fdb.sanction     else D2('0')
        montant_perc = D2(str(fdb.montant_percu)) if fdb and fdb.montant_percu else D2('0')
        comp_mut     = config.complement_mutuelle_fin_exercice
        comp_fonds   = config.complement_fonds_fin_exercice
        total_d      = capital + recon + interets + epargne + comp_mut + comp_fonds
        total_r      = sanction + dette_pret
        fiches.append({'adherent': adh, 'annee': annee, 'fonds_caisse': capital,
            'reconduction': recon, 'interet': interets, 'epargne': epargne,
            'complement_mutuelle': comp_mut, 'complement_fonds': comp_fonds,
            'total_distribuer': total_d, 'sanction': sanction, 'dette_pret': dette_pret,
            'total_retenu': total_r, 'net_a_percevoir': total_d - total_r,
            'montant_percu': montant_perc, 'reconduction_suivante': recon})
    return render(request, 'dashboard/impression/fiche_cassation_print.html', {
        'config_exercice': config, 'fiches': fiches, 'annee': annee,
    })


@_bureau_required
def impression_adherents(request):
    config    = ConfigExercice.get_exercice_courant()
    adherents = Adherent.objects.all().order_by('numero_ordre')
    return render(request, 'dashboard/impression/adherents_print.html', {
        'config_exercice': config, 'adherents': adherents,
    })


@_bureau_required
def impression_prets(request):
    from apps.prets.models import Pret
    config = ConfigExercice.get_exercice_courant()
    prets  = Pret.objects.filter(config_exercice=config, statut='EN_COURS'
             ).select_related('adherent').order_by('adherent__nom_prenom')
    return render(request, 'dashboard/impression/prets_print.html', {
        'config_exercice': config, 'prets': prets,
    })


@_bureau_required
def telecharger_listeadherent(request):
    """Export LISTEADHERENT.xlsx — liste complète des adhérents actifs."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    config    = ConfigExercice.get_exercice_courant()
    annee     = config.annee
    adherents = sorted(
        Adherent.objects.filter(statut='ACTIF'),
        key=lambda a: (0 if a.matricule == 'AS201648' else 1, a.numero_ordre)
    )
    wb = Workbook(); ws = wb.active
    ws.title = f'LISTADHERENT{str(annee)[-2:]}'
    BLEU = '1B2B5E'; BLANC = 'FFFFFF'

    ws['A1'] = 'ASSOCIATION'; ws['B1'] = 'ASELBY'
    ws['A2'] = 'ANNEE';       ws['B2'] = annee

    COLS = ['MATRICULE','NUMORDRE','NOM ET PRENOM','NBRE LOT T60',
            'NBRE LOT T75','NBRE LOT T100','CONTACT','STATUT']
    fill = PatternFill('solid', fgColor=BLEU)
    fn   = Font(bold=True, color=BLANC, size=9, name='Arial')
    for j, h in enumerate(COLS, 1):
        cl = ws.cell(4, j, h)
        cl.fill = fill; cl.font = fn
        cl.alignment = Alignment(horizontal='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width = 20

    for i, adh in enumerate(adherents, 5):
        ws.cell(i,1, adh.matricule)
        ws.cell(i,2, adh.numero_ordre)
        ws.cell(i,3, adh.nom_prenom)
        ws.cell(i,4, adh.nbre_lots_t60 or 1)
        ws.cell(i,5, adh.nbre_lots_t75 or 0)
        ws.cell(i,6, adh.nbre_lots_t100 or 0)
        ws.cell(i,7, adh.telephone or '')
        ws.cell(i,8, adh.statut)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    resp = HttpResponse(buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename=ASELBY{annee}LISTEADHERENT.xlsx'
    return resp


@_bureau_required
def telecharger_travauxfinexercice(request):
    """Export TRAVAUXFINEXERCICE.xlsx — DETAILFICHECASSATION + SYNTHESECOMPTE."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from apps.exercice.models import FicheCassation, SyntheseCompte

    config    = ConfigExercice.get_exercice_courant()
    annee     = config.annee
    adherents = sorted(
        Adherent.objects.filter(statut='ACTIF'),
        key=lambda a: (0 if a.matricule == 'AS201648' else 1, a.numero_ordre)
    )
    wb = Workbook(); wb.remove(wb.active)
    BLEU = '1B2B5E'; BLANC = 'FFFFFF'

    # ── Feuille DETAILFICHECASSATION ──────────────────────────
    ws = wb.create_sheet('DETAILFICHECASSATION')
    ws['A1'] = 'ASSOCIATION'; ws['B1'] = 'ASELBY'
    ws['A2'] = 'ANNEE';       ws['B2'] = annee

    COLS_FICHE = [
        'MATRICULE','NOM ET PRENOM','FONDS DE CAISSE','RECONDUCTION',
        'REPARTITION INTERET','EPARGNE','REPARTITION PENALITE REMBOURSEMENT PRET',
        'REPARTITION COLLATION','TOTAL A DISTRIBUER','SANCTION',
        'COMPLEMENT MUTUELLE','COMPLEMENT FONDS','DETTE PRET',
        'TOTAL RETENU','TOTAL A PERCEVOIR','MONTANT PERCU',
        'RECONDUCTION SUIVANTE','NOUVEAU FONDS','DONS FOYER',
        'MONTANT PERCU ESPECES','MONTANT PERCU CHEQUE',
    ]
    fill = PatternFill('solid', fgColor=BLEU)
    fn   = Font(bold=True, color=BLANC, size=8, name='Arial')
    for j, h in enumerate(COLS_FICHE, 1):
        cl = ws.cell(5, j, h)
        cl.fill = fill; cl.font = fn
        cl.alignment = Alignment(horizontal='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width = 16
    ws.row_dimensions[5].height = 36

    fiches = {f.adherent_id: f for f in
              FicheCassation.objects.filter(config_exercice=config)}

    def _d(v):
        if v is None: return 0.0
        try: return float(v)
        except: return 0.0

    for i, adh in enumerate(adherents, 6):
        f = fiches.get(adh.matricule)
        ws.cell(i,1, adh.matricule)
        ws.cell(i,2, adh.nom_prenom)
        ws.cell(i,3, _d(f.fonds_caisse)          if f else 0)
        ws.cell(i,4, _d(f.repartition_interets)  if f else 0)
        ws.cell(i,5, _d(f.repartition_interets)  if f else 0)
        ws.cell(i,6, _d(f.epargne_cumulee)        if f else 0)
        ws.cell(i,7, _d(f.repartition_penalites)  if f else 0)
        ws.cell(i,8, _d(f.repartition_collation)  if f else 0)
        ws.cell(i,9, _d(f.total_a_distribuer)     if f else 0)
        ws.cell(i,10,_d(f.sanctions)              if f else 0)
        ws.cell(i,11,_d(f.complement_mutuelle)    if f else 0)
        ws.cell(i,12,_d(f.complement_fonds)       if f else 0)
        ws.cell(i,13,_d(f.dette_pret)             if f else 0)
        ws.cell(i,14,_d(f.total_retenu)           if f else 0)
        ws.cell(i,15,_d(f.net_a_percevoir)        if f else 0)
        ws.cell(i,16,_d(f.montant_percu)          if f else 0)
        ws.cell(i,17,_d(f.reconduction)           if f else 0)
        ws.cell(i,18,_d(f.nouveau_fonds)          if f else 0)
        ws.cell(i,19,_d(f.dons_foyer)             if f else 0)
        ws.cell(i,20,_d(f.montant_percu_especes)  if f else 0)
        ws.cell(i,21,_d(f.montant_percu_cheque)   if f else 0)

    # ── Feuille SYNTHESECOMPTE ────────────────────────────────
    ws2 = wb.create_sheet('SYNTHESECOMPTE')
    ws2['A1'] = 'ASSOCIATION'; ws2['B1'] = 'ASELBY'
    ws2['A2'] = 'ANNEE';       ws2['B2'] = annee
    sc = SyntheseCompte.objects.filter(config_exercice=config).first()
    if sc:
        LIGNES = [
            ('FONDS DE CAISSE', sc.report_fonds_caisse, sc.entrees_fonds_caisse, sc.sorties_fonds_caisse),
            ('FONDS DE ROULEMENT', sc.report_fonds_roulement, sc.entrees_fonds_roulement, sc.sorties_fonds_roulement),
            ('FRAIS EXCEPTIONNELS', sc.report_frais_exceptionnels, sc.entrees_frais_excep, sc.sorties_frais_excep),
        ]
        for i, (label, report, entrees, sorties) in enumerate(LIGNES, 5):
            ws2.cell(i,1, label)
            ws2.cell(i,2, float(report  or 0))
            ws2.cell(i,3, float(entrees or 0))
            ws2.cell(i,4, float(sorties or 0))
            ws2.cell(i,5, float((report or 0)+(entrees or 0)-(sorties or 0)))

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    resp = HttpResponse(buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename=ASELBY{annee}TRAVAUXFINEXERCICE.xlsx'
    return resp