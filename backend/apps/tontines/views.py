"""
apps/tontines/views.py
Tontines — lecture seule depuis SaisieMonthly.
Plus de saisie séparée : tout passe par saisie_mensuelle.
Vues : tableau récapitulatif + exports Excel.
"""
from django.shortcuts import render
from django.utils import timezone
from decimal import Decimal
from apps.core.mixins import bureau_required
from apps.parametrage.models import ConfigExercice
from apps.adherents.models import Adherent
from apps.saisie.models import SaisieMonthly

MOIS_FR   = ['','Janvier','Février','Mars','Avril','Mai','Juin',
             'Juillet','Août','Septembre','Octobre','Novembre','Décembre']
MOIS_CODE = ['','JANV','FEV','MARS','AVRIL','MAI','JUIN',
             'JUIL','AOUT','SEPT','OCT','NOV','DEC']
D = Decimal


@bureau_required
def tableau_tontines(request):
    """
    Récapitulatif mensuel des tontines T60 / T75 / T100.
    Données lues depuis SaisieMonthly.
    """
    config = ConfigExercice.get_exercice_courant()
    mois   = int(request.GET.get('mois',  timezone.now().month))
    annee  = int(request.GET.get('annee', config.annee))

    saisies = SaisieMonthly.objects.filter(
        mois=mois, annee=annee, config_exercice=config
    ).select_related('adherent').order_by('adherent__numero_ordre')

    # Totaux par niveau
    tot_t60  = sum(s.tontine_t60  for s in saisies)
    tot_t75  = sum(s.tontine_t75  for s in saisies)
    tot_t100 = sum(s.tontine_t100 for s in saisies)
    tot_mois = tot_t60 + tot_t75 + tot_t100

    # Lots attribués ce mois
    lots_t60  = [s for s in saisies if s.achat_lot_t60  > 0]
    lots_t75  = [s for s in saisies if s.achat_lot_t75  > 0]
    lots_t100 = [s for s in saisies if s.achat_lot_t100 > 0]

    # Ventes petit lot
    ventes_t60  = [s for s in saisies if s.vente_petit_lot_t60  > 0]
    ventes_t75  = [s for s in saisies if s.vente_petit_lot_t75  > 0]
    ventes_t100 = [s for s in saisies if s.vente_petit_lot_t100 > 0]

    prec = (12, annee-1) if mois == 1 else (mois-1, annee)
    suiv = (1,  annee+1) if mois == 12 else (mois+1, annee)

    return render(request, 'dashboard/tontines/tableau.html', {
        'config_exercice': config,
        'saisies':  saisies,
        'mois':     mois,
        'annee':    annee,
        'mois_label': MOIS_FR[mois],
        'mois_fr':  MOIS_FR,
        'prec_mois': prec[0], 'prec_annee': prec[1],
        'suiv_mois': suiv[0], 'suiv_annee': suiv[1],
        'tot_t60':  tot_t60,  'tot_t75':  tot_t75,
        'tot_t100': tot_t100, 'tot_mois': tot_mois,
        'lots_t60':  lots_t60,  'lots_t75':  lots_t75,  'lots_t100':  lots_t100,
        'ventes_t60': ventes_t60, 'ventes_t75': ventes_t75, 'ventes_t100': ventes_t100,
        'nb_total':  Adherent.objects.filter(statut='ACTIF').count(),
        'nb_saisis': saisies.count(),
    })


@bureau_required
def telecharger_tontine(request, niveau):
    """
    Export Excel TONTINE{niveau}.xlsx — structure exacte des fichiers réels.

    T60 (13 cols, 12 feuilles, TOUS les membres):
      L1-L3: en-tête ASSOCIATION/ANNEE/TONTINE
      L5: entêtes 13 colonnes
      L6+: tous les membres (ASELBY en premier, puis actifs par numero_ordre)
      PAS de feuille RESUME

    T75/T100 (16 cols, 13 feuilles avec TONTRESUME, seulement membres avec parts):
      L1-L4: en-tête + TAUX/NBRE PART/DATE
      L5: entêtes 16 colonnes
      L6+: seulement membres ayant nbre_lots > 0 ce mois
      Lignes synthèse en bas (TOTAL PART, NBRE ADHERENT, détail ESPECES/BANQUE/TOTAL)
      + feuille TONTRESUME = cumul annuel

    Sources BD: SaisieMonthly
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse, Http404

    if niveau not in ('T60', 'T75', 'T100'):
        raise Http404

    config    = ConfigExercice.get_exercice_courant()
    annee     = config.annee
    # ASELBY toujours en premier
    adherents_all = sorted(
        Adherent.objects.filter(statut='ACTIF'),
        key=lambda a: (0 if a.matricule == 'AS201648' else 1, a.numero_ordre)
    )

    BLEU  = '1B2B5E'; BLANC = 'FFFFFF'
    MOIS_CODE = {1:'JANV',2:'FEV',3:'MARS',4:'AVRIL',5:'MAI',6:'JUIN',
                 7:'JUIL',8:'AOUT',9:'SEPT',10:'OCT',11:'NOV',12:'DEC'}
    MOIS_NOM  = {1:'JANVIER',2:'FEVRIER',3:'MARS',4:'AVRIL',5:'MAI',6:'JUIN',
                 7:'JUILLET',8:'AOUT',9:'SEPTEMBRE',10:'OCTOBRE',
                 11:'NOVEMBRE',12:'DECEMBRE'}

    # Config selon niveau
    if niveau == 'T60':
        taux     = D(str(config.versement_t35))  # 60 000
        fld_nbre = 'nbre_lots_t60'
        fld_ton  = 'tontine_t60'
        fld_vente = 'vente_petit_lot_t60'
        fld_int   = 'interet_petit_lot_t60'
        fld_achat = 'achat_lot_t60'
        label_tontine = f'TONTINE {int(taux):,}'.replace(',', ' ')
    elif niveau == 'T75':
        taux     = D(str(config.taux_t75))  # 75 000
        fld_nbre = 'nbre_lots_t75'
        fld_ton  = 'tontine_t75'
        fld_vente = 'vente_petit_lot_t75'
        fld_int   = 'interet_petit_lot_t75'
        fld_achat = 'achat_lot_t75'
        label_tontine = f'TONTINE {int(taux):,}'.replace(',', ' ')
    else:  # T100
        taux     = D(str(config.taux_t100))  # 100 000
        fld_nbre = 'nbre_lots_t100'
        fld_ton  = 'tontine_t100'
        fld_vente = 'vente_petit_lot_t100'
        fld_int   = 'interet_petit_lot_t100'
        fld_achat = 'achat_lot_t100'
        label_tontine = f'TONTINE {int(taux):,}'.replace(',', ' ')

    wb = Workbook()
    wb.remove(wb.active)

    # Pré-charger toutes les saisies
    saisies_all = {}
    for s in SaisieMonthly.objects.filter(annee=annee, config_exercice=config):
        saisies_all[(s.adherent_id, s.mois)] = s

    def _d(v):
        if v is None: return D('0')
        try: return D(str(v))
        except: return D('0')

    def _hdr_cell(ws, row, col, val, bg=BLEU):
        c = ws.cell(row, col, val)
        c.fill = PatternFill('solid', fgColor=bg)
        c.font = Font(bold=True, color=BLANC, size=8, name='Arial')
        c.alignment = Alignment(horizontal='center', wrap_text=True)
        return c

    def _set_col_widths(ws, widths):
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

    # ── Cumul annuel pour TONTRESUME ──────────────────────────────
    cumul_annuel = {}  # mat → {nbre, tontine, vente, interet, achat}

    # ── FEUILLES MENSUELLES ───────────────────────────────────────
    for mois in range(1, 13):
        label = f"TONT{MOIS_CODE[mois]}{str(annee)[-2:]}"
        ws    = wb.create_sheet(label)

        # Saisies de ce mois
        saisies_mois = {
            mat: saisies_all[(mat, mois)]
            for mat, m in saisies_all
            if m == mois and (mat, mois) in saisies_all
        }

        if niveau == 'T60':
            # ── T60: TOUS les membres, 13 colonnes ──────────────
            _set_col_widths(ws, {'A':12,'B':5,'C':28,'D':14,'E':14,
                                  'F':10,'G':14,'H':14,'I':14,'J':14,
                                  'K':12,'L':14,'M':14})
            ws.cell(1,1,'ASSOCIATION'); ws.cell(1,2,'ASELBY')
            ws.cell(2,1,'ANNEE');       ws.cell(2,2, annee)
            ws.cell(3,1,'TONTINE');     ws.cell(3,2, niveau)

            ws.row_dimensions[5].height = 32
            COLS_T60 = ['MATRICULE','NBREPART','NOM ET PRENOM','MONTANT TONTINE',
                        'MONTANT VERSE','PENALITE VERSEMENT ESPECES','MODE VERSEMENT',
                        'VENTE PETIT LOT','INTERET PETIT LOT','MODE PAIEMENT LOT',
                        'NUMERO CHEQUE','MONTANT LOT PRINCIPAL','INTERET LOT PRINCIPAL']
            for j, h in enumerate(COLS_T60, 1):
                _hdr_cell(ws, 5, j, h)

            tot_nbre=0; tot_mont=D('0'); tot_verse=D('0'); tot_pen=D('0')
            tot_vente=D('0'); tot_int=D('0'); tot_achat=D('0')

            for i, adh in enumerate(adherents_all, 6):
                s = saisies_mois.get(adh.matricule)
                nbre   = int(_d(getattr(s, fld_nbre,  0))) if s else 0
                tontine= _d(getattr(s, fld_ton,   D('0'))) if s else D('0')
                vente  = _d(getattr(s, fld_vente, D('0'))) if s else D('0')
                interet= _d(getattr(s, fld_int,   D('0'))) if s else D('0')
                achat  = _d(getattr(s, fld_achat, D('0'))) if s else D('0')
                pen    = s.penalite_versement_especes if s else D('0')
                mode   = s.mode_paiement_tontine if s else None
                mode_lot = s.mode_paiement_lot   if s else None
                num_chq  = s.numero_cheque       if s else None

                # Cumuler pour RESUME annuel
                c = cumul_annuel.setdefault(adh.matricule, {
                    'adh':adh,'nbre':0,'tontine':D('0'),
                    'vente':D('0'),'interet':D('0'),'achat':D('0'),'pen':D('0')
                })
                c['nbre'] += nbre; c['tontine'] += tontine
                c['vente'] += vente; c['interet'] += interet
                c['achat'] += achat; c['pen'] += _d(pen)

                tot_nbre += nbre; tot_mont += tontine; tot_verse += tontine
                tot_pen  += _d(pen); tot_vente += vente
                tot_int  += interet; tot_achat += achat

                ws.cell(i,1, adh.matricule)
                ws.cell(i,2, nbre)
                ws.cell(i,3, adh.nom_prenom)
                ws.cell(i,4, float(nbre * taux))
                ws.cell(i,5, float(tontine))
                ws.cell(i,6, float(_d(pen)))
                ws.cell(i,7, mode   or None)
                ws.cell(i,8, float(vente))
                ws.cell(i,9, float(interet))
                ws.cell(i,10, mode_lot or None)
                ws.cell(i,11, num_chq  or None)
                ws.cell(i,12, float(achat))
                ws.cell(i,13, 0)  # intérêt lot principal

        else:
            # ── T75/T100: seulement membres avec parts, 16 colonnes ──
            _set_col_widths(ws, {'A':12,'B':5,'C':28,'D':14,'E':14,
                                  'F':14,'G':14,'H':14,'I':12,'J':14,
                                  'K':14,'L':14,'M':14,'N':12,'O':14,'P':14})
            # L1-L4
            ws.cell(1,1,'ASSOCIATION'); ws.cell(1,2,'ASELBY')
            ws.cell(2,1,'ANNEE');       ws.cell(2,2, annee)
            ws.cell(2,3,'TONIINE');     ws.cell(2,4,'NBRE PART')

            # Compter parts totales et membres ce mois
            membres_mois = [
                (adh, saisies_mois[adh.matricule])
                for adh in adherents_all
                if adh.matricule in saisies_mois
                and int(_d(getattr(saisies_mois[adh.matricule], fld_nbre, 0))) > 0
            ]
            total_parts = sum(int(_d(getattr(s, fld_nbre, 0))) for _, s in membres_mois)
            montant_lot = total_parts * taux  # montant total lot = parts × taux

            ws.cell(3,1,'TAUX TONTINE '); ws.cell(3,2, float(taux))
            ws.cell(3,3, float(montant_lot)); ws.cell(3,4, total_parts)
            from django.utils import timezone
            ws.cell(4,3, timezone.now().replace(day=13, month=mois,
                         year=annee).strftime('%Y-%m-%d'))

            ws.row_dimensions[5].height = 32
            COLS_T75100 = [
                'MATRICULE','NBREPART','NOM ET PRENOM','MONTANT TONTINE',
                'MODE VERSEMENT','VENTE PETIT LOT','INTERET PETIT LOT',
                'MODE VERSEMENT VENTE PETIT LOT','NUMERO CHEQUE',
                'MONTANT A REMBOURSER PETIT LOT','REMBOURSEMENT PETIT LOT',
                'MODE VERSEMENT REMBOURSEMENT','MODE VERSEMENTLOT PRINCIPAL',
                'NUMERO CHEQUE','MONTANT  LOT PRINCIPAL','INTERET LOT PRINCIPAL'
            ]
            for j, h in enumerate(COLS_T75100, 1):
                _hdr_cell(ws, 5, j, h)

            # Totaux
            tot_p=0; tot_t=D('0'); tot_v=D('0'); tot_i=D('0')
            tot_r=D('0'); tot_a=D('0'); tot_ai=D('0')
            tot_esp=D('0'); tot_bq=D('0')

            for i, (adh, s) in enumerate(membres_mois, 6):
                nbre   = int(_d(getattr(s, fld_nbre,  0)))
                tontine= _d(getattr(s, fld_ton,   D('0')))
                vente  = _d(getattr(s, fld_vente, D('0')))
                interet= _d(getattr(s, fld_int,   D('0')))
                achat  = _d(getattr(s, fld_achat, D('0')))
                mode   = s.mode_paiement_tontine or 'ECHEC'
                mode_lot = s.mode_paiement_lot   or 'ECHEC'
                mode_rem = s.mode_remb_petit_lot  or 'ECHEC'
                num_chq  = s.numero_cheque        or 0
                int_achat = D('0')  # intérêt lot principal

                # Montant à rembourser petit lot = vente avec remise
                mont_remb = vente  # simplifié
                remb_pl   = D('0')

                # Cumuler
                c = cumul_annuel.setdefault(adh.matricule, {
                    'adh':adh,'nbre':0,'tontine':D('0'),
                    'vente':D('0'),'interet':D('0'),'achat':D('0'),'pen':D('0')
                })
                c['nbre'] += nbre; c['tontine'] += tontine
                c['vente'] += vente; c['interet'] += interet; c['achat'] += achat

                tot_p+=nbre; tot_t+=tontine; tot_v+=vente; tot_i+=interet
                tot_a+=achat
                if mode == 'ESPECES': tot_esp += tontine
                else:                 tot_bq  += tontine

                ws.cell(i,1, adh.matricule)
                ws.cell(i,2, nbre)
                ws.cell(i,3, adh.nom_prenom)
                ws.cell(i,4, float(tontine))
                ws.cell(i,5, mode)
                ws.cell(i,6, float(vente))
                ws.cell(i,7, float(interet))
                ws.cell(i,8, None)
                ws.cell(i,9, num_chq)
                ws.cell(i,10, None)
                ws.cell(i,11, float(remb_pl))
                ws.cell(i,12, mode_rem)
                ws.cell(i,13, mode_lot)
                ws.cell(i,14, 0)
                ws.cell(i,15, float(achat))
                ws.cell(i,16, float(int_achat))

            # Lignes synthèse
            nb_m = len(membres_mois)
            row_tot = nb_m + 7

            ws.cell(row_tot-1, 1, 'TOTAL PART').font = Font(bold=True, size=9)
            ws.cell(row_tot-1, 2, tot_p)
            ws.cell(row_tot, 1, 'NBRE ADHERENT').font = Font(bold=True, size=9)
            ws.cell(row_tot, 2, nb_m)
            ws.cell(row_tot, 4, float(tot_t))
            ws.cell(row_tot, 6, float(tot_v))
            ws.cell(row_tot, 7, float(tot_i))
            ws.cell(row_tot, 15, float(tot_a))

            # Décomposition ESPECES/BANQUE/TOTAL
            r = row_tot + 2
            ws.cell(r,   3, 'ESPECES'); ws.cell(r,   4, float(tot_esp))
            ws.cell(r+1, 3, 'BANQUE');  ws.cell(r+1, 4, float(tot_bq))
            ws.cell(r+2, 3, 'TOTAL');   ws.cell(r+2, 4, float(tot_t))
            ws.cell(r+3, 3, None)

    # ── FEUILLE TONTRESUME (seulement T75 et T100) ───────────────
    if niveau in ('T75', 'T100'):
        ws_r = wb.create_sheet(f"TONTRESUME{str(annee)[-2:]}")
        ws_r.cell(1,1,'ASSOCIATION'); ws_r.cell(1,2,'ASELBY')
        ws_r.cell(2,1,'ANNEE');       ws_r.cell(2,2, annee)
        ws_r.cell(2,3,'TONIINE');     ws_r.cell(2,4,'NBRE PART')

        # Membres ayant participé dans l'année
        membres_annee = [
            v for v in cumul_annuel.values() if v['nbre'] > 0
        ]
        membres_annee.sort(key=lambda x: (
            0 if x['adh'].matricule == 'AS201648' else 1,
            x['adh'].numero_ordre
        ))

        tot_parts_an = sum(m['nbre'] for m in membres_annee)
        ws_r.cell(3,1,'TAUX TONTINE '); ws_r.cell(3,2, float(taux))
        ws_r.cell(3,3, 0); ws_r.cell(3,4, tot_parts_an)
        ws_r.cell(4,3, 'RECAPITULATIF')

        COLS_RESUME = ['MATRICULE','NBREPART','NOM ET PRENOM','MONTANT TONTINE',
                       'MODE VERSEMENT','VENTE PETIT LOT','INTERET PETIT LOT',
                       'MODE VERSEMENT VENTE PETIT LOT']
        for j, h in enumerate(COLS_RESUME, 1):
            _hdr_cell(ws_r, 5, j, h)

        tot_rn=0; tot_tn=D('0'); tot_vn=D('0'); tot_in=D('0')
        for i, m in enumerate(membres_annee, 6):
            ws_r.cell(i,1, m['adh'].matricule)
            ws_r.cell(i,2, m['nbre'])
            ws_r.cell(i,3, m['adh'].nom_prenom)
            ws_r.cell(i,4, float(m['tontine']))
            ws_r.cell(i,5, None)
            ws_r.cell(i,6, float(m['vente']))
            ws_r.cell(i,7, float(m['interet']))
            ws_r.cell(i,8, None)
            tot_rn += m['nbre']; tot_tn += m['tontine']
            tot_vn += m['vente']; tot_in += m['interet']

        nb_r = len(membres_annee)
        ws_r.cell(nb_r+7, 1, 'TOTAL PART').font = Font(bold=True, size=9)
        ws_r.cell(nb_r+7, 2, tot_rn)
        ws_r.cell(nb_r+8, 1, 'NBRE ADHERENT').font = Font(bold=True, size=9)
        ws_r.cell(nb_r+8, 2, nb_r)
        ws_r.cell(nb_r+8, 4, float(tot_tn))
        ws_r.cell(nb_r+8, 6, float(tot_vn))
        ws_r.cell(nb_r+8, 7, float(tot_in))

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    taux_label = {'T60':'60000','T75':'75000','T100':'100000'}[niveau]
    resp = HttpResponse(buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = (
        f'attachment; filename=ASELBY{annee}TONTINE{taux_label}.xlsx')
    return resp