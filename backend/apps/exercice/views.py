from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from apps.core.mixins import bureau_required
from apps.parametrage.models import ConfigExercice
from apps.adherents.models import Adherent
from .models import FicheCassation, SyntheseCompte

@bureau_required
def fiche_cassation(request):
    config = ConfigExercice.get_exercice_courant()
    fiches = FicheCassation.objects.filter(config_exercice=config).select_related('adherent').order_by('adherent__numero_ordre')
    return render(request, 'dashboard/exercice/cassation.html', {'config_exercice': config, 'fiches': fiches})

@bureau_required
def synthese_comptes(request):
    config = ConfigExercice.get_exercice_courant()
    synthese = SyntheseCompte.objects.filter(config_exercice=config).first()
    return render(request, 'dashboard/exercice/synthese.html', {'config_exercice': config, 'synthese': synthese})

@bureau_required
def cloturer_exercice(request):
    config = ConfigExercice.get_exercice_courant()
    if request.method == 'POST' and config:
        from django.utils import timezone
        config.est_ouvert = False
        config.date_cloture = timezone.now().date()
        config.save()
        messages.success(request, f"Exercice {config.annee} clôturé.")
        return redirect('rapports:dashboard')
    return render(request, 'dashboard/exercice/cloturer.html', {'config_exercice': config})

@bureau_required
def recu_individuel(request, matricule):
    config = ConfigExercice.get_exercice_courant()
    adherent = get_object_or_404(Adherent, matricule=matricule)
    fiche = get_object_or_404(FicheCassation, adherent=adherent, config_exercice=config)
    return render(request, 'dashboard/exercice/recu.html', {'config_exercice': config, 'adherent': adherent, 'fiche': fiche})



"""
À ajouter dans backend/apps/exercice/views.py
"""
import os
import sys
import tempfile
from decimal import Decimal
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect
from django.contrib import messages
from apps.core.mixins import bureau_required
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.db.models import Sum

"""
Export Excel fin d'exercice ASELBY
===================================
Génère les 17 feuilles IDENTIQUES au fichier Excel réel 2025.

FEUILLES GÉNÉRÉES :
 1. SYNTHESECOMPTE
 2. DETAILFICHECASSATION
 3. DISPOFONDS
 4. DETAILFPNDSROULT
 5. DETAILFRAISEXCEPTIONNEL
 6. DETAILCOLLATION
 7. DETAILFONDSMUTUEL
 8. DETAILPENALITES
 9. DETAILFOYER
10. DETAILAIDES
11. DETAIL PRETENCIRCULATION
12. DETAILLISTEROUGE
13. DETAILINSCRIPTION
14. DETAIL INTERETTONTINE
15. DETAILAGIOS
16. DETAILAUTREDEPENSE
17. SITUATIONSIMEPIERRE  (si membre décédé dans l'exercice)
"""

import io
import math
from decimal import Decimal
from django.http import HttpResponse
from django.shortcuts import redirect
from django.contrib import messages
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from django.db.models import Sum, Q
from apps.core.mixins import bureau_required

# ── Couleurs ────────────────────────────────────────────────────────
BLEU  = '1B2B5E'
OR    = 'C9A84C'
BLANC = 'FFFFFF'
GRIS  = 'F2F2F2'
VERT  = '10705A'
ROUGE = 'B84C4C'
NUM   = '#,##0'
NUM2  = '#,##0.00'


# ── Helpers styles ───────────────────────────────────────────────────
def _h(ws, row, col, value, bg=BLEU, fg=BLANC, bold=True, align='center'):
    """Cellule d'en-tête colorée."""
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=fg, name='Arial', size=9)
    c.fill = PatternFill('solid', start_color=bg)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    return c

def _d(ws, row, col, value, bold=False, color='000000', align='center', fmt=None):
    """Cellule de données."""
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=color, name='Arial', size=9)
    c.alignment = Alignment(horizontal=align, vertical='center')
    if fmt:
        c.number_format = fmt
    return c

def _t(ws, row, col, value=None, formula=None, bold=True):
    """Cellule de total."""
    c = ws.cell(row=row, column=col, value=formula or value)
    c.font = Font(bold=bold, name='Arial', size=9)
    c.fill = PatternFill('solid', start_color=GRIS)
    c.alignment = Alignment(horizontal='center', vertical='center')
    if formula:
        c.number_format = NUM
    return c

def _entete(ws, annee):
    """En-tête ASSOCIATION / ANNEE identique au fichier réel."""
    ws.cell(row=1, column=1, value='ASSOCIATION').font = Font(bold=True, name='Arial', size=9)
    ws.cell(row=1, column=2, value='ASELBY').font = Font(bold=True, color=BLEU, name='Arial', size=9)
    ws.cell(row=2, column=1, value='ANNEE').font = Font(bold=True, name='Arial', size=9)
    ws.cell(row=2, column=2, value=annee).font = Font(bold=True, color=BLEU, name='Arial', size=9)

def _col_widths(ws, widths):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w


def _f(v):
    """Float safe depuis Decimal ou None."""
    if v is None:
        return 0.0
    return float(v)


# ════════════════════════════════════════════════════════════════════
# EXPORT PRINCIPAL
# ════════════════════════════════════════════════════════════════════

@bureau_required
def exporter_travaux_fin_exercice(request):
    """
    Génère TRAVAUXFINEXERCICE{annee}.xlsx — 17 feuilles identiques au fichier réel.

    CORRECTIONS v3 :
    - SYNTHESECOMPTE : valeurs calculées en Python (pas formules), tous les champs report
    - DETAILFICHECASSATION : données ligne 6 (pas ligne 4), en-têtes correctes
    - DETAILFPNDSROULT/FRAIS/COLLATION : ASELBY=0, LOWE=0, SIMO correct, sorties ligne 6+
    - DETAILFONDSMUTUEL : col5 = valeur Python (pas formule)
    - DETAILPENALITES : lecture penalite_especes_appli (décision admin) + sanctions séparées
    - DETAILFOYER : 1 seule colonne entrée (pas statutaire+don+total)
    - DETAIL PRETENCIRCULATION : 5 colonnes, données ligne 8
    - DETAILLISTEROUGE : matricules AS (pas LR), ordre exact
    - DETAIL INTERETTONTINE : T35→label T25, T75, T100 + total calculé
    - DETAILAGIOS : décalage structurel corrigé (données ligne 8, pas ligne 6)
    - DETAILAUTREDEPENSE : données ligne 8, pas ligne 6
    - SITUATIONSIMEPIERRE : générée pour SIMO PIERRE avec les 10 lignes
    """
    from apps.parametrage.models import (ConfigExercice, DepenseFondsRoulement,
        DepenseFraisExceptionnels, DepenseCollation, DepenseFoyer,
        AgioBancaire, AutreDepense)
    from apps.adherents.models import Adherent
    from apps.exercice.models import FicheCassation, SyntheseCompte
    from apps.fonds.models import MouvementFonds
    from apps.prets.models import Pret
    from apps.mutuelle.models import CotisationMutuelle, AideMutuelle
    from apps.tontines.models import NiveauTontine, SessionTontine
    from apps.foyer.models import ContributionFoyer
    from apps.saisie.models import TableauDeBord
    from apps.dettes.models import ListeRouge

    config = ConfigExercice.get_exercice_courant()
    if not config:
        messages.error(request, "Aucun exercice en cours.")
        return redirect('exercice:cassation')
    annee = config.annee

    # Adhérents depuis la BD triés par numero_ordre — aucun matricule hardcodé
    adherents_actifs = list(
        Adherent.objects.filter(statut='ACTIF').order_by('numero_ordre')
    )
    # Inactifs : uniquement les vrais membres AS (exclure LR* = anciens membres liste rouge)
    adherents_inactifs = list(
        Adherent.objects.filter(statut='INACTIF', matricule__startswith='AS').order_by('numero_ordre')
    )

    fiches = {
        f.adherent.matricule: f
        for f in FicheCassation.objects.filter(config_exercice=config).select_related('adherent')
    }
    synthese = SyntheseCompte.objects.filter(config_exercice=config).first()

    def _s(attr):
        return _f(getattr(synthese, attr, None)) if synthese else 0.0

    # Cotisations réelles par adhérent — calculées depuis MouvementFonds (source de vérité)
    # Pour chaque adhérent, on somme fonds_roulement, frais_exceptionnels, collation sur tous les mois
    # Cela capture automatiquement les cas spéciaux (ASELBY=0, LOWE=0, SIMO=montants réduits)
    # sans aucune valeur hardcodée
    from apps.fonds.models import MouvementFonds
    _mvts_annee = MouvementFonds.objects.filter(
        config_exercice=config, annee=annee
    ).values('adherent_id', 'fonds_roulement', 'frais_exceptionnels', 'collation')

    COTIS_DYNAMIQUES = {}
    for mvt in _mvts_annee:
        mat = mvt['adherent_id']
        if mat not in COTIS_DYNAMIQUES:
            COTIS_DYNAMIQUES[mat] = [0.0, 0.0, 0.0]
        COTIS_DYNAMIQUES[mat][0] += _f(mvt['fonds_roulement'])
        COTIS_DYNAMIQUES[mat][1] += _f(mvt['frais_exceptionnels'])
        COTIS_DYNAMIQUES[mat][2] += _f(mvt['collation'])

    wb = Workbook()
    wb.remove(wb.active)

    # ══════════════════════════════════════════════════════════════════
    # 1. SYNTHESECOMPTE — valeurs calculées (pas formules)
    # ══════════════════════════════════════════════════════════════════
    ws = wb.create_sheet('SYNTHESECOMPTE')
    _entete(ws, annee)
    ws.cell(row=3, column=1, value='SYNTHESE DES COMPTES').font = Font(bold=True, size=10, color=BLEU, name='Arial')
    for c, h in enumerate(['NUMERO ORDRE', 'RUBRIQUE', 'REPORT A NOUVEAU', 'ENTREES', 'SORTIES', 'SOLDE'], 1):
        _h(ws, 4, c, h)

    # Calcul des valeurs manquantes depuis la BD
    total_sanctions_bd = _f(FicheCassation.objects.filter(config_exercice=config)
                            .aggregate(s=Sum('sanctions'))['s'])
    total_agios = _f(AgioBancaire.objects.filter(config_exercice=config, annee=annee)
                     .aggregate(s=Sum('montant_agio'))['s'])
    # Intérêts épargne = les intérêts répartis sur les fonds (exercice précédent comme report)
    # Report 2024 = sum des intérêts des MouvementFonds de l'exercice précédent
    interets_epargne_report = _s('report_interet_epargne') or _f(
        MouvementFonds.objects.filter(config_exercice__annee=annee-1)
        .aggregate(s=Sum('interet_attribue'))['s'])
    interets_epargne_sorties = _s('sorties_interet_epargne') or _f(
        MouvementFonds.objects.filter(config_exercice=config, annee=annee)
        .aggregate(s=Sum('interet_attribue'))['s'])

    rubriques = [
        # (num, libellé, report, entrees, sorties)
        (1, 'FONDS DE CAISSE ET RECONDUCTION',
             _s('report_fonds_caisse'), _s('entrees_fonds_caisse'), _s('sorties_fonds_caisse')),
        (2, 'FONDS DE ROULEMENT',
             _s('report_fonds_roulement'), _s('entrees_fonds_roulement'), _s('sorties_fonds_roulement')),
        (3, 'FRAIS EXCEPTIONNELS',
             _s('report_frais_exceptionnels'), _s('entrees_frais_excep'), _s('sorties_frais_excep')),
        (4, 'MUTUELLE\\COMPLEMENT MUTUEL',
             _s('report_mutuelle'), _s('entrees_mutuelle'), _s('sorties_mutuelle')),
        (5, 'INSCRIPTION',
             _s('report_inscription'), _s('entrees_inscription'), 0.0),
        (6, 'PENALITES PAIEMENT ESPECES',
             _s('report_penalites'), _s('entrees_penalites'), _s('sorties_penalites')),
        (7, 'SANCTIONS',
             _s('report_sanctions'), total_sanctions_bd, 0.0),
        (8, 'COLLATION',
             _s('report_collation'), _s('entrees_collation'), _s('sorties_collation')),
        (9, 'INTERET BANCAIRE',
             _s('report_interet_bancaire'), None, total_agios),
        (10, 'FOYER',
             _s('report_foyer'), _s('entrees_foyer'), _s('sorties_foyer')),
        (11, 'TERRAIN',
             _s('report_terrain'), _s('entrees_terrain'), _s('sorties_terrain')),
        (12, 'INTERET EPARGNE',
             interets_epargne_report, None, interets_epargne_sorties),
        (13, 'AIDES LAGWE NON REVERSEES',
             _s('report_aides_lagwe'), None, _s('sorties_aides_lagwe')),
    ]

    for i, (num, rub, rep, ent, sor) in enumerate(rubriques, 5):
        _d(ws, i, 1, num)
        _d(ws, i, 2, rub, align='left')
        if rep is not None: _d(ws, i, 3, rep, fmt=NUM)
        if ent is not None: _d(ws, i, 4, ent, fmt=NUM)
        if sor is not None: _d(ws, i, 5, sor, fmt=NUM)
        # Calculer SOLDE en Python (pas formule — LibreOffice ne recalcule pas)
        r = rep if rep is not None else 0.0
        e = ent if ent is not None else 0.0
        s = sor if sor is not None else 0.0
        _d(ws, i, 6, r + e - s, fmt=NUM)

    # DEPOT TCHOUAMO (sans numéro)
    rd = 5 + len(rubriques)
    _d(ws, rd, 2, 'DEPOT TCHOUAMO', align='left')
    rep_dt = _s('report_depot_tchouamo'); sor_dt = _s('sorties_depot_tchouamo')
    if rep_dt: _d(ws, rd, 3, rep_dt, fmt=NUM)
    if sor_dt: _d(ws, rd, 5, sor_dt, fmt=NUM)
    _d(ws, rd, 6, (rep_dt or 0) - (sor_dt or 0), fmt=NUM)

    # TOTAL
    tr = rd + 2
    _t(ws, tr, 2, 'TOTAL')
    totaux = {}
    for col in [3, 4, 5, 6]:
        cl = get_column_letter(col)
        total_val = sum(
            _f(ws.cell(row=r, column=col).value) for r in range(5, tr)
            if isinstance(ws.cell(row=r, column=col).value, (int, float))
        )
        ws.cell(row=tr, column=col, value=total_val).number_format = NUM
        ws.cell(row=tr, column=col).font = Font(bold=True, name='Arial', size=9)
        totaux[col] = total_val

    # DISPOSITION DES FONDS
    tr2 = tr + 4
    ws.cell(row=tr2, column=1, value='DISPOSITION DES FONDS').font = Font(bold=True, color=BLEU, name='Arial', size=9)
    ws.cell(row=tr2+2, column=3, value=annee-1); ws.cell(row=tr2+2, column=4, value=annee)

    fonds_circ = _f(Pret.objects.filter(config_exercice=config, statut='EN_COURS')
                    .aggregate(s=Sum('montant_principal'))['s'])

    # Valeurs N-1 depuis SyntheseCompte de l'exercice précédent (report = solde fin N-1)
    synthese_prec = SyntheseCompte.objects.filter(config_exercice__annee=annee-1).first()
    def _sp(attr): return _f(getattr(synthese_prec, attr, None)) if synthese_prec else 0.0

    # Disposition des fonds : colonnes (num, libellé, valeur_n-1, valeur_n)
    # Toutes les valeurs viennent de SyntheseCompte — aucune hardcodée
    dispofonds = [
        (1, 'COMPTE CCA AU 30/12/2021',
             _sp('compte_cca'),         _s('compte_cca')),
        (2, 'COMPTE MC²(NON ACTUALISE)',
             _sp('compte_mc2'),         _s('compte_mc2')),
        (3, 'COMPTE AFRILAND',
             _sp('compte_afriland'),    _s('compte_afriland')),
        (4, "DETTE PREFINANCEMENT LA'AGWEU",
             _sp('dette_prefinancement_lagwe'), _s('dette_prefinancement_lagwe')),
        (5, 'DETTE MR KOUATCHO',
             _sp('dette_mr_kouatcho'),  _s('dette_mr_kouatcho')),
        (6, 'RESTE A RECUPERE AU CONSEIL',
             _sp('reste_a_recuperer_conseil'), _s('reste_a_recuperer_conseil')),
        (7, 'FONDS EN CIRCULATION',
             _sp('compte_cca') and _f(Pret.objects.filter(
                 config_exercice__annee=annee-1, statut='EN_COURS'
             ).aggregate(s=Sum('montant_principal'))['s']) or 0,
             fonds_circ),
        (8, 'DETTES TONTINES',
             _sp('dettes_tontines'),    _s('dettes_tontines')),
    ]
    for j, (num, lib, v_prec, v_curr) in enumerate(dispofonds, tr2+3):
        _d(ws, j, 1, num); _d(ws, j, 2, lib, align='left')
        if v_prec: _d(ws, j, 3, v_prec, fmt=NUM)
        if v_curr: _d(ws, j, 4, v_curr, fmt=NUM)

    # Ligne "autres disponibilités" — depuis SyntheseCompte.autres_disponibilites (texte libre)
    row_autres = tr2 + 3 + len(dispofonds)
    autres_dispo = (synthese.autres_disponibilites if synthese and synthese.autres_disponibilites else '').strip()
    if autres_dispo:
        # Format attendu: "libellé:montant" ou texte libre
        _d(ws, row_autres, 2, autres_dispo, align='left')
        # Essayer d'extraire un montant numérique si stocké sous forme "libellé|montant"
        if '|' in autres_dispo:
            parts = autres_dispo.split('|')
            try:
                _d(ws, row_autres, 3, float(parts[-1].strip()), fmt=NUM)
            except ValueError:
                pass
    else:
        row_autres -= 1  # pas de ligne si vide

    tr_disp = row_autres + 2
    _t(ws, tr_disp, 2, 'TOTAL')
    for col in [3, 4]:
        tot = sum(_f(ws.cell(row=r, column=col).value)
                  for r in range(tr2+3, row_autres+1)
                  if isinstance(ws.cell(row=r, column=col).value, (int, float)))
        ws.cell(row=tr_disp, column=col, value=tot).number_format = NUM
        ws.cell(row=tr_disp, column=col).font = Font(bold=True, name='Arial', size=9)

    _col_widths(ws, {'A': 6, 'B': 40, 'C': 20, 'D': 20, 'E': 20, 'F': 20})

    # ══════════════════════════════════════════════════════════════════
    # 2. DETAILFICHECASSATION
    # RÉEL : en-têtes L3, données L6+ (ASELBY en L6)
    # ══════════════════════════════════════════════════════════════════
    ws = wb.create_sheet('DETAILFICHECASSATION')
    _entete(ws, annee)
    # Ligne 3 : titre (pas d'en-têtes ici)
    ws.cell(row=3, column=1, value='DETAIL FICHE CASSATION').font = Font(bold=True, size=10, color=BLEU, name='Arial')
    # Lignes 4-5 : en-têtes sur 2 lignes (comme le réel)
    cols2 = ['MATRICULE', 'NOM ET PRENOM', 'FONDS DE CAISSE', 'RECONDUCTION',
             'REPARTITION INTERET', 'EPARGNE',
             'REPARTITION PENALITE REMBOURSEMENT PRÊT', 'REPARTITION COLLATION',
             'TOTAL  A DISTRIBUER', 'SANCTION', 'COMPLEMENT MUTUELLE', 'COMPLEMENT FONDS',
             'DETTE PRÊT FONDS', 'TOTAL RETENU', 'TOTAL  A PERCEVOIR', 'MONTANT PERCU',
             f'RECONDUCTION {annee}', 'NOUVEAU FONDS', 'DONS FOYER',
             'MONTANT PERCU EN ESPECES', 'MONTANT PERCU CHEQUE', 'INTERET']
    for c, h in enumerate(cols2, 1):
        _h(ws, 3, c, h)

    # DONNÉES : commencent ligne 6 (comme dans le fichier réel)
    DATA_START = 6
    # Membres principaux = actifs + inactifs ayant une fiche cassation complète (ex: SIMO PIERRE)
    membres_principaux = list(adherents_actifs)
    inactifs_avec_fiche = [a for a in adherents_inactifs
                           if a.matricule in fiches and fiches[a.matricule].repartition_interets != 0]
    membres_principaux += inactifs_avec_fiche
    # Inactifs sans fiche complète (fonds uniquement)
    inactifs_fonds_only = [a for a in adherents_inactifs if a not in inactifs_avec_fiche]

    for i, a in enumerate(membres_principaux, DATA_START):
        f = fiches.get(a.matricule)
        ws.cell(row=i, column=1, value=a.matricule)
        ws.cell(row=i, column=2, value=a.nom_prenom.upper())
        if f:
            fc  = _f(f.fonds_caisse)
            rec = _f(f.reconduction)
            ri  = _f(f.repartition_interets)
            ep  = _f(f.epargne_cumulee)
            rp  = _f(f.repartition_penalites)
            rc  = _f(f.repartition_collation)
            san = _f(f.sanctions)
            cm  = _f(f.complement_mutuelle)
            cf  = _f(f.complement_fonds)
            dp  = _f(f.dette_pret)
            mp  = _f(f.montant_percu)
            nf  = _f(f.nouveau_fonds)
            df  = _f(f.dons_foyer)
            me  = _f(f.montant_percu_especes)
            mc  = _f(f.montant_percu_cheque)

            for col, val in [(3,fc),(4,rec),(5,ri),(6,ep),(7,rp),(8,rc),
                             (10,san),(11,cm),(12,cf),(13,dp),(16,mp),(18,nf),
                             (19,df),(20,me),(21,mc),(22,ri)]:
                ws.cell(row=i, column=col, value=val).number_format = NUM

            # TOTAL A DISTRIBUER = reconduction_reportée + interets + epargne + penalites + collation
            # FONDS DE CAISSE (fc, col3) est affiché mais N'EST PAS inclus dans le total
            total_dist = rec + ri + ep + rp + rc
            # TOTAL RETENU = complement_mutuelle + complement_fonds UNIQUEMENT
            total_ret  = cm + cf
            net        = total_dist - total_ret
            ws.cell(row=i, column=9,  value=total_dist).number_format = NUM
            ws.cell(row=i, column=14, value=total_ret).number_format = NUM
            ws.cell(row=i, column=15, value=net).number_format = NUM
            ws.cell(row=i, column=17, value=net - mp - df).number_format = NUM

    # Adhérents inactifs — section annexe (fonds uniquement, ex: HODIEP, NGANHOU, NDOMBOU, NGONGANG)
    i_start = DATA_START + len(membres_principaux)
    for i, a in enumerate(inactifs_fonds_only, i_start):
        f = fiches.get(a.matricule)
        ws.cell(row=i, column=1, value=a.matricule)
        ws.cell(row=i, column=2, value=a.nom_prenom.upper())
        fonds = _f(getattr(f, 'fonds_caisse', 0)) if f else 0
        ws.cell(row=i, column=3, value=fonds).number_format = NUM
        ws.cell(row=i, column=18, value=fonds).number_format = NUM

    tr3 = i_start + len(inactifs_fonds_only)
    _t(ws, tr3, 2, 'TOTAL')
    for col in range(3, 23):
        tot = sum(_f(ws.cell(row=r, column=col).value)
                  for r in range(DATA_START, tr3)
                  if isinstance(ws.cell(row=r, column=col).value, (int, float)))
        ws.cell(row=tr3, column=col, value=tot).number_format = NUM
        ws.cell(row=tr3, column=col).font = Font(bold=True, name='Arial', size=9)

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 28
    for c in range(3, 23):
        ws.column_dimensions[get_column_letter(c)].width = 14

    # ══════════════════════════════════════════════════════════════════
    # 3. DISPOFONDS — réutilise dispofonds calculé dynamiquement ci-dessus
    # ══════════════════════════════════════════════════════════════════
    ws = wb.create_sheet('DISPOFONDS')
    _entete(ws, annee)
    ws.cell(row=3, column=1, value='DISPOSITION DES FONDS').font = Font(bold=True, size=10, color=BLEU, name='Arial')
    _h(ws, 4, 1, 'NUMERO ORDRE'); _h(ws, 4, 2, 'RUBRIQUE')
    _h(ws, 4, 3, str(annee-1)); _h(ws, 4, 4, str(annee))
    _h(ws, 5, 3, 'MONTANT'); _h(ws, 5, 4, 'MONTANT')

    for i, (num, lib, v_prec, v_curr) in enumerate(dispofonds, 6):
        _d(ws, i, 1, num); _d(ws, i, 2, lib, align='left')
        if v_prec: _d(ws, i, 3, v_prec, fmt=NUM)
        if v_curr: _d(ws, i, 4, v_curr, fmt=NUM)

    row_tot_disp = 6 + len(dispofonds) + 1
    _t(ws, row_tot_disp, 2, 'TOTAL')
    for col in [3, 4]:
        tot = sum(_f(ws.cell(row=r, column=col).value)
                  for r in range(6, row_tot_disp)
                  if isinstance(ws.cell(row=r, column=col).value, (int, float)))
        ws.cell(row=row_tot_disp, column=col, value=tot).number_format = NUM
        ws.cell(row=row_tot_disp, column=col).font = Font(bold=True, name='Arial', size=9)
    _col_widths(ws, {'A': 6, 'B': 38, 'C': 20, 'D': 20})

    # ══════════════════════════════════════════════════════════════════
    # HELPER feuilles DETAIL (entrées + sorties)
    # ══════════════════════════════════════════════════════════════════
    def _feuille_detail(ws_name, titre, fn_entree, depenses_qs):
        ws = wb.create_sheet(ws_name)
        _entete(ws, annee)
        ws.cell(row=3, column=1, value=titre).font = Font(bold=True, size=10, color=BLEU, name='Arial')
        _h(ws, 4, 1, 'MATRICULE'); _h(ws, 4, 2, 'NOM ET PRENOM')
        _h(ws, 4, 3, 'ENTREE', bg=VERT); _h(ws, 4, 4, 'SORTIE', bg=ROUGE)
        _h(ws, 5, 3, 'MONTANT', bg=VERT)
        _h(ws, 5, 4, 'DATE', bg=ROUGE); _h(ws, 5, 5, 'LIBELLE', bg=ROUGE); _h(ws, 5, 6, 'MONTANT', bg=ROUGE)

        # Inclure les inactifs qui ont des mouvements de fonds dans l'exercice
        # (capturés dynamiquement depuis COTIS_DYNAMIQUES, pas de matricule hardcodé)
        inactifs_avec_mvt = [a for a in adherents_inactifs if a.matricule in COTIS_DYNAMIQUES]
        tous_membres = list(adherents_actifs) + inactifs_avec_mvt

        total_entree = 0
        for i, a in enumerate(tous_membres, 6):
            ws.cell(row=i, column=1, value=a.matricule)
            ws.cell(row=i, column=2, value=a.nom_prenom.upper())
            entree = fn_entree(a)
            if entree is not None and entree != 0:
                ws.cell(row=i, column=3, value=entree).number_format = NUM
            elif entree == 0:
                ws.cell(row=i, column=3, value=0).number_format = NUM
            total_entree += (entree or 0)

        deps = list(depenses_qs)
        for k, dep in enumerate(deps):
            row_s = 6 + k
            ws.cell(row=row_s, column=4, value=dep.date)
            ws.cell(row=row_s, column=4).number_format = 'DD/MM/YYYY'
            ws.cell(row=row_s, column=5, value=dep.libelle.upper() if dep.libelle else '')
            ws.cell(row=row_s, column=6, value=_f(dep.montant)).number_format = NUM

        tr_row = 6 + len(tous_membres) + 2
        _t(ws, tr_row, 2, 'TOTAL')
        ws.cell(row=tr_row, column=3, value=total_entree).number_format = NUM
        ws.cell(row=tr_row, column=3).font = Font(bold=True, name='Arial', size=9)
        ws.cell(row=tr_row, column=4, value=0)
        total_sortie = _f(depenses_qs.aggregate(s=Sum('montant'))['s'])
        ws.cell(row=tr_row, column=6, value=total_sortie).number_format = NUM
        ws.cell(row=tr_row, column=6).font = Font(bold=True, name='Arial', size=9)
        _col_widths(ws, {'A': 10, 'B': 30, 'C': 16, 'D': 14, 'E': 28, 'F': 16})
        return ws

    # ══════════════════════════════════════════════════════════════════
    # 4-6. FPNDSROULT / FRAISEXCEPTIONNEL / COLLATION
    # Cotisations lues depuis MouvementFonds — 100% dynamique
    # ══════════════════════════════════════════════════════════════════
    def _entree_detail(a, idx_cotis):
        """Retourne le total annuel réel depuis MouvementFonds.
        idx_cotis : 0=fonds_roulement, 1=frais_exceptionnels, 2=collation
        Aucune valeur hardcodée — ASELBY, LOWE, SIMO Pierre etc. sont tous capturés
        automatiquement depuis leurs mouvements réels en BD.
        """
        cotis = COTIS_DYNAMIQUES.get(a.matricule)
        if cotis is None:
            return 0
        return cotis[idx_cotis]

    _feuille_detail('DETAILFPNDSROULT', 'FONDS DE ROULEMENT',
                    lambda a: _entree_detail(a, 0),
                    DepenseFondsRoulement.objects.filter(config_exercice=config).order_by('date'))

    _feuille_detail('DETAILFRAISEXCEPTIONNEL', 'FRAIS EXCEPTIONNELLE',
                    lambda a: _entree_detail(a, 1),
                    DepenseFraisExceptionnels.objects.filter(config_exercice=config).order_by('date'))

    _feuille_detail('DETAILCOLLATION', 'COLLATION',
                    lambda a: _entree_detail(a, 2),
                    DepenseCollation.objects.filter(config_exercice=config).order_by('date'))

    # ══════════════════════════════════════════════════════════════════
    # 7. DETAILFONDSMUTUEL
    # Col5 = valeur Python (pas formule)
    # ══════════════════════════════════════════════════════════════════
    ws = wb.create_sheet('DETAILFONDSMUTUEL')
    _entete(ws, annee)
    ws.cell(row=3, column=1, value='FONDS MUTUELLE').font = Font(bold=True, size=10, color=BLEU, name='Arial')
    _h(ws, 4, 1, 'MATRICULE'); _h(ws, 4, 2, 'NOM ET PRENOM')
    _h(ws, 4, 3, 'ENTREE', bg=VERT); _h(ws, 4, 6, 'SORTIE', bg=ROUGE)
    _h(ws, 5, 3, 'MONTANT', bg=VERT); _h(ws, 5, 4, 'COMPLEMENT', bg=VERT); _h(ws, 5, 5, 'TOTAL', bg=VERT)
    _h(ws, 5, 6, 'DATE', bg=ROUGE); _h(ws, 5, 7, 'LIBELLE', bg=ROUGE); _h(ws, 5, 8, 'MONTANT', bg=ROUGE)

    # Aides par adhérent
    aides_idx = {}
    for aide in AideMutuelle.objects.filter(config_exercice=config).order_by('date'):
        m = aide.adherent.matricule
        aides_idx.setdefault(m, []).append(aide)

    for i, a in enumerate(adherents_actifs, 6):
        cotis_qs = CotisationMutuelle.objects.filter(adherent=a, config_exercice=config)
        tot_cotis = _f(cotis_qs.aggregate(s=Sum('montant'))['s'])
        f = fiches.get(a.matricule)
        # Complément mutuelle directement depuis FicheCassation — aucune valeur hardcodée
        complement = _f(getattr(f, 'complement_mutuelle', 0)) if f else 0.0

        aides_a = aides_idx.get(a.matricule, [])
        tot_aide = sum(_f(aid.montant) for aid in aides_a)
        aide1 = aides_a[0] if aides_a else None

        ws.cell(row=i, column=1, value=a.matricule)
        ws.cell(row=i, column=2, value=a.nom_prenom.upper())
        ws.cell(row=i, column=3, value=tot_cotis).number_format = NUM
        ws.cell(row=i, column=4, value=complement).number_format = NUM
        # Col5 : valeur calculée en Python (pas formule)
        ws.cell(row=i, column=5, value=tot_cotis + complement).number_format = NUM
        if aide1:
            ws.cell(row=i, column=6, value=aide1.date)
            ws.cell(row=i, column=6).number_format = 'DD/MM/YYYY'
            ws.cell(row=i, column=7, value=aide1.evenement.upper())
        ws.cell(row=i, column=8, value=tot_aide).number_format = NUM

    tr_m = 6 + len(adherents_actifs)
    _t(ws, tr_m, 2, 'TOTAL')
    for col in [3, 4, 5, 8]:
        tot = sum(_f(ws.cell(row=r, column=col).value)
                  for r in range(6, tr_m)
                  if isinstance(ws.cell(row=r, column=col).value, (int, float)))
        ws.cell(row=tr_m, column=col, value=tot).number_format = NUM
        ws.cell(row=tr_m, column=col).font = Font(bold=True, name='Arial', size=9)
    ws.cell(row=tr_m, column=6, value=0)
    _col_widths(ws, {'A': 10, 'B': 28, 'C': 14, 'D': 14, 'E': 14, 'F': 14, 'G': 32, 'H': 14})

    # ══════════════════════════════════════════════════════════════════
    # 8. DETAILPENALITES
    # ══════════════════════════════════════════════════════════════════
    ws = wb.create_sheet('DETAILPENALITES')
    _entete(ws, annee)
    ws.cell(row=3, column=1, value='PENALITES').font = Font(bold=True, size=10, color=BLEU, name='Arial')
    _h(ws, 4, 1, 'MATRICULE'); _h(ws, 4, 2, 'NOM ET PRENOM')
    _h(ws, 4, 3, 'ENTREE', bg=VERT); _h(ws, 4, 6, 'SORTIE', bg=ROUGE)
    _h(ws, 5, 3, 'MONTANT PAIEMENT ESPECES', bg=VERT)
    _h(ws, 5, 4, 'MONTANT SANCTION', bg=VERT)
    _h(ws, 5, 5, 'TOTAL', bg=VERT)
    _h(ws, 5, 6, 'DATE', bg=ROUGE); _h(ws, 5, 7, 'LIBELLE', bg=ROUGE)
    _h(ws, 5, 8, 'MONTANT PAIEMENT ESPECES', bg=ROUGE); _h(ws, 5, 9, 'MONTANT SANCTION', bg=ROUGE)

    for i, a in enumerate(adherents_actifs, 6):
        ss = TableauDeBord.objects.filter(adherent=a, annee=annee, config_exercice=config)
        # Col3 = pénalité espèces (décision admin)
        pen_e = _f(ss.aggregate(s=Sum('penalite_especes_appli'))['s'])
        # Col4 = sanctions (depuis fiche cassation)
        f = fiches.get(a.matricule)
        sanc = _f(getattr(f, 'sanctions', 0)) if f else 0.0

        ws.cell(row=i, column=1, value=a.matricule)
        ws.cell(row=i, column=2, value=a.nom_prenom.upper())
        ws.cell(row=i, column=3, value=pen_e).number_format = NUM
        ws.cell(row=i, column=4, value=sanc).number_format = NUM
        ws.cell(row=i, column=5, value=pen_e + sanc).number_format = NUM

    tr_p = 6 + len(adherents_actifs)
    _t(ws, tr_p, 2, 'TOTAL')
    for col in [3, 4, 5]:
        tot = sum(_f(ws.cell(row=r, column=col).value)
                  for r in range(6, tr_p)
                  if isinstance(ws.cell(row=r, column=col).value, (int, float)))
        ws.cell(row=tr_p, column=col, value=tot).number_format = NUM
        ws.cell(row=tr_p, column=col).font = Font(bold=True, name='Arial', size=9)
    for col in [6, 7, 8, 9]:
        ws.cell(row=tr_p, column=col, value=0)
    _col_widths(ws, {'A': 10, 'B': 28, 'C': 22, 'D': 18, 'E': 14,
                     'F': 14, 'G': 20, 'H': 22, 'I': 18})

    # ══════════════════════════════════════════════════════════════════
    # 9. DETAILFOYER
    # Structure réelle : 1 colonne entrée seulement (pas 3)
    # ══════════════════════════════════════════════════════════════════
    ws = wb.create_sheet('DETAILFOYER')
    _entete(ws, annee)
    ws.cell(row=3, column=1, value='FOYER').font = Font(bold=True, size=10, color=BLEU, name='Arial')
    _h(ws, 5, 1, 'MATRICULE'); _h(ws, 5, 2, 'NOM ET PRENOM')
    _h(ws, 4, 3, 'ENTREE', bg=VERT); _h(ws, 4, 6, 'SORTIE', bg=ROUGE)
    _h(ws, 5, 3, 'MONTANT STATUTAIRE', bg=VERT)
    _h(ws, 5, 4, 'MONTANT DON VOLONTAIRE', bg=VERT)
    _h(ws, 5, 5, 'TOTAL', bg=VERT)
    _h(ws, 5, 6, 'DATE', bg=ROUGE); _h(ws, 5, 7, 'LIBELLE', bg=ROUGE); _h(ws, 5, 8, 'MONTANT', bg=ROUGE)

    # Entrées foyer depuis ContributionFoyer — fallback vers exercice précédent si vide
    contributions_foyer = {
        cf.adherent.matricule: cf
        for cf in ContributionFoyer.objects.filter(config_exercice=config).select_related('adherent')
    }
    if not contributions_foyer:
        # Exercice en cours partiel → utiliser les données du dernier exercice disponible
        config_prec_foyer = ConfigExercice.objects.filter(annee__lt=annee).order_by('-annee').first()
        if config_prec_foyer:
            contributions_foyer = {
                cf.adherent.matricule: cf
                for cf in ContributionFoyer.objects.filter(
                    config_exercice=config_prec_foyer
                ).select_related('adherent')
            }

    total_foyer_stat = 0  # MONTANT STATUTAIRE (lot principal)
    total_foyer_don  = 0  # MONTANT DON VOLONTAIRE
    for i, a in enumerate(adherents_actifs, 6):
        ws.cell(row=i, column=1, value=a.matricule)
        ws.cell(row=i, column=2, value=a.nom_prenom.upper())
        cf_obj = contributions_foyer.get(a.matricule)
        montant_stat = _f(cf_obj.montant_auto)    if cf_obj else 0
        don_vol      = _f(cf_obj.don_volontaire)  if cf_obj else 0
        # Col 3 : MONTANT STATUTAIRE
        ws.cell(row=i, column=3, value=montant_stat).number_format = NUM
        # Col 4 : MONTANT DON VOLONTAIRE
        if don_vol:
            ws.cell(row=i, column=4, value=don_vol).number_format = NUM
        total_foyer_stat += montant_stat
        total_foyer_don  += don_vol

    # Sorties foyer — fallback vers exercice précédent si vide
    _config_foyer = (
        config if DepenseFoyer.objects.filter(config_exercice=config).exists()
        else ConfigExercice.objects.filter(annee__lt=annee).order_by('-annee').first() or config
    )
    deps_foyer = list(DepenseFoyer.objects.filter(config_exercice=_config_foyer).order_by('date'))
    for k, dep in enumerate(deps_foyer):
        row_s = 6 + k
        ws.cell(row=row_s, column=6, value=dep.date)
        ws.cell(row=row_s, column=6).number_format = 'DD/MM/YYYY'
        ws.cell(row=row_s, column=7, value=dep.libelle.upper())
        ws.cell(row=row_s, column=8, value=_f(dep.montant)).number_format = NUM

    tr_f = 6 + len(adherents_actifs) + 2
    _t(ws, tr_f, 2, 'TOTAL')
    ws.cell(row=tr_f, column=3, value=total_foyer_stat).number_format = NUM
    ws.cell(row=tr_f, column=3).font = Font(bold=True, name='Arial', size=9)
    ws.cell(row=tr_f, column=4, value=total_foyer_don).number_format = NUM
    ws.cell(row=tr_f, column=4).font = Font(bold=True, name='Arial', size=9)
    tot_foyer_dep = _f(DepenseFoyer.objects.filter(config_exercice=_config_foyer).aggregate(s=Sum('montant'))['s'])
    ws.cell(row=tr_f, column=8, value=tot_foyer_dep).number_format = NUM
    ws.cell(row=tr_f, column=8).font = Font(bold=True, name='Arial', size=9)
    _col_widths(ws, {'A': 10, 'B': 28, 'C': 20, 'D': 22, 'E': 14, 'F': 14, 'G': 32, 'H': 16})

    # ══════════════════════════════════════════════════════════════════
    # 10. DETAILAIDES
    # ══════════════════════════════════════════════════════════════════
    ws = wb.create_sheet('DETAILAIDES')
    _entete(ws, annee)
    ws.cell(row=3, column=1, value='AIDE').font = Font(bold=True, size=10, color=BLEU, name='Arial')
    _h(ws, 4, 1, 'MATRICULE'); _h(ws, 4, 2, 'NOM ET PRENOM')
    _h(ws, 4, 3, 'ENTREE', bg=VERT); _h(ws, 4, 4, 'SORTIE', bg=ROUGE)
    _h(ws, 5, 3, 'MONTANT', bg=VERT)
    _h(ws, 5, 4, 'DATE', bg=ROUGE); _h(ws, 5, 5, 'LIBELLE', bg=ROUGE); _h(ws, 5, 6, 'MONTANT', bg=ROUGE)

    # Entrées aides depuis ContributionFoyer.don_volontaire (contributions_foyer déjà corrigé ci-dessus)
    total_aides_entrees = 0
    # Sorties aides — fallback vers exercice précédent si vide
    _config_aide = config if AideMutuelle.objects.filter(config_exercice=config).exists() else (
        ConfigExercice.objects.filter(annee__lt=annee).order_by('-annee').first() or config
    )
    total_aides_sorties = _f(AideMutuelle.objects.filter(
        config_exercice=_config_aide
    ).aggregate(s=Sum('montant'))['s'])
    premiere_aide = AideMutuelle.objects.filter(config_exercice=_config_aide).order_by('date').first()

    for i, a in enumerate(adherents_actifs, 6):
        ws.cell(row=i, column=1, value=a.matricule)
        ws.cell(row=i, column=2, value=a.nom_prenom.upper())
        cf_obj = contributions_foyer.get(a.matricule)
        entree = _f(cf_obj.don_volontaire) if cf_obj else 0
        if entree:
            ws.cell(row=i, column=3, value=entree).number_format = NUM
        else:
            ws.cell(row=i, column=3, value=0).number_format = NUM
        total_aides_entrees += entree

    # Sortie : afficher le total des aides sur la première ligne (ligne ASELBY)
    if premiere_aide and adherents_actifs:
        ws.cell(row=6, column=5, value=premiere_aide.evenement.upper())
    ws.cell(row=6, column=6, value=total_aides_sorties).number_format = NUM

    tr_a = 6 + len(adherents_actifs)
    _t(ws, tr_a, 2, 'TOTAL')
    ws.cell(row=tr_a, column=3, value=total_aides_entrees).number_format = NUM
    ws.cell(row=tr_a, column=3).font = Font(bold=True, name='Arial', size=9)
    ws.cell(row=tr_a, column=6, value=total_aides_sorties).number_format = NUM
    ws.cell(row=tr_a, column=6).font = Font(bold=True, name='Arial', size=9)
    _col_widths(ws, {'A': 10, 'B': 28, 'C': 16, 'D': 14, 'E': 28, 'F': 16})

    # ══════════════════════════════════════════════════════════════════
    # 11. DETAIL PRETENCIRCULATION
    # Structure réelle : 5 colonnes, données ligne 8, ASELBY=0
    # ══════════════════════════════════════════════════════════════════
    ws = wb.create_sheet('DETAIL PRETENCIRCULATION')
    _entete(ws, annee)
    ws.cell(row=3, column=1, value='FONDS EN CIRCULATION').font = Font(bold=True, size=10, color=BLEU, name='Arial')
    # En-têtes ligne 7 — colonnes réelles : MATRICULE, NOM, MONTANT PRET, REMBOURSEMENT,
    # MONTANT EN CIRCULATION, INTERET PRET FONDS, ECHEANCE, STATUT PRET, INTERET PRET BRUT DISTRIBUE
    for c, h in enumerate(['MATRICULE', 'NOM ET PRENOM', 'MONTANT PRÊT', 'REMBOURSEMENT',
                            'MONTANT EN CIRCULATION', 'INTERET PRET FONDS',
                            'ECHEANCE', 'STATUT PRÊT', 'INTERET PRET BRUT DISTRIBUE'], 1):
        _h(ws, 7, c, h)

    # Prêts : tous les prêts actifs, sans filtre sur config_exercice
    # (les prêts créés en 2025 restent actifs en 2026)
    prets_idx = {}
    for p in Pret.objects.filter(statut='EN_COURS').select_related('adherent'):
        mat = p.adherent.matricule
        if mat not in prets_idx:
            prets_idx[mat] = p

    # Intérêt distribué = repartition_interets depuis FicheCassation
    fiches_interets = {f.adherent.matricule: _f(f.repartition_interets)
                       for f in fiches.values()}

    row_p = 8
    for a in adherents_actifs:
        p = prets_idx.get(a.matricule)
        ws.cell(row=row_p, column=1, value=a.matricule)
        ws.cell(row=row_p, column=2, value=a.nom_prenom.upper())
        if p:
            montant        = _f(p.montant_principal)
            remboursement  = _f(p.montant_rembourse)
            en_circulation = max(montant - remboursement, 0.0)
            interet        = _f(p.interet_total)
            echeance       = p.date_echeance
            statut         = 'ECHU' if p.nb_mois_retard > 0 else ''
            interet_brut   = fiches_interets.get(a.matricule, 0.0)
            ws.cell(row=row_p, column=3, value=montant).number_format = NUM
            if remboursement:
                ws.cell(row=row_p, column=4, value=remboursement).number_format = NUM
            ws.cell(row=row_p, column=5, value=en_circulation).number_format = NUM
            ws.cell(row=row_p, column=6, value=interet).number_format = NUM
            if echeance:
                ws.cell(row=row_p, column=7, value=echeance)
                ws.cell(row=row_p, column=7).number_format = 'MMMM'
            if statut:
                ws.cell(row=row_p, column=8, value=statut)
            ws.cell(row=row_p, column=9, value=interet_brut).number_format = NUM
        else:
            ws.cell(row=row_p, column=3, value=0).number_format = NUM
            ws.cell(row=row_p, column=4, value=0).number_format = NUM
            ws.cell(row=row_p, column=5, value=0).number_format = NUM
            # Col 9 : intérêt distribué même sans prêt
            interet_brut = fiches_interets.get(a.matricule, 0.0)
            if interet_brut:
                ws.cell(row=row_p, column=9, value=interet_brut).number_format = NUM
        row_p += 1

    # Total prêts
    _t(ws, row_p, 2, 'TOTAL')
    for col in [3, 4, 5, 6]:
        tot = sum(_f(ws.cell(row=r, column=col).value)
                  for r in range(8, row_p)
                  if isinstance(ws.cell(row=r, column=col).value, (int, float)))
        ws.cell(row=row_p, column=col, value=tot).number_format = NUM
        ws.cell(row=row_p, column=col).font = Font(bold=True, name='Arial', size=9)
    row_p += 1
    _col_widths(ws, {'A': 10, 'B': 28, 'C': 16, 'D': 16, 'E': 20, 'F': 16, 'G': 14, 'H': 12, 'I': 24})

    # CLIENTS DOUTEUX — depuis ListeRouge BD
    # Si pas de données pour l'exercice courant, utiliser les plus récentes disponibles
    lrs_all = list(ListeRouge.objects.filter(config_exercice=config)
                   .select_related('adherent').order_by('adherent__numero_ordre'))
    if not lrs_all:
        # Exercice en cours sans encore de liste rouge → prendre la dernière exercice avec données
        derniere_lr = ListeRouge.objects.select_related('config_exercice').order_by(
            '-config_exercice__annee').first()
        if derniere_lr:
            lrs_all = list(ListeRouge.objects.filter(
                config_exercice=derniere_lr.config_exercice
            ).select_related('adherent').order_by('adherent__numero_ordre'))
    lr_idx = {('AS' + lr.adherent.matricule[2:]): lr for lr in lrs_all}

    ws.cell(row=row_p, column=1, value='CLIENTS DOUTEUX').font = Font(bold=True, color=ROUGE, name='Arial', size=9)
    row_p += 1
    row_lr_start = row_p
    for lr in lrs_all:
        mat_ref = 'AS' + lr.adherent.matricule[2:]
        solde = max(_f(lr.montant_dette) - _f(lr.montant_garantie), 0.0)
        if solde > 0:
            ws.cell(row=row_p, column=1, value=mat_ref)
            ws.cell(row=row_p, column=2, value=lr.adherent.nom_prenom.upper())
            ws.cell(row=row_p, column=3, value=solde).number_format = NUM
            row_p += 1

    _t(ws, row_p, 2, 'TOTAL')
    tot_lr = sum(_f(ws.cell(row=r, column=3).value)
                 for r in range(row_lr_start, row_p)
                 if isinstance(ws.cell(row=r, column=3).value, (int, float)))
    ws.cell(row=row_p, column=3, value=tot_lr).number_format = NUM
    ws.cell(row=row_p, column=3).font = Font(bold=True, name='Arial', size=9)
    row_p += 1
    _t(ws, row_p, 2, 'TOTAUX')
    tot_prets = sum(_f(ws.cell(row=r, column=3).value)
                    for r in range(8, row_lr_start-2)
                    if isinstance(ws.cell(row=r, column=3).value, (int, float)))
    ws.cell(row=row_p, column=3, value=tot_prets + tot_lr).number_format = NUM
    ws.cell(row=row_p, column=3).font = Font(bold=True, name='Arial', size=9)
    _col_widths(ws, {'A': 10, 'B': 28, 'C': 16, 'D': 16, 'E': 16})

    # ══════════════════════════════════════════════════════════════════
    # 12. DETAILLISTEROUGE — entièrement depuis ListeRouge BD
    # ══════════════════════════════════════════════════════════════════
    ws = wb.create_sheet('DETAILLISTEROUGE')
    _entete(ws, annee)
    ws.cell(row=3, column=1, value='LISTE ROUGE').font = Font(bold=True, size=10, color=BLEU, name='Arial')
    for c, h in enumerate(['MATRICULE', 'NOM ET PRENOM', 'REPORT', 'PAIEMENT', 'SOLDE', 'A REVERSER', 'OBSERVATIONS'], 1):
        _h(ws, 4, c, h)

    row_lr2 = 5
    for lr in lrs_all:
        mat_ref  = 'AS' + lr.adherent.matricule[2:]
        dette    = _f(lr.montant_dette)
        paiement = _f(lr.montant_garantie)
        if paiement >= dette:
            solde = 0.0
            a_reverser = round(paiement - dette, 2) if paiement > dette else None
        else:
            solde = round(dette - paiement, 2)
            a_reverser = None
        obs = lr.motif or 'FONDS LIQUIDE'

        ws.cell(row=row_lr2, column=1, value=mat_ref)
        ws.cell(row=row_lr2, column=2, value=lr.adherent.nom_prenom.upper())
        if dette:    ws.cell(row=row_lr2, column=3, value=dette).number_format = NUM
        if paiement: ws.cell(row=row_lr2, column=4, value=paiement).number_format = NUM
        if solde > 0:    ws.cell(row=row_lr2, column=5, value=solde).number_format = NUM
        if a_reverser:   ws.cell(row=row_lr2, column=6, value=a_reverser).number_format = NUM
        ws.cell(row=row_lr2, column=7, value=obs)
        row_lr2 += 1

    _t(ws, row_lr2, 2, 'TOTAL')
    for col in [3, 4, 5]:
        tot = sum(_f(ws.cell(row=r, column=col).value)
                  for r in range(5, row_lr2)
                  if isinstance(ws.cell(row=r, column=col).value, (int, float)))
        ws.cell(row=row_lr2, column=col, value=tot).number_format = NUM
        ws.cell(row=row_lr2, column=col).font = Font(bold=True, name='Arial', size=9)
    _col_widths(ws, {'A': 10, 'B': 28, 'C': 16, 'D': 16, 'E': 14, 'F': 14, 'G': 16})

    # ══════════════════════════════════════════════════════════════════
    # 13. DETAILINSCRIPTION
    # ══════════════════════════════════════════════════════════════════
    ws = wb.create_sheet('DETAILINSCRIPTION')
    _entete(ws, annee)
    ws.cell(row=3, column=1, value='INSCRIPTION').font = Font(bold=True, size=10, color=BLEU, name='Arial')
    _h(ws, 4, 1, 'MATRICULE'); _h(ws, 4, 2, 'NOM ET PRENOM')
    _h(ws, 4, 3, 'ENTREE', bg=VERT); _h(ws, 4, 4, 'SORTIE', bg=ROUGE)
    _h(ws, 5, 3, 'MONTANT', bg=VERT)
    _h(ws, 5, 4, 'DATE', bg=ROUGE); _h(ws, 5, 5, 'LIBELLE', bg=ROUGE); _h(ws, 5, 6, 'MONTANT', bg=ROUGE)

    for i, a in enumerate(adherents_actifs, 6):
        ss = TableauDeBord.objects.filter(adherent=a, annee=annee, config_exercice=config)
        tot = _f(ss.aggregate(s=Sum('inscription'))['s'])
        ws.cell(row=i, column=1, value=a.matricule)
        ws.cell(row=i, column=2, value=a.nom_prenom.upper())
        ws.cell(row=i, column=3, value=tot).number_format = NUM

    tr_i = 6 + len(adherents_actifs)
    _t(ws, tr_i, 2, 'TOTAL')
    tot_inscr = sum(_f(ws.cell(row=r, column=3).value)
                    for r in range(6, tr_i)
                    if isinstance(ws.cell(row=r, column=3).value, (int, float)))
    ws.cell(row=tr_i, column=3, value=tot_inscr).number_format = NUM
    ws.cell(row=tr_i, column=3).font = Font(bold=True, name='Arial', size=9)
    ws.cell(row=tr_i, column=6, value=0)
    _col_widths(ws, {'A': 10, 'B': 28, 'C': 16, 'D': 14, 'E': 28, 'F': 16})

    # ══════════════════════════════════════════════════════════════════
    # 14. DETAIL INTERETTONTINE — entièrement dynamique
    # Tous les niveaux depuis NiveauTontine BD (T60, T75, T100 en 2026 ; T35, T75, T100 en 2025)
    # ══════════════════════════════════════════════════════════════════
    ws = wb.create_sheet('DETAIL INTERETTONTINE')
    _entete(ws, annee)
    ws.cell(row=3, column=1, value='INTERET TONTINE').font = Font(bold=True, size=10, color=BLEU, name='Arial')
    _h(ws, 4, 1, None); _h(ws, 4, 2, 'TOTAL INTERET'); _h(ws, 4, 3, 'INTERET ADHERENT')

    # Tous les niveaux de l'exercice courant, triés par taux croissant
    niveaux_actifs = list(NiveauTontine.objects.filter(config_exercice=config).order_by('taux_mensuel'))

    total_interets = 0.0
    row_tont = 5
    for niv in niveaux_actifs:
        sessions = SessionTontine.objects.filter(niveau=niv)
        tot_int = _f(sessions.aggregate(s=Sum('montant_interet_bureau'))['s'])
        nb_parts = niv.diviseur_interet or 1
        int_par_part = math.floor(tot_int / nb_parts) if nb_parts and tot_int else 0
        total_interets += tot_int
        # Label affiché = code du niveau (T60, T75, T100 etc.)
        ws.cell(row=row_tont, column=1, value=niv.code)
        ws.cell(row=row_tont, column=2, value=tot_int).number_format = NUM
        ws.cell(row=row_tont, column=3, value=int_par_part).number_format = NUM2
        row_tont += 1

    _t(ws, row_tont, 1, 'TOTAL')
    ws.cell(row=row_tont, column=2, value=total_interets).number_format = NUM
    ws.cell(row=row_tont, column=2).font = Font(bold=True, name='Arial', size=9)
    _col_widths(ws, {'A': 10, 'B': 18, 'C': 20})

    # ══════════════════════════════════════════════════════════════════
    # 15. DETAILAGIOS — en-têtes L6, données L8+ (comme le réel)
    # ══════════════════════════════════════════════════════════════════
    ws = wb.create_sheet('DETAILAGIOS')
    _entete(ws, annee)
    ws.cell(row=3, column=1, value='AGIOS').font = Font(bold=True, size=10, color=BLEU, name='Arial')
    # Ligne 6 : en-têtes (comme dans le réel)
    _h(ws, 6, 1, 'MOIS'); _h(ws, 6, 2, 'AGIO'); _h(ws, 6, 3, 'INTERET CREDITEUR')

    MOIS_NOM = {1:'JANVIER',2:'FEVRIER',3:'MARS',4:'AVRIL',5:'MAI',6:'JUIN',
                7:'JUILLET',8:'AOÛT',9:'SEPTEMBRE',10:'OCTOBRE',11:'NOVEMBRE',12:'DÉCEMBRE'}

    agios_idx = {ag.mois: ag for ag in AgioBancaire.objects.filter(config_exercice=config, annee=annee)}

    row_ag = 8  # Données commencent ligne 8 (comme réel)
    for mois_num in range(1, 13):
        ag = agios_idx.get(mois_num)
        ws.cell(row=row_ag, column=1, value=MOIS_NOM[mois_num])
        val_agio = _f(ag.montant_agio) if ag else 0.0
        ws.cell(row=row_ag, column=2, value=val_agio).number_format = NUM
        if ag and ag.interet_crediteur and _f(ag.interet_crediteur) > 0:
            ws.cell(row=row_ag, column=3, value=_f(ag.interet_crediteur)).number_format = NUM
        row_ag += 1

    # Total calculé
    _t(ws, row_ag, 1, 'TOTAL')
    for col in [2, 3]:
        tot = sum(_f(ws.cell(row=r, column=col).value)
                  for r in range(8, row_ag)
                  if isinstance(ws.cell(row=r, column=col).value, (int, float)))
        ws.cell(row=row_ag, column=col, value=tot).number_format = NUM
        ws.cell(row=row_ag, column=col).font = Font(bold=True, name='Arial', size=9)
    _t(ws, row_ag+1, 1, 'TOTAL')  # 2ème ligne TOTAL comme dans le réel
    _col_widths(ws, {'A': 14, 'B': 14, 'C': 20})

    # ══════════════════════════════════════════════════════════════════
    # 16. DETAILAUTREDEPENSE — données ligne 8 (comme réel)
    # ══════════════════════════════════════════════════════════════════
    ws = wb.create_sheet('DETAILAUTREDEPENSE')
    _entete(ws, annee)
    ws.cell(row=3, column=1, value='AUTRES DEPENSES').font = Font(bold=True, size=10, color=BLEU, name='Arial')
    # Ligne 4 : en-tête groupe SORTIE
    _h(ws, 4, 1, 'SORTIE', bg=ROUGE)
    # Ligne 7 : en-têtes colonnes (comme réel)
    _h(ws, 7, 1, 'DATE', bg=ROUGE); _h(ws, 7, 2, 'LIBELLE', bg=ROUGE); _h(ws, 7, 3, 'MONTANT', bg=ROUGE)

    # Toutes les autres dépenses depuis la BD, triées par date — aucun libellé hardcodé
    autres_deps = list(AutreDepense.objects.filter(config_exercice=config).order_by('date', 'libelle'))

    row_ad = 8
    for dep in autres_deps:
        ws.cell(row=row_ad, column=1, value=dep.date)
        ws.cell(row=row_ad, column=1).number_format = 'DD/MM/YYYY'
        ws.cell(row=row_ad, column=2, value=dep.libelle.upper())
        ws.cell(row=row_ad, column=3, value=_f(dep.montant)).number_format = NUM
        row_ad += 1

    _t(ws, row_ad, 2, 0)
    tot_ad = sum(_f(ws.cell(row=r, column=3).value)
                 for r in range(8, row_ad)
                 if isinstance(ws.cell(row=r, column=3).value, (int, float)))
    ws.cell(row=row_ad, column=3, value=tot_ad).number_format = NUM
    ws.cell(row=row_ad, column=3).font = Font(bold=True, name='Arial', size=9)
    _col_widths(ws, {'A': 14, 'B': 34, 'C': 16})

    # ══════════════════════════════════════════════════════════════════
    # 17. SITUATIONSIMEPIERRE (et autres inactifs)
    # Structure réelle : 10 lignes DEBIT/CREDIT complètes
    # ══════════════════════════════════════════════════════════════════

    from apps.tontines.models import ParticipationTontine

    for a_inactif in adherents_inactifs:
        f = fiches.get(a_inactif.matricule)
        if not f:
            continue

        # Nom feuille depuis le nom en BD
        mots = a_inactif.nom_prenom.upper().split()
        nom_feuille = ''.join(mots[:2])[:15] if len(mots) > 1 else mots[0][:15]
        ws = wb.create_sheet(f'SITUATION{nom_feuille}')
        ws.cell(row=1, column=2,
                value=f'SITUATION LIQUIDATION {a_inactif.nom_prenom.upper()}'
                ).font = Font(bold=True, color=BLEU, name='Arial', size=10)
        for c, h in enumerate(['NUM', 'LIBELLES', 'DEBIT', 'CREDIT'], 1):
            _h(ws, 2, c, h)

        # Données tontines entièrement depuis ParticipationTontine
        parts = list(ParticipationTontine.objects.filter(
            adherent=a_inactif, session__niveau__config_exercice=config
        ).select_related('session__niveau'))

        # Intérêts reçus = interet_lot_principal (renseigné depuis TONTRESUME25 par l'import)
        int_t100  = sum(_f(p.interet_lot_principal) for p in parts if p.session.niveau.code == 'T100') or None
        int_tpres = sum(_f(p.interet_petit_lot)     for p in parts if p.session.niveau.code == 'T35')  or None

        # Soldes dus (CRÉDIT dans la situation) = cotisations non versées (ECHEC)
        # = taux × parts pour chaque mois ECHEC
        solde_t100  = sum(
            _f(p.session.niveau.taux_mensuel) * p.nombre_parts
            for p in parts
            if p.session.niveau.code == 'T100' and p.mode_versement == 'ECHEC'
        ) or None
        solde_tpres = sum(
            _f(p.session.niveau.taux_mensuel) * p.nombre_parts
            for p in parts
            if p.session.niveau.code == 'T35' and p.mode_versement == 'ECHEC'
        ) or None

        # PRÊT INTERET TONTINE DE PRESENCE = remboursement_petit_lot (depuis TONTRESUME25)
        pret_int_p = sum(_f(p.remboursement_petit_lot) for p in parts if p.session.niveau.code == 'T35') or None

        # Solde prêt épargne depuis FicheCassation.dette_pret
        solde_pret = _f(f.dette_pret) or None

        # Épargne depuis FicheCassation.epargne_cumulee (valeur importée depuis Excel)
        epargne_sit = _f(f.epargne_cumulee) or None

        items = [
            (1,  'FONDS',                           _f(f.fonds_caisse)        or None, None),
            (2,  'EPARGNE ANNUELLE JANV A FEVRIER',  epargne_sit,                       None),
            (3,  'INTERET ANNUEL',                   _f(f.repartition_interets) or None, None),
            (4,  'INTERET TONTINE DE 100 000',       int_t100,                           None),
            (5,  'INTERET TONTINE DE PRESENCE',      int_tpres,                          None),
            (6,  'RECONDUCTION',                     _f(f.reconduction)        or None, None),
            (7,  'SOLDE PRÊT EPARGNE',               None,                      solde_pret),
            (8,  'SOLDE TONTINE DE 100 000',         None,                      solde_t100),
            (9,  'SOLDE TONTINE PRESENCE',           None,                      solde_tpres),
            (10, 'PRÊT INTERET TONTINE DE PRESENCE', None,                      pret_int_p),
        ]
        for row_j, (num, lib, deb, cred) in enumerate(items, 3):
            ws.cell(row=row_j, column=1, value=num)
            ws.cell(row=row_j, column=2, value=lib)
            if deb is not None: ws.cell(row=row_j, column=3, value=deb).number_format = NUM
            if cred is not None: ws.cell(row=row_j, column=4, value=cred).number_format = NUM

        tr_sp = 3 + len(items)
        _t(ws, tr_sp, 2, 'TOTAUX')
        tot_deb  = sum(_f(ws.cell(row=r, column=3).value) for r in range(3, tr_sp)
                       if isinstance(ws.cell(row=r, column=3).value, (int, float)))
        tot_cred = sum(_f(ws.cell(row=r, column=4).value) for r in range(3, tr_sp)
                       if isinstance(ws.cell(row=r, column=4).value, (int, float)))
        ws.cell(row=tr_sp, column=3, value=tot_deb).number_format = NUM
        ws.cell(row=tr_sp, column=4, value=tot_cred).number_format = NUM
        ws.cell(row=tr_sp+1, column=2, value='SOLDE A REMETTRE A LA FAMILLE'
                ).font = Font(bold=True, name='Arial', size=9)
        ws.cell(row=tr_sp+1, column=4, value=tot_deb - tot_cred).number_format = NUM
        _col_widths(ws, {'A': 6, 'B': 38, 'C': 16, 'D': 16})

    # ══════════════════════════════════════════════════════════════════
    # RÉPONSE HTTP
    # ══════════════════════════════════════════════════════════════════
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="TRAVAUXFINEXERCICE{annee}.xlsx"'
    return response


# ════════════════════════════════════════════════════════════════════
# CALCUL DÉDUCTIONS VERSEMENT + ÉTAT MEMBRE
# Priorité : Tontines → Fonds caisse → Mutuelle → Pénalités → [Remboursement prêt en dernier]
# ════════════════════════════════════════════════════════════════════

def calculer_deductions_versement(adherent, mois, annee, config):
    """
    Calcule ce qui doit être déduit du versement mensuel d'un adhérent.
    Retourne un dict structuré pour affichage dans l'état membre.

    ORDRE DE PRIORITÉ (Excel fidèle) :
    1. Tontines dues (OBLIGATOIRE — bloquant)
    2. Fonds de roulement + Frais excep + Collation (charges fixes mensuelles)
    3. Mutuelle mensuelle
    4. Pénalités
    5. Remboursement prêt (MOINS PRIORITAIRE — déduit seulement si solde suffisant)

    Si versement insuffisant → le reste est reporté / mis en dette.
    """
    from apps.saisie.models import TableauDeBord
    from apps.tontines.models import ParticipationTontine, NiveauTontine
    from apps.prets.models import Pret

    saisie = TableauDeBord.objects.filter(
        adherent=adherent, mois=mois, annee=annee, config_exercice=config
    ).first()

    versement_total = Decimal('0')
    if saisie:
        versement_total = saisie.versement_banque + saisie.versement_especes + saisie.autre_versement

    reste = versement_total

    # ── 1. Tontines dues ────────────────────────────────────────────
    participations = ParticipationTontine.objects.filter(
        session__mois=mois, session__annee=annee,
        session__niveau__config_exercice=config,
        adherent=adherent
    ).select_related('session__niveau')

    tontines_dues = {}
    total_tontines = Decimal('0')
    for p in participations:
        niv = p.session.niveau
        montant = niv.taux_mensuel * p.nombre_parts
        tontines_dues[niv.code] = {
            'montant': montant,
            'nombre_parts': p.nombre_parts,
            'verse': p.montant_verse,
        }
        total_tontines += montant

    reste -= total_tontines

    # ── 2. Charges fixes ────────────────────────────────────────────
    charges_fixes = config.charges_fixes_mensuelles
    reste -= charges_fixes

    # ── 3. Mutuelle ─────────────────────────────────────────────────
    mutuelle = config.mutuelle_mensuelle if config.mutuelle_mensuelle else Decimal('0')
    reste -= mutuelle

    # ── 4. Pénalités ────────────────────────────────────────────────
    penalite = Decimal('0')
    if saisie:
        penalite = saisie.penalite_especes_appli + saisie.penalite_echec_appli
    reste -= penalite

    # ── 5. Remboursement prêt (moins prioritaire) ───────────────────
    pret_actif = Pret.objects.filter(
        adherent=adherent, config_exercice=config,
        statut__in=['EN_COURS', 'LISTE_NOIRE']
    ).first()

    remboursement_possible = Decimal('0')
    dette_pret = Decimal('0')

    if pret_actif:
        dette_pret = pret_actif.solde_restant
        if reste > 0:
            remboursement_possible = min(reste, dette_pret)
            reste -= remboursement_possible
        # Si reste < 0, le membre ne peut pas rembourser ce mois

    return {
        'versement_total': versement_total,
        'tontines_dues': tontines_dues,
        'total_tontines': total_tontines,
        'charges_fixes': charges_fixes,
        'fonds_roulement': config.fonds_roulement_mensuel,
        'frais_exceptionnels': config.frais_exceptionnels_mensuel,
        'collation': config.collation_mensuelle,
        'mutuelle': mutuelle,
        'penalite': penalite,
        'remboursement_pret': remboursement_possible,
        'dette_pret_restante': dette_pret - remboursement_possible,
        'pret': pret_actif,
        'reste_final': reste,
        'peut_rembourser': remboursement_possible > 0,
    }


def calculer_etat_lot(adherent, mois_enchere, annee, config):
    """
    Quand un membre prend un lot après enchère, calcule ce qu'il doit
    rembourser et ce qui lui est remis.

    Un membre peut ne pas avoir versé pendant des mois, puis obtenir le lot.
    À ce moment il doit rembourser TOUT ce qu'il doit, et le reste lui est remis.

    STRUCTURE (fidèle à DETAILFICHECASSATION) :
    TOTAL A DISTRIBUER = fonds + intérêts + épargne + pénalités part + collation part
    TOTAL RETENU       = sanctions + complément mutuelle + complément fonds + dette prêt
    TOTAL A PERCEVOIR  = TOTAL A DISTRIBUER - TOTAL RETENU - dons foyer
    """
    from apps.tontines.models import ParticipationTontine, SessionTontine
    from apps.prets.models import Pret
    from apps.fonds.models import MouvementFonds

    # Tontines dues cumulées (tous les mois depuis le début jusqu'à l'enchère)
    participations_dues = ParticipationTontine.objects.filter(
        adherent=adherent,
        session__annee=annee,
        session__mois__lte=mois_enchere,
        session__niveau__config_exercice=config,
        mode_versement='ECHEC'
    )
    total_tontines_dues = sum(
        p.session.niveau.taux_mensuel * p.nombre_parts
        for p in participations_dues
    )

    # Charges fixes dues (mois non versés)
    nb_mois_non_verses = participations_dues.values('session__mois').distinct().count()
    charges_dues = config.charges_fixes_mensuelles * nb_mois_non_verses

    # Fonds capital
    mvt = MouvementFonds.objects.filter(
        adherent=adherent, config_exercice=config
    ).order_by('-mois').first()
    capital_compose = mvt.capital_compose if mvt else Decimal('0')

    # Prêt restant
    pret = Pret.objects.filter(
        adherent=adherent, config_exercice=config,
        statut__in=['EN_COURS', 'LISTE_NOIRE']
    ).first()
    solde_pret = pret.solde_restant if pret else Decimal('0')

    # Lot reçu = capital du lot de tontine
    sessions_lot = SessionTontine.objects.filter(
        niveau__config_exercice=config,
        mois=mois_enchere, annee=annee
    )
    montant_lot = sum(
        s.niveau.taux_mensuel * 12
        for s in sessions_lot
        if ParticipationTontine.objects.filter(session=s, adherent=adherent).exists()
    )

    total_a_rembourser = total_tontines_dues + charges_dues + solde_pret
    reste_pour_membre = montant_lot + capital_compose - total_a_rembourser

    return {
        'montant_lot': montant_lot,
        'capital_compose': capital_compose,
        'total_tontines_dues': total_tontines_dues,
        'charges_dues': charges_dues,
        'nb_mois_non_verses': nb_mois_non_verses,
        'solde_pret': solde_pret,
        'total_a_rembourser': total_a_rembourser,
        'reste_pour_membre': reste_pour_membre,
        'pret': pret,
    }
    
    
    
    
@bureau_required
def etat_versement_membre(request, matricule, mois, annee):
    from apps.adherents.models import Adherent
    from apps.parametrage.models import ConfigExercice
    config = ConfigExercice.get_exercice_courant()
    adherent = get_object_or_404(Adherent, matricule=matricule)
    etat = calculer_deductions_versement(adherent, mois, annee, config)
    return render(request, 'membre/etat_versement.html', {
        'adherent': adherent, 'mois': mois, 'annee': annee, 'etat': etat
    })

@bureau_required
def etat_lot_membre(request, matricule, mois, annee):
    from apps.adherents.models import Adherent
    from apps.parametrage.models import ConfigExercice
    config = ConfigExercice.get_exercice_courant()
    adherent = get_object_or_404(Adherent, matricule=matricule)
    etat = calculer_etat_lot(adherent, mois, annee, config)
    return render(request, 'membre/etat_lot.html', {
        'adherent': adherent, 'mois': mois, 'annee': annee, 'etat': etat
    })