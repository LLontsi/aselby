from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.core.paginator import Paginator
from apps.core.mixins import bureau_required
from apps.parametrage.models import ConfigExercice
from .models import Adherent
from .forms import AdherentForm

@bureau_required
def liste_adherents(request):
    config = ConfigExercice.get_exercice_courant()
    qs = Adherent.objects.filter(statut='ACTIF').order_by('numero_ordre')
    q = request.GET.get('q', '')
    if q:
        qs = qs.filter(nom_prenom__icontains=q)
    paginator = Paginator(qs, 20)
    page_obj = paginator.get_page(request.GET.get('page'))
    ctx = {'config_exercice': config, 'page_obj': page_obj, 'adherents': page_obj.object_list, 'q': q}
    return render(request, 'dashboard/adherents/liste.html', ctx)

@bureau_required
def fiche_adherent(request, matricule):
    config = ConfigExercice.get_exercice_courant()
    adherent = get_object_or_404(Adherent, matricule=matricule)
    from apps.fonds.models import MouvementFonds
    from apps.tontines.models import ParticipationTontine
    from apps.prets.models import Pret
    mouvements = MouvementFonds.objects.filter(adherent=adherent, annee=config.annee).order_by('mois')
    participations = ParticipationTontine.objects.filter(adherent=adherent, session__annee=config.annee).select_related('session__niveau')
    prets = Pret.objects.filter(adherent=adherent).order_by('-date_octroi')
    ctx = {
        'config_exercice': config, 'adherent': adherent,
        'mouvements': mouvements, 'participations': participations, 'prets': prets,
        'liste_rouge': getattr(adherent, 'liste_rouge', None),
    }
    return render(request, 'dashboard/adherents/fiche.html', ctx)

@bureau_required
def nouvel_adherent(request):
    config = ConfigExercice.get_exercice_courant()
    form = AdherentForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        adherent = form.save()
        messages.success(request, f"Adhérent {adherent.nom_prenom} créé avec succès.")
        return redirect('adherents:fiche', matricule=adherent.matricule)
    return render(request, 'dashboard/adherents/form.html', {'config_exercice': config, 'form': form})

@bureau_required
def modifier_adherent(request, matricule):
    config = ConfigExercice.get_exercice_courant()
    adherent = get_object_or_404(Adherent, matricule=matricule)
    form = AdherentForm(request.POST or None, instance=adherent)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "Fiche mise à jour.")
        return redirect('adherents:fiche', matricule=matricule)
    return render(request, 'dashboard/adherents/form.html', {'config_exercice': config, 'form': form, 'adherent': adherent})

@bureau_required
def liste_inactifs(request):
    config = ConfigExercice.get_exercice_courant()
    adherents = Adherent.objects.filter(statut='INACTIF').order_by('nom_prenom')
    return render(request, 'dashboard/adherents/inactifs.html', {'config_exercice': config, 'adherents': adherents})

# ══════════════════════════════════════════════════════════════
# AJOUTER à la fin de apps/adherents/views.py
# ══════════════════════════════════════════════════════════════

@bureau_required
def telecharger_listeadherent(request):
    """Génère LISTEADHERENT.xlsx avec toutes les feuilles du fichier original."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from apps.parametrage.models import ConfigExercice

    config    = ConfigExercice.get_exercice_courant()
    annee     = config.annee
    adherents = Adherent.objects.filter(statut='ACTIF').order_by('numero_ordre')
    tous      = Adherent.objects.all().order_by('numero_ordre')

    wb   = Workbook()
    wb.remove(wb.active)
    BLEU = '1B2B5E'; OR = 'C9A84C'; BLANC = 'FFFFFF'

    def _hdr(ws, cols, titre):
        ws.cell(1, 1, 'ASSOCIATION'); ws.cell(1, 2, 'ASELBY')
        ws.cell(2, 1, 'ANNEE');      ws.cell(2, 2, annee)
        ws.cell(3, 1, titre)
        ws.cell(3, 1).font = Font(bold=True, size=11, name='Arial')
        f  = PatternFill('solid', fgColor=BLEU)
        fn = Font(bold=True, color=BLANC, size=8, name='Arial')
        for j, c in enumerate(cols, 1):
            cl = ws.cell(5, j, c)
            cl.fill = f; cl.font = fn
            cl.alignment = Alignment(wrap_text=True, horizontal='center')
            ws.column_dimensions[get_column_letter(j)].width = 18
        ws.row_dimensions[5].height = 28

    # ── Feuille 1 : LISTADHERENT ──────────────────────────────
    ws1 = wb.create_sheet('LISTADHERENT')
    COLS1 = ['MLE', 'NUMORDRE', 'NOM ET  PRENOM', 'FONCTION',
             'DATE ADHESION', 'CONTACT', 'CONTACT 2', 'RESIDENCE', 'STATUT', 'DATE RECEPTION']
    _hdr(ws1, COLS1, 'LISTE DES ADHERENTS')
    for i, adh in enumerate(tous, 6):
        ws1.cell(i, 1, adh.matricule)
        ws1.cell(i, 2, adh.numero_ordre)
        ws1.cell(i, 3, adh.nom_prenom)
        ws1.cell(i, 4, adh.fonction)
        ws1.cell(i, 5, adh.date_adhesion)
        ws1.cell(i, 6, adh.telephone1)
        ws1.cell(i, 7, adh.telephone2)
        ws1.cell(i, 8, adh.residence)
        ws1.cell(i, 9, adh.statut)
        ws1.cell(i, 10, adh.date_reception)

    # ── Feuille 2 : DATERECEPTION ─────────────────────────────
    ws2 = wb.create_sheet('DATERECEPTION')
    COLS2 = ['MLE', 'NUMORDRE', 'NOM ET  PRENOM', 'DATE RECEPTION']
    _hdr(ws2, COLS2, 'DATES DE RECEPTION')
    actifs_reception = tous.exclude(date_reception__isnull=True).order_by('date_reception')
    for i, adh in enumerate(actifs_reception, 6):
        ws2.cell(i, 1, adh.matricule)
        ws2.cell(i, 2, adh.numero_ordre)
        ws2.cell(i, 3, adh.nom_prenom)
        ws2.cell(i, 4, adh.date_reception)

    # ── Feuille 3 : LISTADHERENTRECEPTION ─────────────────────
    ws3 = wb.create_sheet('LISTADHERENTRECEPTION')
    COLS3 = ['MLE', 'NUMORDRE', 'NOM ET  PRENOM', 'STATUT', 'DATE ADHESION',
             'DATE RECEPTION', 'POSTE BUREAU', 'NOTES']
    _hdr(ws3, COLS3, 'LISTE ADHERENTS + RECEPTION')
    for i, adh in enumerate(tous, 6):
        ws3.cell(i, 1, adh.matricule)
        ws3.cell(i, 2, adh.numero_ordre)
        ws3.cell(i, 3, adh.nom_prenom)
        ws3.cell(i, 4, adh.statut)
        ws3.cell(i, 5, adh.date_adhesion)
        ws3.cell(i, 6, adh.date_reception)
        ws3.cell(i, 7, adh.poste_bureau)
        ws3.cell(i, 8, adh.notes)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f'attachment; filename="ASELBY{annee}LISTEADHERENT.xlsx"'
    return resp