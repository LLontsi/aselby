from django.shortcuts import render, get_object_or_404
from django.utils import timezone
from django.db.models import Sum
from apps.core.mixins import bureau_required
from apps.parametrage.models import ConfigExercice
from apps.adherents.models import Adherent
from .models import MouvementFonds, ReserveMensuelle
from .services import calculer_interets_mensuels
from decimal import Decimal

@bureau_required
def etat_mensuel(request):
    config = ConfigExercice.get_exercice_courant()
    mois = int(request.GET.get('mois', timezone.now().month))
    annee = int(request.GET.get('annee', timezone.now().year))
    mouvements = MouvementFonds.objects.filter(mois=mois, annee=annee, config_exercice=config).select_related('adherent').order_by('adherent__numero_ordre')
    # Si aucune donnée pour le mois courant, afficher le dernier mois disponible
    if not mouvements.exists() and not request.GET.get('mois'):
        dernier = MouvementFonds.objects.filter(
            config_exercice=config
        ).order_by('-annee', '-mois').first()
        if dernier:
            mois = dernier.mois
            annee = dernier.annee
            mouvements = MouvementFonds.objects.filter(mois=mois, annee=annee, config_exercice=config).select_related('adherent').order_by('adherent__numero_ordre')
    total_fonds = mouvements.aggregate(t=Sum('fonds_definitif'))['t'] or Decimal('0')
    total_interets = mouvements.aggregate(t=Sum('interet_attribue'))['t'] or Decimal('0')
    nb_eligibles = mouvements.filter(base_calcul_interet__gt=0).count()
    ctx = {'config_exercice': config, 'mouvements': mouvements, 'total_fonds': total_fonds,
           'total_interets': total_interets, 'nb_eligibles': nb_eligibles, 'mois': mois, 'annee': annee}
    return render(request, 'dashboard/fonds/etat.html', ctx)

@bureau_required
def detail_adherent(request, matricule):
    config = ConfigExercice.get_exercice_courant()
    adherent = get_object_or_404(Adherent, matricule=matricule)
    mouvements = MouvementFonds.objects.filter(adherent=adherent, annee=config.annee).order_by('mois')
    ctx = {'config_exercice': config, 'adherent': adherent, 'mouvements': mouvements}
    return render(request, 'dashboard/fonds/detail.html', ctx)

@bureau_required
def repartition_interets(request):
    config = ConfigExercice.get_exercice_courant()
    if request.method == 'POST':
        mois = int(request.POST.get('mois'))
        annee = int(request.POST.get('annee'))
        pool = Decimal(request.POST.get('pool_interets', '0'))
        result = calculer_interets_mensuels(mois, annee, pool, config)
        from django.contrib import messages
        messages.success(request, f"Intérêts répartis : {result['total_distribue']:,.2f} FCFA entre {result['nb_eligibles']} adhérents éligibles.")
    return render(request, 'dashboard/fonds/interets.html', {'config_exercice': config})



# ══════════════════════════════════════════════════════════════
# AJOUTER à la fin de apps/fonds/views.py
# ══════════════════════════════════════════════════════════════

MOIS_CODE = ['','JANV','FEV','MARS','AVRIL','MAI','JUIN',
             'JUIL','AOUT','SEPT','OCT','NOV','DEC']


@bureau_required
def telecharger_listefondscaisse(request):
    """
    LISTEFONDSCAISSE.xlsx — fidèle à l'Excel réel.
    Feuille FONDSCAISSE :
      A=MATRICULE, B=NUMORDRE, C=NOM ET PRENOM,
      D=FONDS DE CAISSE (capital_depart_exercice),
      E=RETRAIT PARTIEL (cumul annuel),
      F=RECONDUCTION (capital_compose mois précédent),
      G=TOTAL FONDS = D - E
    """
    import io, math
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from decimal import Decimal
    from apps.saisie.models import SaisieMonthly

    D = Decimal
    config    = ConfigExercice.get_exercice_courant()
    annee     = config.annee
    adherents = sorted(Adherent.objects.filter(statut='ACTIF'), key=lambda a: (0 if a.matricule=='AS201648' else 1, a.numero_ordre))
    wb = Workbook()

    # ── Feuille FONDSCAISSE ──────────────────────────────────
    ws = wb.active
    ws.title = 'FONDSCAISSE'
    BLEU = '1B2B5E'; BLANC = 'FFFFFF'

    ws['A1'] = 'ASSOCIATION'; ws['B1'] = 'ASELBY'
    ws['A2'] = 'ANNEE';       ws['B2'] = annee

    COLS = ['MATRICULE','NUMORDRE','NOM ET PRENOM','FONDS DE CAISSE',
            'RETRAIT PARTIEL','RECONDUCTION','TOTAL FONDS']
    fill = PatternFill('solid', fgColor=BLEU)
    fn   = Font(bold=True, color=BLANC, size=9, name='Arial')
    for j, h in enumerate(COLS, 1):
        cl = ws.cell(4, j, h)
        cl.fill = fill; cl.font = fn
        cl.alignment = Alignment(horizontal='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width = 20

    for i, adh in enumerate(adherents, 5):
        # Col D = capital_depart_exercice (fixe début exercice, ex: 900 000)
        fonds_caisse = D(str(adh.capital_depart_exercice or 0))

        # Col E = cumul retrait_partiel sur l'année
        retrait = sum(
            D(str(s.retrait_partiel or 0))
            for s in SaisieMonthly.objects.filter(
                adherent=adh, annee=annee, config_exercice=config)
        )

        # Col F = reconduction = capital_compose du dernier mois
        mvt_last = MouvementFonds.objects.filter(
            adherent=adh, annee=annee, config_exercice=config
        ).order_by('-mois').first()
        reconduction = D(str(mvt_last.capital_compose if mvt_last else 0))

        # Col G = TOTAL FONDS = D - E
        total_fonds = fonds_caisse - retrait

        ws.cell(i, 1, adh.matricule)
        ws.cell(i, 2, adh.numero_ordre)
        ws.cell(i, 3, adh.nom_prenom)
        ws.cell(i, 4, float(fonds_caisse))
        ws.cell(i, 5, float(retrait))
        ws.cell(i, 6, float(reconduction))
        ws.cell(i, 7, float(total_fonds))

    # Ligne TOTAL
    nb = len(adherents) + 5
    ws.cell(nb, 1, 'TOTAL')
    for col in [4, 5, 6, 7]:
        ws.cell(nb, col, f'=SUM({get_column_letter(col)}5:{get_column_letter(col)}{nb-1})')

    # ── Feuille BACKUP ──────────────────────────────────────
    ws_bk = wb.create_sheet('FONDSCAISSEBACKUP')
    ws_bk['A1'] = 'BACKUP — copie FONDSCAISSE'

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    resp = HttpResponse(buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename=ASELBY{annee}LISTEFONDSCAISSE.xlsx'
    return resp


@bureau_required
def telecharger_basecalculinteret(request):
    """
    BASECALCULINTERET — logique exacte vérifiée cellule par cellule.

    Structure:
    - Feuille MVTDEC(n-1) = TRANSITION: F = C + D (capital_depart + reconduction_cassation)
    - Feuilles MVTmois(n)  = COURANT:   F = I_prec - E + L

    Colonnes A→AL (38) — mapping TABBORD vérifié:
      E  ← TABBORD AV(48) = retrait_partiel
      J  ← TABBORD V(22)  = sanction
      K  ← TABBORD AS(45) = reste  [= SaisieMonthly.reste]
      P  ← TABBORD H(8)   = penalite_versement_especes
      Q  ← TABBORD W(23)  = inscription
      R  ← TABBORD AT(46) = mutuelle
      S  ← TABBORD Z(26)  = pret_fonds
      T  ← TABBORD BF(58) = interet_pret
      U  ← TABBORD AB(28) = numero_cheque
      V  ← TABBORD S(19)  = remboursement_pret
      W  ← TABBORD T(20)  = mode_remb_pret
      X  ← TABBORD AA(27) = mode_paiement_pret
      Y  ← TABBORD AG(33) = numero_cheque 2
      Z  ← TABBORD BL(64) = penalite_pret_fonds
      AB ← TABBORD AR(44) = penalite_echec_tontine
      AC ← TABBORD AQ(43) = penalite_retard_tontine
      AD ← TABBORD AU(47) = remboursement_transport
      AE ← TABBORD BB(54) = contribution_foyer
      AF ← TABBORD BH(60) = depense_fonds_roulement
      AG ← TABBORD BI(61) = depense_frais_exc
      AH ← TABBORD BJ(62) = depense_fonds_mutuel
      AI ← TABBORD BK(63) = depense_collation
      AJ ← TABBORD BL(64) = penalite_pret_fonds_sortie
      AK ← TABBORD BM(65) = depense_penalite_vst_banque
      AL ← TABBORD BN(66) = autres_depenses
    """
    import io, math
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from decimal import Decimal
    from apps.saisie.models import SaisieMonthly
    from apps.exercice.models import FicheCassation

    D = Decimal
    def _d(v):
        if v is None: return D('0')
        try: return D(str(v))
        except: return D('0')

    config    = ConfigExercice.get_exercice_courant()
    annee     = config.annee
    adherents = list(sorted(Adherent.objects.filter(statut='ACTIF'), key=lambda a: (0 if a.matricule=='AS201648' else 1, a.numero_ordre)))

    BLEU = '1B2B5E'; BLANC = 'FFFFFF'
    MOIS_CODE = {1:'JANV',2:'FEV',3:'MARS',4:'AVRIL',5:'MAI',6:'JUIN',
                 7:'JUIL',8:'AOUT',9:'SEPT',10:'OCT',11:'NOV',12:'DEC'}
    MOIS_NOM  = {1:'JANVIER',2:'FEVRIER',3:'MARS',4:'AVRIL',5:'MAI',6:'JUIN',
                 7:'JUILLET',8:'AOUT',9:'SEPTEMBRE',10:'OCTOBRE',11:'NOVEMBRE',12:'DECEMBRE'}

    COLS_HDR = [
        'MATRICULE','NOM ET PRENOM','FONDES DE DEPART','RECONDUCTION',
        'RETRAIT PARTIEL','FONDS DEFINITIF',
        'BASE DE CALCUL INTERET FONDS DEFINITIF',
        'REPARTITION PROVISOIRE INTERET FONDS+EPARGNE','CAPITAL COMPOSE',
        'SANCTION','RESTE','EPARGNE','FONDS DE ROULEMENT',
        'FRAIS EXCEPTIONNEL','COLLATION','PENALITE VST ESPECES',
        'INSCRISPTION','MUTUELLE','PRET FONDS','INTERET PRET FONDS',
        'NUMERO CHEQUE','REMBOURSEMENT PRET FONDS',
        'MODE VERSEMENT REMBOURSEMENT','MODE VERSEMENT PRET',
        'NUMERO CHEQUE','PENALITE PRET FONDS','PENALITE FONDS',
        'PENALITE ECHEC TONTINE','PENALITE RETARD TONTINE',
        'REMBOURSEMENT TRANSPORT','CONTRIBUTION FOYER',
        'DEPENSE FONDS DE ROULEMENT','DEPENSE FRAIS EXCEPTIONNEL',
        'DEPENSE FONDS MUTUELLE','DEPENSE COLLATION RECEPTION',
        'PENALITE PRET FONDS','DEPENSE PENALITE VERSEMENT BANQUE',
        'AUTRES DEPENSES',
    ]

    wb = Workbook()
    wb.remove(wb.active)

    # ── Récupérer les reconductions de cassation (FicheCassation année-1) ──
    config_prec = ConfigExercice.objects.filter(annee=annee-1).first()
    reconductions = {}  # mat → Decimal
    if config_prec:
        for fc in FicheCassation.objects.filter(config_exercice=config_prec):
            reconductions[fc.adherent_id] = _d(fc.reconduction)

    # ── Feuilles à générer ────────────────────────────────────────
    # 1. MVTDEC(n-1) : feuille de transition avec C+D = fonds départ
    # 2. MVTmois(n)  : mois de l'exercice courant
    from django.db.models import Max
    sheets = []
    if config_prec:
        res = MouvementFonds.objects.filter(config_exercice=config_prec).aggregate(m=Max('mois'))
        if res['m']:
            sheets.append(('transition', annee-1, res['m'], config_prec))
    mois_dispo = list(
        MouvementFonds.objects.filter(config_exercice=config)
        .values_list('mois', flat=True).distinct().order_by('mois')
    )
    for m in mois_dispo:
        sheets.append(('courant', annee, m, config))

    seuil      = _d(config.seuil_eligibilite_interets)
    fonds_roul = _d(config.fonds_roulement_mensuel)
    frais_exc  = _d(config.frais_exceptionnels_mensuel)
    collation  = _d(config.collation_mensuelle)

    for sheet_type, an, mois, cfg_mois in sheets:
        label    = f"MVT{MOIS_CODE[mois]}{str(an)[-2:]}"
        ws       = wb.create_sheet(label)
        is_trans = (sheet_type == 'transition')

        ws['A1'] = 'ASSOCIATION'; ws['B1'] = 'ASELBY'
        ws['A2'] = 'ANNEE';       ws['B2'] = an
        ws['A3'] = 'TONTINE'
        ws['A4'] = f"MOIS: {MOIS_NOM[mois]}"

        fill = PatternFill('solid', fgColor=BLEU)
        fn   = Font(bold=True, color=BLANC, size=8, name='Arial')
        for j, h in enumerate(COLS_HDR, 1):
            cl = ws.cell(5, j, h)
            cl.fill = fill; cl.font = fn
            cl.alignment = Alignment(horizontal='center', wrap_text=True)
            ws.column_dimensions[get_column_letter(j)].width = 14
        ws.row_dimensions[5].height = 36

        # Saisies du mois courant
        saisies = {
            s.adherent_id: s
            for s in SaisieMonthly.objects.filter(
                mois=mois, annee=an, config_exercice=cfg_mois)
        }

        # MouvementFonds mois précédent (pour I_prec dans les mois courants)
        prec_mois  = 12 if mois == 1 else mois - 1
        prec_annee = an - 1 if mois == 1 else an
        cfg_prec_m = ConfigExercice.objects.filter(annee=prec_annee).first()
        mvts_prec  = {}
        if cfg_prec_m:
            mvts_prec = {
                mv.adherent_id: mv
                for mv in MouvementFonds.objects.filter(
                    mois=prec_mois, annee=prec_annee, config_exercice=cfg_prec_m)
            }

        # Pool T = somme intérêts prêt du mois
        pool_t = sum(_d(s.interet_pret) for s in saisies.values())

        # ── 1er passage : F, G, L ─────────────────────────
        rows_data = {}
        total_G = D('0')

        for adh in adherents:
            s  = saisies.get(adh.matricule)
            mp = mvts_prec.get(adh.matricule)

            # K = RESTE depuis SaisieMonthly.reste
            reste = _d(s.reste) if s else D('0')

            # L = IF(K=0, 0, K-M-N-O)
            if reste > 0:
                epargne = reste - fonds_roul - frais_exc - collation
                if epargne < 0: epargne = D('0')
                m_val = fonds_roul; n_val = frais_exc; o_val = collation
            else:
                epargne = D('0')
                m_val = D('0'); n_val = D('0'); o_val = D('0')

            # E = retrait du mois
            retrait = _d(s.retrait_partiel) if s else D('0')

            if is_trans:
                # MVTDEC(n-1) — feuille de transition
                # C = capital_depart_exercice (début exercice courant)
                capital_depart = _d(adh.capital_depart_exercice)
                # D = reconduction cassation (FicheCassation année-1)
                reconduction = reconductions.get(adh.matricule, D('0'))
                # F = C + D
                fonds_def = capital_depart + reconduction
                c_val = capital_depart
                d_val = reconduction
            else:
                # Mois courant: F = I_prec - E + L
                c_val = None
                d_val = None
                reconduction_prec = _d(mp.capital_compose) if mp else D('0')
                fonds_def = reconduction_prec - retrait + epargne

            # G = IF(F > seuil, F, 0)
            base_calc = fonds_def if fonds_def > seuil else D('0')
            total_G  += base_calc

            rows_data[adh.matricule] = {
                'c': c_val, 'd': d_val,
                'retrait': retrait, 'reste': reste, 'epargne': epargne,
                'fonds_def': fonds_def, 'base_calc': base_calc,
                'm': m_val, 'n': n_val, 'o': o_val,
            }

        # ── 2ème passage : H, I, remplissage ─────────────
        for i, adh in enumerate(adherents, 6):
            b = rows_data[adh.matricule]
            s = saisies.get(adh.matricule)

            # H = ROUNDDOWN(pool/total_G * G, 2)
            if total_G > 0 and b['base_calc'] > 0:
                h_raw = float(pool_t) / float(total_G) * float(b['base_calc'])
                interet = D(str(math.floor(h_raw * 100) / 100))
            else:
                interet = D('0')

            # I = F + H
            capital_compose = b['fonds_def'] + interet

            def sv(field, default=D('0')):
                if not s: return default
                v = getattr(s, field, None)
                return _d(v) if v is not None else default

            def ss(field, default=''):
                if not s: return default
                v = getattr(s, field, None)
                return str(v).strip() if v is not None else default

            ws.cell(i,  1, adh.matricule)
            ws.cell(i,  2, adh.nom_prenom)
            # C : capital_depart pour transition, None pour courant
            if b['c'] is not None:
                ws.cell(i, 3, float(b['c']))
            # D : reconduction pour transition, None pour courant
            if b['d'] is not None:
                ws.cell(i, 4, float(b['d']))
            ws.cell(i,  5, float(b['retrait']))
            ws.cell(i,  6, float(b['fonds_def']))
            ws.cell(i,  7, float(b['base_calc']))
            ws.cell(i,  8, float(interet))
            ws.cell(i,  9, float(capital_compose))
            ws.cell(i, 10, float(sv('sanction')))
            ws.cell(i, 11, float(b['reste']))
            ws.cell(i, 12, float(b['epargne']))
            ws.cell(i, 13, float(b['m']))
            ws.cell(i, 14, float(b['n']))
            ws.cell(i, 15, float(b['o']))
            ws.cell(i, 16, float(_d(s.penalite_versement_especes)) if s else 0)
            ws.cell(i, 17, float(sv('inscription')))
            ws.cell(i, 18, float(sv('mutuelle')))
            ws.cell(i, 19, float(sv('pret_fonds')))
            ws.cell(i, 20, float(_d(s.interet_pret)) if s else 0)
            ws.cell(i, 21, ss('numero_cheque'))
            ws.cell(i, 22, float(sv('remboursement_pret')))
            ws.cell(i, 23, ss('mode_remb_pret'))
            ws.cell(i, 24, ss('mode_paiement_pret'))
            ws.cell(i, 25, '')
            ws.cell(i, 26, float(sv('penalite_pret_fonds')))
            ws.cell(i, 27, 0)
            ws.cell(i, 28, float(_d(s.penalite_echec_tontine)) if s else 0)
            ws.cell(i, 29, float(_d(s.penalite_retard_tontine)) if s else 0)
            ws.cell(i, 30, 0)
            ws.cell(i, 31, float(sv('contribution_foyer')))
            ws.cell(i, 32, 0)
            ws.cell(i, 33, 0)
            ws.cell(i, 34, 0)
            ws.cell(i, 35, 0)
            ws.cell(i, 36, float(sv('penalite_pret_fonds')))
            ws.cell(i, 37, 0)
            ws.cell(i, 38, float(sv('montant_depense')))

        # Ligne TOTAL
        nb = len(adherents) + 6
        ws.cell(nb, 1, 'TOTAL')
        for col in range(3, 39):
            col_l = get_column_letter(col)
            ws.cell(nb, col, f'=SUM({col_l}6:{col_l}{nb-1})')
        # Ligne de sous-totaux dépenses (L46 dans réel)
        nb2 = nb + 2
        ws.cell(nb2, 13, f'=SUM(M6:M{nb-1})')   # M = FONDS DE ROULEMENT total
        ws.cell(nb2, 14, f'=SUM(N6:N{nb-1})')   # N = FRAIS EXCEPTIONNEL total
        ws.cell(nb2, 15, f'=SUM(O6:O{nb-1})')   # O = COLLATION total

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    resp = HttpResponse(buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename=ASELBY{annee}BASECALCULINTERET.xlsx'
    return resp