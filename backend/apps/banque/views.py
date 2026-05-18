"""
apps/banque/views.py
Formulaire 1 : Relevé bancaire mensuel (TABBHISTOBQUE).
À saisir EN PREMIER avant la saisie mensuelle.
"""
from django.shortcuts import render, redirect
from django.contrib import messages
from django.utils import timezone
from django.db import transaction
from decimal import Decimal
from apps.core.mixins import bureau_required
from apps.parametrage.models import ConfigExercice
from apps.adherents.models import Adherent
from .models import ReleveBancaire, AgioBancaire

MOIS_FR = ['','Janvier','Février','Mars','Avril','Mai','Juin',
           'Juillet','Août','Septembre','Octobre','Novembre','Décembre']
D = Decimal


@bureau_required
def releve_bancaire(request):
    """
    Grille : tous les adhérents × 4 champs (banque, espèces, autre, agio).
    Sauvegarde tous les adhérents en une seule soumission.
    """
    config = ConfigExercice.get_exercice_courant()
    mois   = int(request.GET.get('mois',  timezone.now().month))
    annee  = int(request.GET.get('annee', config.annee))

    adherents = sorted(Adherent.objects.filter(statut='ACTIF'), key=lambda a: (0 if a.matricule=='AS201648' else 1, a.numero_ordre))

    if request.method == 'POST':
        mois  = int(request.POST.get('mois',  mois))
        annee = int(request.POST.get('annee', annee))

        def d(key):
            v = request.POST.get(key, '0') or '0'
            try:
                return D(str(v).replace(' ', '').replace('\xa0', ''))
            except Exception:
                return D('0')

        nb_saisis = 0
        with transaction.atomic():
            for adh in adherents:
                pfx = f"{adh.matricule}_"
                banque  = d(f"{pfx}banque")
                especes = d(f"{pfx}especes")
                autre   = d(f"{pfx}autre")
                agio    = d(f"{pfx}agio")

                # Ne sauvegarder que si au moins une valeur non nulle
                if banque or especes or autre or agio:
                    ReleveBancaire.objects.update_or_create(
                        adherent=adh, mois=mois, annee=annee,
                        defaults={
                            'config_exercice':  config,
                            'versement_banque':  banque,
                            'versement_especes': especes,
                            'autre_versement':   autre,
                            'agio':              agio,
                        }
                    )
                    nb_saisis += 1

        messages.success(
            request,
            f"Relevé bancaire {MOIS_FR[mois]} {annee} enregistré "
            f"({nb_saisis} adhérent(s)). "
            f"Passez maintenant à la Saisie mensuelle."
        )
        return redirect(f"{request.path}?mois={mois}&annee={annee}")

    # GET — charger les données existantes
    releves = {
        r.adherent_id: r
        for r in ReleveBancaire.objects.filter(
            mois=mois, annee=annee, config_exercice=config)
    }

    # Navigation mois
    prec = (12, annee-1) if mois == 1 else (mois-1, annee)
    suiv = (1,  annee+1) if mois == 12 else (mois+1, annee)

    nb_saisis    = len(releves)
    total_banque = sum(r.versement_banque  for r in releves.values())
    total_esp    = sum(r.versement_especes for r in releves.values())

    return render(request, 'dashboard/banque/releve_bancaire.html', {
        'config_exercice': config,
        'adherents':       adherents,
        'releves':         releves,
        'mois':            mois,
        'annee':           annee,
        'mois_label':      MOIS_FR[mois],
        'mois_fr':         MOIS_FR,
        'prec_mois':       prec[0], 'prec_annee': prec[1],
        'suiv_mois':       suiv[0], 'suiv_annee': suiv[1],
        'nb_total':        len(adherents),
        'nb_saisis':       nb_saisis,
        'total_banque':    total_banque,
        'total_especes':   total_esp,
    })


@bureau_required
def tresorerie(request):
    """Vue synthèse trésorerie — inchangée."""
    from apps.banque.models import ReleveBancaire
    from django.db.models import Sum
    config = ConfigExercice.get_exercice_courant()
    mois   = int(request.GET.get('mois', timezone.now().month))
    annee  = int(request.GET.get('annee', config.annee))

    releves = ReleveBancaire.objects.filter(
        mois=mois, annee=annee, config_exercice=config
    ).select_related('adherent').order_by('adherent__numero_ordre')

    totaux = releves.aggregate(
        total_banque  = Sum('versement_banque'),
        total_especes = Sum('versement_especes'),
        total_autre   = Sum('autre_versement'),
        total_agio    = Sum('agio'),
    )

    return render(request, 'dashboard/banque/tresorerie.html', {
        'config_exercice': config,
        'releves':  releves,
        'totaux':   totaux,
        'mois':     mois,
        'annee':    annee,
        'mois_label': MOIS_FR[mois],
        'mois_fr':  MOIS_FR,
    })


@bureau_required
def telecharger_tabbhistobque(request):
    """
    TABBHISTOBQUE — 10 colonnes fidèles au réel:
      A = MATRICULE
      B = NOM ET PRENOM
      C = HISTORIQUE TONTINE = tontine_t60 + tontine_t75 + tontine_t100
      D = HISTORIQUE ESPECES = versement_especes
      E = HISTORIQUE BANQUE  = versement_banque
      F = AUTRE VERSEMENT    = autre_versement
      G = EN COMPTE          = autre_versement (même valeur que F)
      H = MONTANT ENGAGEMENT = D + E + F
      I = MONTANT A JUSTIFIER= H - C  (positif=excédent, négatif=manque)
      J = AGIO               = ReleveBancaire.agio (intérêts bancaires)

    Structure:
      L1: ASSOCIATION / ASELBY
      L2: MOIS / mois-annee
      L3: numérotation 1..10
      L4: HISTORIQUE BANCAIRE (titre centré)
      L5: entêtes
      L6+: données (seulement membres avec au moins un versement)
      Ln: TOTAUX
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from apps.saisie.models import SaisieMonthly

    D = __import__('decimal').Decimal

    def _d(v):
        if v is None: return D('0')
        try: return D(str(v))
        except: return D('0')

    config    = ConfigExercice.get_exercice_courant()
    annee     = config.annee
    adherents = sorted(
        Adherent.objects.filter(statut='ACTIF'),
        key=lambda a: (0 if a.matricule == 'AS201648' else 1, a.numero_ordre)
    )

    BLEU = '1B2B5E'; BLANC = 'FFFFFF'
    COLS = ['MATRICULE','NOM ET PRENOM','HISTORIQUE TONTINE',
            'HITORIQUE ESPECES','HITORIQUE BANQUE','AUTRE VERSEMENT',
            'EN COMPTE','MONTANT ENGAGEMENT','MONTANT A JUSTIFIER','AGIO']

    MOIS_CODE = {1:'JANV',2:'FEV',3:'MARS',4:'AVRIL',5:'MAI',6:'JUIN',
                 7:'JUIL',8:'AOUT',9:'SEPT',10:'OCT',11:'NOV',12:'DEC'}
    MOIS_FR_COURT = {
        1:'janv',2:'févr',3:'mars',4:'avr',5:'mai',6:'juin',
        7:'juil',8:'août',9:'sept',10:'oct',11:'nov',12:'déc'
    }

    wb = Workbook()
    wb.remove(wb.active)

    # Pré-charger toutes les données
    releves_all = {}
    for r in ReleveBancaire.objects.filter(annee=annee, config_exercice=config):
        releves_all[(r.adherent_id, r.mois)] = r
    saisies_all = {}
    for s in SaisieMonthly.objects.filter(annee=annee, config_exercice=config):
        saisies_all[(s.adherent_id, s.mois)] = s

    # Cumul annuel pour RESUME
    cumul = {adh.matricule: [D('0')]*8 for adh in adherents}

    for mois in range(1, 13):
        label = f"HISTOBQUE{MOIS_CODE[mois]}{str(annee)[-2:]}"
        ws    = wb.create_sheet(label)
        mois_label = f"{MOIS_FR_COURT[mois]}-{str(annee)[-2:]}"

        # En-têtes
        ws['A1'] = 'ASSOCIATION'; ws['B1'] = 'ASELBY'
        ws['A2'] = 'MOIS';        ws['B2'] = mois_label

        # L3: numérotation
        for j in range(1, len(COLS)+1):
            ws.cell(3, j, j)

        # L4: titre HISTORIQUE BANCAIRE
        ws.merge_cells('A4:J4')
        titre = ws['A4']
        titre.value = 'HISTORIQUE BANCAIRE'
        titre.font = Font(bold=True, size=14, name='Arial')
        titre.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[4].height = 24

        # L5: entêtes colonnes
        fill = PatternFill('solid', fgColor=BLEU)
        fn   = Font(bold=True, color=BLANC, size=8, name='Arial')
        for j, h in enumerate(COLS, 1):
            cl = ws.cell(5, j, h)
            cl.fill = fill; cl.font = fn
            cl.alignment = Alignment(horizontal='center', wrap_text=True)
            ws.column_dimensions[get_column_letter(j)].width = 18
        ws.row_dimensions[5].height = 32

        # Largeur colonnes
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 26

        # Données
        row_n = 6
        tot = [D('0')] * 8  # totaux colonnes C-J

        for adh in adherents:
            r = releves_all.get((adh.matricule, mois))
            s = saisies_all.get((adh.matricule, mois))

            D_esp = _d(r.versement_especes)  if r else D('0')
            E_bq  = _d(r.versement_banque)   if r else D('0')
            F_aut = _d(r.autre_versement)    if r else D('0')
            agio  = _d(r.agio)              if r else D('0')

            if s:
                C = _d(s.tontine_t60) + _d(s.tontine_t75) + _d(s.tontine_t100)
            else:
                C = D('0')

            H = D_esp + E_bq + F_aut      # MONTANT ENGAGEMENT
            I = H - C                      # MONTANT A JUSTIFIER

            vals = [C, D_esp, E_bq, F_aut, F_aut, H, I, agio]

            # Cumuler
            for k, v in enumerate(vals):
                cumul[adh.matricule][k] += v
                tot[k] += v

            ws.cell(row_n, 1, adh.matricule)
            ws.cell(row_n, 2, adh.nom_prenom)
            for k, v in enumerate(vals, 3):
                cell = ws.cell(row_n, k, float(v) if v != 0 else 0)
                if v < 0:
                    cell.font = Font(color='FF0000', size=9, name='Arial')
                elif k in [3, 4]:  # C=tontine, D=espèces en rouge dans l'image
                    cell.font = Font(color='FF0000', size=9, name='Arial')
            row_n += 1

        # Ligne TOTAUX
        ws.cell(row_n, 2, 'TOTAUX').font = Font(bold=True, size=9, name='Arial')
        for k, v in enumerate(tot, 3):
            ws.cell(row_n, k, float(v)).font = Font(bold=True, size=9, name='Arial')

    # Feuille RESUME (cumul annuel)
    label_r = f"HISTOBQUERESUME{str(annee)[-2:]}"
    ws_r = wb.create_sheet(label_r)
    ws_r['A1'] = 'ASSOCIATION'; ws_r['B1'] = 'ASELBY'
    ws_r['A2'] = 'MOIS';        ws_r['B2'] = 'RESUME'
    for j in range(1, len(COLS)+1):
        ws_r.cell(3, j, j)
    ws_r.merge_cells('A4:J4')
    titre = ws_r['A4']
    titre.value = 'HISTORIQUE BANCAIRE — RESUME ANNUEL'
    titre.font = Font(bold=True, size=14, name='Arial')
    titre.alignment = Alignment(horizontal='center', vertical='center')
    ws_r.row_dimensions[4].height = 24
    fill = PatternFill('solid', fgColor=BLEU)
    fn   = Font(bold=True, color=BLANC, size=8, name='Arial')
    for j, h in enumerate(COLS, 1):
        cl = ws_r.cell(5, j, h)
        cl.fill = fill; cl.font = fn
        cl.alignment = Alignment(horizontal='center', wrap_text=True)
        ws_r.column_dimensions[get_column_letter(j)].width = 18
    ws_r.column_dimensions['A'].width = 12
    ws_r.column_dimensions['B'].width = 26
    ws_r.row_dimensions[5].height = 32

    row_n = 6
    tot_r = [D('0')] * 8
    for adh in adherents:
        c_vals = cumul[adh.matricule]
        ws_r.cell(row_n, 1, adh.matricule)
        ws_r.cell(row_n, 2, adh.nom_prenom)
        for k, v in enumerate(c_vals, 3):
            ws_r.cell(row_n, k, float(v) if v != 0 else 0)
            tot_r[k-3] += v
        row_n += 1

    ws_r.cell(row_n, 2, 'TOTAUX').font = Font(bold=True, size=9)
    for k, v in enumerate(tot_r, 3):
        ws_r.cell(row_n, k, float(v)).font = Font(bold=True, size=9)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    resp = HttpResponse(buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = (
        f'attachment; filename=ASELBY{annee}TABBHISTOBQUE.xlsx')
    return resp